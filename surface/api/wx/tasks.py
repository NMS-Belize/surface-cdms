from __future__ import absolute_import, unicode_literals

import hashlib
import json
import logging
import os
import socket
import uuid
from datetime import datetime, timedelta, timezone, date
import calendar
from ftplib import FTP, error_perm, error_reply
from time import sleep, time
from croniter import croniter

import csv
import tempfile
import urllib3
from minio import Minio

import math, cmath
from matplotlib.transforms import Bbox
from metpy.interpolate import interpolate_to_grid
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import cronex
import dateutil.parser
import pandas
from geopandas import geopandas
import cv2
import matplotlib.pyplot as plt
from pathlib import Path
import psycopg2
from psycopg2 import sql
import pytz
import requests
import re
import subprocess
import shutil
import tarfile
import wx.export_surface_oscar as exso
from celery import shared_task, current_app
from celery.utils.log import get_task_logger
from django.core.cache import cache
from django.db import connection
from django.utils.timezone import now
from django.db.models import Count, Q
from django.db import transaction
from django_celery_beat.models import PeriodicTask


from tempestas_api import settings
from wx.decoders.flash import read_data as read_data_flash
from wx.decoders.hobo import read_file as read_file_hobo
from wx.decoders.hydro import read_file as read_file_hydrology
from wx.decoders.manual_data import read_file as read_file_manual_data
from wx.decoders.manual_data_hourly import read_file as read_file_manual_data_hourly
from wx.decoders.f2000 import read_file as read_file_f2000
from wx.decoders.R_format import read_file as read_file_rformat
from wx.decoders.nesa import read_data as read_data_nesa
from wx.decoders.sat_tx325 import read_data as read_data_sat_tx325
from wx.decoders.surface import read_file as read_file_surface
from wx.decoders.toa5 import read_file as read_file_toa5
from wx.decoders.weatherlink_json import read_file as read_file_weatherlink_json
from wx.decoders.f2000 import read_file as read_file_f2000
from wx.decoders.davis import read_file as read_file_davis
from wx.decoders.R_format import read_file as read_file_R_format
from wx.models import DataFile, CombineDataFile
from wx.models import Document
from wx.models import NoaaDcp
from wx.models import Station, Country, WMOReportingStatus, WMORegion, WMOStationType
from wx.models import StationFileIngestion, StationDataFile, ManualStationDataFile
from wx.models import HourlySummaryTask, DailySummaryTask
from wx.models import HydroMLPredictionStation, HydroMLPredictionMapping
from wx.models import HighFrequencyData, HFSummaryTask
from wx.models import LocalWisCredentials, RegionalWisCredentials, Wis2BoxPublishLogs, Wis2BoxPublish
from wx.decoders.insert_raw_data import insert as insert_rd
from wx.decoders.insert_raw_data_daily import insert as insert_rd_daily
from wx.decoders.insert_raw_data_synop import insert as insert_rd_synop

from django.core.exceptions import ObjectDoesNotExist
import numpy as np
import pandas as pd
from wx.models import Variable
from wx.models import QcPersistThreshold, BackupTask, BackupLog, StagedRawData, StagedRawDataTask
from wx.enums import QualityFlagEnum

NOT_CHECKED = QualityFlagEnum.NOT_CHECKED.id
SUSPICIOUS = QualityFlagEnum.SUSPICIOUS.id
BAD = QualityFlagEnum.BAD.id
GOOD = QualityFlagEnum.GOOD.id

logger = get_task_logger(__name__)
db_logger = get_task_logger('db')

def get_connection():
    return psycopg2.connect(settings.SURFACE_CONNECTION_STRING)

############################################################


from wx.wave_data_generator import gen_dataframe_and_filename, format_and_send_data
from wx.models import FTPServer

@shared_task
def gen_toa5_file():
    logger.info('Inside dcp_tasks_scheduler')
    df, filename = gen_dataframe_and_filename()
    format_and_send_data(df, filename)


############################################################


from sqlalchemy import create_engine

ENGINE = create_engine(
    f"postgresql+psycopg2://{settings.SURFACE_DB_USER}:{settings.SURFACE_DB_PASSWORD}@{settings.SURFACE_DB_HOST}:5432/{settings.SURFACE_DB_NAME}",
    pool_pre_ping=True
)


############################################################

def backup_set_running(backup_task, started_at, file_path):
    status = 'RUNNING'
    message = 'Backup process is currently running.'
    new_log = BackupLog(backup_task=backup_task,
                        started_at=started_at,
                        status=status,
                        message=message,
                        file_path=file_path)

    new_log.save()
    return new_log.id


def backup_create(file_path):
    """
        Create a full PostgreSQL cluster backup using pg_basebackup.

        This creates a temporary base backup directory, compresses it into
        a .tar.gz file, and then removes the temporary directory.

        Notes:
        - This backs up the entire PostgreSQL cluster, not just one database.
        - Retention/deletion are handled separately by backup_free().
    """

    db = settings.DATABASES['default']

    db_host = db['HOST']
    db_port = str(db['PORT'])
    db_user = db['USER']
    db_pass = db['PASSWORD']

    # Copy the current environment and add PGPASSWORD.
    # This avoids placing the password directly inside a shell command.
    env = os.environ.copy()
    env["PGPASSWORD"] = db_pass

    temp_backup_dir = None

    try:
        # Create a unique temporary directory inside /data.
        # This prevents two backup jobs from conflicting with each other.
        temp_backup_dir = tempfile.mkdtemp(
            prefix="pgBackupTmpDir_",
            dir="/data"
        )

        # Run pg_basebackup without shell=True.
        # This is safer because arguments are passed directly to the program,
        # not interpreted by the shell.
        subprocess.run(
            [
                "/usr/bin/pg_basebackup",
                "--host", db_host,
                "--port", db_port,
                "--username", db_user,
                "--pgdata", temp_backup_dir,   # write backup files here
                "--format", "p",              # plain directory format
                "--wal-method", "stream",     # stream WAL files for consistency
                "--checkpoint", "fast",       # request a fast checkpoint
            ],
            env=env,
            check=True,
        )

        # Compress the temporary backup directory into the final .tar.gz file.
        # arcname="." means the contents of the temp directory are stored
        # at the root of the archive instead of including the temp folder name.
        with tarfile.open(file_path, "w:gz") as tar:
            tar.add(temp_backup_dir, arcname=".")

    except subprocess.CalledProcessError as e:
        logger.error(f"PostgreSQL backup failed during pg_basebackup: {e}")
        raise

    except Exception as e:
        logger.error(f"Backup failed: {e}")
        raise

    finally:
        # Always remove the temporary backup directory, whether the backup
        # succeeded or failed. This helps prevent /data from filling up.
        if temp_backup_dir and os.path.exists(temp_backup_dir):
            shutil.rmtree(temp_backup_dir, ignore_errors=True)


def backup_ftp(file_name, file_path, ftp_server, remote_folder):
    # remote_file_path = os.path.join(remote_folder, file_name)

    with FTP() as ftp:
        ftp.connect(host=ftp_server.host, port=ftp_server.port)
        ftp.login(user=ftp_server.username, passwd=ftp_server.password)
        # ftp.set_pasv(val = not ftp_server.is_active_mode)

        # Make sure we are in the target folder
        ftp.cwd(remote_folder)

        with open(file_path, "rb") as file:
            ftp.storbinary(f"STOR {file_name}", file)

        # with open(file_path, "rb") as file:
        #     ftp.storbinary(f"STOR {remote_file_path}", file)

        ftp.dir()
        ftp.quit()


def backup_log(log_id, backup_task, started_at, finished_at, status, message, file_path):
    message = str(message)

    if status in ['SUCCESS', 'FTP ERROR'] and file_path and os.path.exists(file_path):
        file_size = os.stat(file_path).st_size / 1024 / 1024  # Bytes -> MB

        BackupLog.objects.filter(id=log_id).update(
            finished_at=finished_at,
            status=status,
            message=message,
            file_size=file_size,
        )

    else:
        if log_id is None:
            new_log = BackupLog(
                backup_task=backup_task,
                started_at=started_at,
                finished_at=finished_at,
                status=status,
                message=message,
                file_path=file_path,
            )
            new_log.save()
        else:
            BackupLog.objects.filter(id=log_id).update(
                finished_at=finished_at,
                status=status,
                message=message,
            )


def backup_free(backup_task, now):
    """
    Delete backup files and logs older than the backup task retention period.

    Retention is measured in days.
    Only logs/files for the given backup_task are affected.
    """

    delete_datetime = now - timedelta(days=backup_task.retention)

    old_backup_logs = BackupLog.objects.filter(
        backup_task=backup_task,
        created_at__lt=delete_datetime,
    )

    deleted_files = 0
    missing_files = 0
    failed_files = 0

    for backup_log in old_backup_logs:
        file_path = backup_log.file_path

        if not file_path:
            continue

        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_files += 1
            else:
                missing_files += 1

        except Exception as e:
            failed_files += 1
            logger.warning(
                f"Could not delete old backup file '{file_path}': {e}"
            )

    deleted_logs_count, _ = old_backup_logs.delete()

    logger.info(
        f"Backup cleanup completed for task {backup_task.id}. "
        f"Deleted files: {deleted_files}. "
        f"Missing files: {missing_files}. "
        f"Failed file deletes: {failed_files}. "
        f"Deleted logs: {deleted_logs_count}."
    )


def backup_process(_entry):
    backup_dir = '/data/backup/'

    log_id = None
    file_name = None
    file_path = None    

    started_at = pytz.UTC.localize(datetime.now())
    
    try:
        file_name = f"{started_at.strftime('%Y%m%d_%I%M_%p')}_{_entry.file_name}.tar.gz"

        file_path = os.path.join(backup_dir, file_name)

        log_id = backup_set_running(_entry, started_at, file_path)

        backup_create(file_path)

        if _entry.ftp_server is not None:
            try:
                
                # send backup via FTP
                backup_ftp(file_name, file_path, _entry.ftp_server, _entry.remote_folder)

                status = 'SUCCESS'
                message = 'Backup created and successfully send via FTP.'
            except Exception as e:
                status = 'FTP ERROR'
                message = str(e)
                logger.error('Exception happened during backup ftp %s' %(e))                    
        else:
            status = 'SUCCESS'
            message = 'Backup created.'

    except Exception as e:
        status = 'BACKUP ERROR'
        message = str(e)
        logger.error('Exception happened during backup creation %s' %(e))
    finally:
        finished_at = pytz.UTC.localize(datetime.now())

        backup_log(log_id, _entry, started_at, finished_at, status, message, file_path)

        backup_free(_entry, started_at)


def backup_free_orphaned_files(orphan_retention_days=30):
    """
    Delete old backup files that are no longer connected to an existing BackupTask.

    Why this exists:
    ----------------
    Normal cleanup uses BackupTask.retention.

    But if a BackupTask is deleted, old backup files may remain in /data/backup/
    because there is no BackupTask left to define their retention schedule.

    This function handles those orphaned files using a global orphan retention
    period, such as 30 days.

    Safety rules:
    -------------
    1. Only delete files older than orphan_retention_days.
    2. Only delete .tar.gz files from the backup directory.
    3. Do not delete files that are still referenced by logs belonging to
       existing BackupTasks.
    4. Keep orphaned BackupLog records until their file is deleted or missing.
    """

    backup_dir = Path("/data/backup/")
    cutoff_datetime = datetime.now(timezone.utc) - timedelta(days=orphan_retention_days)

    deleted_files = 0
    skipped_files = 0
    failed_files = 0
    deleted_file_paths = []

    if not backup_dir.exists():
        logger.warning(f"Backup directory does not exist: {backup_dir}")
        return {
            "deleted_files": deleted_files,
            "skipped_files": skipped_files,
            "failed_files": failed_files,
            "deleted_orphan_logs": 0,
        }

    # Get IDs of BackupTasks that still exist.
    existing_backup_task_ids = set(
        BackupTask.objects.values_list("id", flat=True)
    )

    # Get file paths that still belong to existing BackupTasks.
    # These should be handled by backup_free(), not this orphan cleanup.
    valid_backup_file_paths = set(
        BackupLog.objects.filter(
            backup_task_id__in=existing_backup_task_ids,
            file_path__isnull=False,
        ).values_list("file_path", flat=True)
    )

    # Find old .tar.gz files in the backup directory.
    for backup_file in backup_dir.glob("*.tar.gz"):
        try:
            if not backup_file.is_file():
                skipped_files += 1
                continue

            file_path = str(backup_file)

            # If this file still belongs to an existing BackupTask,
            # do not touch it here.
            if file_path in valid_backup_file_paths:
                skipped_files += 1
                continue

            # Use modified time, not ctime.
            # On Linux, ctime is metadata-change time, not true creation time.
            file_modified_datetime = datetime.fromtimestamp(
                backup_file.stat().st_mtime,
                tz=timezone.utc,
            )

            if file_modified_datetime < cutoff_datetime:
                backup_file.unlink()
                deleted_files += 1
                deleted_file_paths.append(file_path)
            else:
                skipped_files += 1

        except Exception as e:
            failed_files += 1
            logger.warning(
                f"Could not delete orphaned backup file '{backup_file}': {e}"
            )

    # Now clean orphaned logs only when their file has actually been deleted
    # or when the file path is missing/empty.
    orphan_logs = BackupLog.objects.exclude(
        backup_task_id__in=existing_backup_task_ids
    )

    deleted_orphan_logs_count = 0

    if deleted_file_paths:
        count, _ = orphan_logs.filter(file_path__in=deleted_file_paths).delete()

        deleted_orphan_logs_count += count

    count, _ = orphan_logs.filter(file_path__isnull=True).delete()

    deleted_orphan_logs_count += count

    logger.info(
        f"Orphaned backup cleanup completed. "
        f"Deleted files: {deleted_files}. "
        f"Skipped files: {skipped_files}. "
        f"Failed deletes: {failed_files}. "
        f"Deleted orphaned logs: {deleted_orphan_logs_count}."
    )

    return {
        "deleted_files": deleted_files,
        "skipped_files": skipped_files,
        "failed_files": failed_files,
        "deleted_orphan_logs": deleted_orphan_logs_count,
    }


@shared_task
def backup_postgres():

    lock_key = "backup_postgres_lock"
    lock_timeout = 60 * 60 * 6  # 6 hours

    # If lock already exists, another backup_postgres is still running.
    if not cache.add(lock_key, "running", timeout=lock_timeout):
        logger.warning("backup_postgres skipped because another backup_postgres is already running.")
        return

    try:
        now = pytz.UTC.localize(datetime.now())

        _backup_tasks = BackupTask.objects.filter(is_active=True)

        for _entry in _backup_tasks:
            if croniter.match(_entry.cron_schedule, now):
                backup_process(_entry)

        # Clean up old backup files that no longer belong to any existing BackupTask.
        backup_free_orphaned_files(orphan_retention_days=30)

    finally:
        cache.delete(lock_key)

############################################################


@shared_task
def calculate_hourly_summary(start_datetime=None, end_datetime=None, station_id_list=None):
    # IMPORTANT: Hourly summaries are ran on SUSPICIOUS, GOOD and NOT CHECKED
    start_at = time()

    if not start_datetime:
        start_datetime = datetime.today() - timedelta(hours=48)

    if not end_datetime:
        end_datetime = datetime.today()

    if start_datetime.tzinfo is None:
        start_datetime = pytz.UTC.localize(start_datetime)

    if end_datetime.tzinfo is None:
        end_datetime = pytz.UTC.localize(end_datetime)

    if start_datetime > end_datetime:
        print('Error - start date is more recent than end date.')
        return
    
    # ✅ Normalize to the hour bucket
    hour_start = start_datetime.replace(minute=0, second=0, microsecond=0)
    hour_end = hour_start + timedelta(hours=1)

    if station_id_list is None:
        station_ids = Station.objects.filter(is_active=True).values_list('id', flat=True)
    else:
        station_ids = station_id_list

    station_ids = list(station_ids)

    logger.info('Hourly summary started at {}'.format(datetime.today()))
    logger.info('Hourly summary parameters: {} {} {}'.format(start_datetime, end_datetime, station_id_list))

    delete_sql = """
        DELETE FROM hourly_summary
        WHERE station_id = ANY(%(station_ids)s)
          AND datetime = %(hour_start)s
    """

    insert_sql = """
        INSERT INTO hourly_summary (
            datetime,
            station_id,
            variable_id,
            min_value,
            max_value,
            avg_value,
            sum_value,
            num_records,
            created_at,
            updated_at
        )
        SELECT 
            values.datetime,
            values.station_id,
            values.variable_id,
            values.min_value,
            values.max_value,
            values.avg_value,
            values.sum_value,
            values.num_records,
            now(),
            now()
        FROM
            (SELECT 
                CASE
                  WHEN rd.datetime = rd.datetime::date
                  THEN date_trunc('hour', rd.datetime - '1 second'::interval)
                  ELSE date_trunc('hour', rd.datetime)
                END as datetime,
                station_id,
                variable_id,
                min(calc.value) AS min_value,
                max(calc.value) AS max_value,
                avg(calc.value) AS avg_value,
                sum(calc.value) AS sum_value,
                count(calc.value) AS num_records
            FROM 
                raw_data rd
                ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
            WHERE rd.datetime >= %(hour_start)s
              AND rd.datetime <  %(hour_end)s
              AND (rd.manual_flag in (1,4) OR (rd.manual_flag IS NULL AND rd.quality_flag in (1,2,4)))
              AND NOT rd.is_daily
              AND calc.value != %(MISSING_VALUE)s
              AND station_id = ANY(%(station_ids)s) 
            GROUP BY 1,2,3) values
        WHERE values.datetime = %(hour_start)s

        UNION ALL

        SELECT 
            values.datetime,
            values.station_id,
            values.variable_id,
            values.min_value,
            values.max_value,
            values.avg_value,
            values.sum_value,
            values.num_records,
            now(),
            now()
        FROM
            (SELECT date_trunc('hour', rd.datetime) as datetime,
                station_id,
                variable_id,
                min(calc.value) AS min_value,
                max(calc.value) AS max_value,
                avg(calc.value) AS avg_value,
                sum(calc.value) AS sum_value,
                count(calc.value) AS num_records
            FROM 
                raw_data rd
                ,LATERAL (SELECT CASE WHEN rd.consisted IS NOT NULL THEN rd.consisted ELSE rd.measured END as value) AS calc
            WHERE rd.datetime >= %(hour_start)s
              AND rd.datetime <  %(hour_end)s
              AND (rd.manual_flag in (1,4) OR (rd.manual_flag IS NULL AND rd.quality_flag in (1,2,4)))
              AND rd.is_daily
              AND calc.value != %(MISSING_VALUE)s
              AND station_id = ANY(%(station_ids)s) 
            GROUP BY 1,2,3) values
        WHERE values.datetime = %(hour_start)s
    """


    conn = get_connection()

    with conn.cursor() as cursor:
        cursor.execute(delete_sql, {
            "station_ids": station_ids,
            "hour_start": hour_start,
        })
        cursor.execute(insert_sql, {
            "station_ids": station_ids,
            "hour_start": hour_start,
            "hour_end": hour_end,
            "MISSING_VALUE": settings.MISSING_VALUE,
        })
    conn.commit()
    conn.close()
    logger.info(f'Hourly summary finished at {datetime.now(pytz.UTC)}. Took {time() - start_at} seconds.')


@shared_task
def calculate_daily_summary(start_date=None, end_date=None, station_id_list=None):
    # IMPORTANT: Daily summaries are run on SUSPICIOUS, GOOD and NOT CHECKED.
    #
    # IMPORTANT DAILY WINDOW BEHAVIOR:
    # This function intentionally uses:
    #
    #     rd.datetime >  datetime_start
    #     rd.datetime <= datetime_end


    """
    Calculate (or recalculate) DAILY summaries.

    IMPORTANT DESIGN RULES (DO NOT BREAK):
    --------------------------------------
    1. DAILY summaries are keyed by a *logical local DATE* (daily_summary.day).

    2. DELETES must ALWAYS be done using DATE equality, NEVER datetime ranges.

    This is important because daily_summary.day is a DATE column, while raw_data
    selection uses timezone-adjusted datetime ranges.

    3. Datetime ranges are ONLY for selecting raw_data, NOT for deleting summaries.

    4. This function may be called:
    - as a normal scheduled daily run
    - as a QC-triggered recalculation for past dates

    Therefore it MUST be safe to run repeatedly and out-of-order.

    =======================================================================================

    DAILY SUMMARY TIME WINDOW BEHAVIOR (READ CAREFULLY):
    ----------------------------------------------------

    This function intentionally uses this raw_data selection window:

        rd.datetime >  datetime_start
        rd.datetime <= datetime_end

    Where:

        datetime_start = local DATE at 00:00 converted to UTC
        datetime_end   = datetime_start + 1 day

    This is a non-standard window. It is effectively:

        (datetime_start, datetime_end]

    instead of the more common:

        [datetime_start, datetime_end)

    EFFECTS:
    --------

    1. A raw_data record exactly at local 00:00:00 of the target date
    (rd.datetime == datetime_start) is EXCLUDED.

    2. A raw_data record exactly at local 00:00:00 of the next day
    (rd.datetime == datetime_end) is INCLUDED.

    3. For NON-DAILY records:

    The logical date is calculated using:

        local_datetime - 1 second

    This means a non-daily record exactly at local midnight of the next day
    is assigned to the PREVIOUS logical date.

    Example:
        2026-05-09 00:00:00 local
        minus 1 second
        -> 2026-05-08 23:59:59 local
        -> daily_summary.day = 2026-05-08

    4. For DAILY records:

    The logical date is calculated WITHOUT subtracting 1 second.

    This means a daily record exactly at local midnight of the next day
    is assigned to the NEXT logical date.

    Example:
        2026-05-09 00:00:00 local
        -> daily_summary.day = 2026-05-09

    5. This behavior is intentional based on the current system logic and should
    not be changed casually, because it affects how midnight-boundary records
    are assigned to daily summaries.

    WARNING:
    --------

    Do NOT change the inequality operators:

        rd.datetime >  datetime_start
        rd.datetime <= datetime_end

    or the non-daily '- 1 second' adjustment unless you also update the logical
    definition of what "daily" means for this system.

    If you want the standard behavior where:

        [datetime_start, datetime_end)

    is used instead, then BOTH the WHERE clause and the date-casting logic must
    be reviewed and changed together.
    """


    logger.info(
        f'DAILY SUMMARY started at {datetime.now(tz=pytz.UTC)} with parameters: '
        f'start_date={start_date} end_date={end_date} '
        f'station_id_list={station_id_list}'
    )

    start_at = time()

    # If no dates are provided, default to today only.
    if start_date is None or end_date is None:
        start_date = datetime.now(pytz.UTC).date()
        end_date = start_date + timedelta(days=1)

    # end_date is treated as exclusive in the Python loop.
    if start_date >= end_date:
        logger.error("Invalid date range: start_date must be < end_date")
        return

    conn = get_connection()

    try:
        with conn.cursor() as cursor:
            # Determine which stations should be processed.
            if station_id_list is None:
                stations = Station.objects.filter(is_active=True)
            else:
                stations = Station.objects.filter(id__in=station_id_list)

            # Group stations by UTC offset.
            # This is necessary because "daily" is based on each station's local day.
            offsets = set(stations.values_list("utc_offset_minutes", flat=True))

            for offset in offsets:
                station_ids = list(
                    stations
                    .filter(utc_offset_minutes=offset)
                    .values_list("id", flat=True)
                )

                if not station_ids:
                    continue

                fixed_offset = pytz.FixedOffset(offset)

                current_date = start_date

                # Process one logical date at a time.
                # This keeps each transaction smaller and safer.
                while current_date < end_date:
                    # Convert the station-local date window to UTC.
                    #
                    # Example:
                    # If station local day starts at 2026-05-08 00:00 local,
                    # this converts that local midnight into UTC.
                    datetime_start = datetime(
                        current_date.year,
                        current_date.month,
                        current_date.day,
                        0, 0, 0,
                        tzinfo=fixed_offset
                    ).astimezone(pytz.UTC)

                    datetime_end = datetime_start + timedelta(days=1)

                    logger.info(
                        "Daily Summary | date=%s offset=%s stations=%d "
                        "utc_window=(%s, %s]",
                        current_date,
                        offset,
                        len(station_ids),
                        datetime_start,
                        datetime_end,
                    )

                    # Delete only the current logical date for the selected stations.
                    #
                    # Important:
                    # daily_summary.day is a DATE column, so we delete by date equality.
                    # Do not delete using datetime ranges here.
                    delete_sql = """
                        DELETE FROM daily_summary
                        WHERE station_id = ANY(%(station_ids)s)
                          AND day = %(target_date)s
                    """

                    cursor.execute(
                        delete_sql,
                        {
                            "station_ids": station_ids,
                            "target_date": current_date,
                        }
                    )

                    # Optimized insert:
                    #
                    # The old version scanned raw_data twice:
                    #   1. once for NOT rd.is_daily
                    #   2. once for rd.is_daily
                    #
                    # This version scans the target raw_data window once and uses
                    # CASE to calculate the correct logical day.
                    #
                    # Non-daily records subtract 1 second before casting to DATE.
                    # Daily records do not subtract 1 second.
                    #
                    # This preserves the midnight-boundary behavior from the old query.
                    insert_sql = """
                        INSERT INTO daily_summary (
                            "day",
                            station_id,
                            variable_id,
                            min_value,
                            max_value,
                            avg_value,
                            sum_value,
                            num_records,
                            created_at,
                            updated_at
                        )
                        WITH filtered AS (
                            SELECT
                                CASE
                                    WHEN rd.is_daily THEN
                                        CAST(
                                            (rd.datetime + interval '%(offset)s minutes')
                                            AT TIME ZONE 'utc'
                                            AS DATE
                                        )
                                    ELSE
                                        CAST(
                                            (rd.datetime + interval '%(offset)s minutes')
                                            AT TIME ZONE 'utc'
                                            - '1 second'::interval
                                            AS DATE
                                        )
                                END AS day,

                                rd.station_id,
                                rd.variable_id,

                                CASE
                                    WHEN rd.consisted IS NOT NULL THEN rd.consisted
                                    ELSE rd.measured
                                END AS value

                            FROM raw_data rd
                            WHERE rd.datetime > %(datetime_start)s
                              AND rd.datetime <= %(datetime_end)s
                              AND rd.station_id = ANY(%(station_ids)s)

                              -- Daily summaries are calculated using:
                              -- GOOD, SUSPICIOUS, and NOT CHECKED values,
                              -- while also respecting accepted manual flags.
                              AND (
                                  rd.manual_flag IN (1, 4)
                                  OR (
                                      rd.manual_flag IS NULL
                                      AND rd.quality_flag IN (1, 2, 4)
                                  )
                              )
                        )
                        SELECT
                            day,
                            station_id,
                            variable_id,
                            MIN(value),
                            MAX(value),
                            AVG(value),
                            SUM(value),
                            COUNT(value),
                            now(),
                            now()
                        FROM filtered
                        WHERE value != %(MISSING_VALUE)s
                        GROUP BY
                            day,
                            station_id,
                            variable_id

                        ON CONFLICT (day, station_id, variable_id) DO UPDATE
                        SET min_value   = EXCLUDED.min_value,
                            max_value   = EXCLUDED.max_value,
                            avg_value   = EXCLUDED.avg_value,
                            sum_value   = EXCLUDED.sum_value,
                            num_records = EXCLUDED.num_records,
                            updated_at  = now();
                    """

                    cursor.execute(
                        insert_sql,
                        {
                            "datetime_start": datetime_start,
                            "datetime_end": datetime_end,
                            "station_ids": station_ids,
                            "offset": offset,
                            "MISSING_VALUE": settings.MISSING_VALUE,
                        }
                    )

                    # Commit after each logical date.
                    # This keeps transactions smaller and reduces how long locks are held.
                    conn.commit()

                    current_date += timedelta(days=1)

    except Exception:
        conn.rollback()
        logger.exception("Error calculating daily summary.")
        raise

    finally:
        conn.close()

    cache.set("daily_summary_last_run", datetime.now(pytz.UTC), None)

    logger.info(
        f'Daily summary finished at {datetime.now(pytz.UTC)}. '
        f'Took {time() - start_at:.2f} seconds.'
    )



@shared_task
def calculate_station_minimum_interval(start_date=None, end_date=None, station_id_list=None):
    logger.info(
        f'CALCULATE STATION MINIMUM INTERVAL started at {datetime.now(tz=pytz.UTC)} '
        f'with parameters: start_date={start_date} end_date={end_date} '
        f'station_id_list={station_id_list}'
    )

    start_at = time()

    if start_date is None or end_date is None:
        start_date = datetime.now(pytz.UTC).date()
        end_date = (datetime.now(pytz.UTC) + timedelta(days=1)).date()

    if start_date > end_date:
        logger.error('Error - start_date is more recent than end_date.')
        return

    conn = get_connection()

    try:
        with conn.cursor() as cursor:
            if station_id_list is None:
                stations = Station.objects.filter(is_active=True)
            else:
                stations = Station.objects.filter(id__in=station_id_list)

            offsets = list(set(stations.values_list('utc_offset_minutes', flat=True)))

            insert_minimum_data_interval = """
                INSERT INTO wx_stationdataminimuminterval (
                     datetime
                    ,station_id
                    ,variable_id
                    ,minimum_interval
                    ,record_count
                    ,ideal_record_count
                    ,record_count_percentage
                    ,created_at
                    ,updated_at
                )
                WITH station_variables AS (
                    SELECT
                        sv.station_id,
                        sv.variable_id
                    FROM wx_stationvariable sv
                    WHERE sv.station_id = ANY(%(station_ids)s)
                ),
                raw_day AS (
                    SELECT
                        rd.station_id,
                        rd.variable_id,
                        rd.datetime,
                        rd.is_daily,
                        CASE
                            WHEN rd.is_daily THEN INTERVAL '24 hours'
                            ELSE LEAD(rd.datetime, 1) OVER (
                                PARTITION BY rd.station_id, rd.variable_id
                                ORDER BY rd.datetime
                            ) - rd.datetime
                        END AS data_interval
                    FROM raw_data rd
                    INNER JOIN station_variables sv
                        ON sv.station_id = rd.station_id
                       AND sv.variable_id = rd.variable_id
                    WHERE rd.datetime > %(raw_start)s
                      AND rd.datetime <= %(raw_end)s
                ),
                grouped_data AS (
                    SELECT
                        sv.station_id,
                        sv.variable_id,
                        MIN(rd.data_interval) AS minimum_interval,
                        COUNT(rd.datetime) AS record_count
                    FROM station_variables sv
                    LEFT JOIN raw_day rd
                        ON rd.station_id = sv.station_id
                       AND rd.variable_id = sv.variable_id
                    GROUP BY
                        sv.station_id,
                        sv.variable_id
                )
                SELECT
                     %(summary_day)s AS datetime
                    ,station_id
                    ,variable_id
                    ,minimum_interval
                    ,record_count
                    ,COALESCE(
                        EXTRACT(EPOCH FROM INTERVAL '1 day')
                        / NULLIF(EXTRACT(EPOCH FROM minimum_interval), 0),
                        0
                    ) AS ideal_record_count
                    ,COALESCE(
                        record_count / NULLIF(
                            EXTRACT(EPOCH FROM INTERVAL '1 day')
                            / NULLIF(EXTRACT(EPOCH FROM minimum_interval), 0),
                            0
                        ) * 100,
                        0
                    ) AS record_count_percentage
                    ,NOW()
                    ,NOW()
                FROM grouped_data
                ON CONFLICT (datetime, station_id, variable_id)
                DO UPDATE SET
                     minimum_interval = excluded.minimum_interval
                    ,record_count = excluded.record_count
                    ,ideal_record_count = excluded.ideal_record_count
                    ,record_count_percentage = excluded.record_count_percentage
                    ,updated_at = NOW()
            """

            for offset in offsets:
                station_ids = list(
                    stations
                    .filter(utc_offset_minutes=offset)
                    .values_list('id', flat=True)
                )

                if not station_ids:
                    continue

                fixed_offset = pytz.FixedOffset(offset)

                logger.info(
                    f'Processing minimum interval for offset={offset}, '
                    f'station_count={len(station_ids)}, station_ids={station_ids}'
                )

                current_date = start_date

                # Preserve the original generate_series behavior:
                # generate_series(start_date, end_date, '1 day') includes end_date.
                while current_date <= end_date:
                    next_date = current_date + timedelta(days=1)

                    # This is the datetime stored in wx_stationdataminimuminterval.datetime.
                    # It matches the old query's current_day value, which was UTC midnight.
                    summary_day = datetime(
                        current_date.year,
                        current_date.month,
                        current_date.day,
                        0, 0, 0,
                        tzinfo=pytz.UTC
                    )

                    # These preserve the old local-day window logic:
                    # old lower bound: rd.datetime > current_day - offset_interval
                    # old upper bound: rd.datetime <= current_day + 1 day - offset_interval
                    raw_start = datetime(
                        current_date.year,
                        current_date.month,
                        current_date.day,
                        0, 0, 0,
                        tzinfo=fixed_offset
                    ).astimezone(pytz.UTC)

                    raw_end = datetime(
                        next_date.year,
                        next_date.month,
                        next_date.day,
                        0, 0, 0,
                        tzinfo=fixed_offset
                    ).astimezone(pytz.UTC)

                    logger.info(
                        f'Calculating minimum interval | summary_day={summary_day} '
                        f'raw_start={raw_start} raw_end={raw_end} '
                        f'offset={offset} station_count={len(station_ids)}'
                    )

                    cursor.execute(
                        insert_minimum_data_interval,
                        {
                            "summary_day": summary_day,
                            "raw_start": raw_start,
                            "raw_end": raw_end,
                            "station_ids": station_ids,
                        }
                    )

                    conn.commit()

                    current_date = next_date

    except Exception:
        conn.rollback()
        logger.exception('Error calculating station minimum interval.')
        raise

    finally:
        conn.close()

    logger.info(
        f'Calculate minimum interval finished at {datetime.now(pytz.UTC)}. '
        f'Took {time() - start_at} seconds.'
    )



# retrieved data inventory (Individual MONTH VIEW)
@shared_task
def data_inventory_month_view(year, month, station_id, variable_id):
    """
    Build the monthly data inventory for one station + one variable.

    This returns one dictionary per day in the selected month.

    IMPORTANT:
    ----------
    This version intentionally keeps the original date behavior.

    We are still using Python date objects:
        query_start_datetime = date(year, month, 1)
        query_end_datetime   = first day of next month

    So this does NOT intentionally change timezone handling.

    Performance design:
    -------------------
    1. Use datetime range filters instead of EXTRACT(YEAR/MONTH) in WHERE.
       This allows PostgreSQL to use the existing datetime indexes.

    2. Aggregate raw_data once for the month in raw_qc.
       This avoids counting raw_data separately for each day.

    3. Use generate_series() to create the days of the month.

    4. Add detailed timing logs so we can see whether the delay is:
       - opening the DB cursor
       - executing SQL
       - fetching rows
       - building the Python response
    """

    task_start = time()

    logger.info(
        f"data_inventory_month_view STARTED: "
        f"year={year}, month={month}, station_id={station_id}, variable_id={variable_id}"
    )

    # ---------------------------------------------------------------------
    # Calculate query dates.
    #
    # We use a half-open range:
    #
    #   datetime >= query_start_datetime
    #   datetime <  query_end_datetime
    #
    # This is safer than BETWEEN because BETWEEN includes the end value too.
    # ---------------------------------------------------------------------
    query_start_datetime = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)

    if month == 12:
        query_end_datetime = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    else:
        query_end_datetime = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=timezone.utc)

    # ---------------------------------------------------------------------
    # Optimized SQL.
    #
    # available_days:
    #   Creates one row per day in the requested month.
    #
    # station_days:
    #   Gets the monthly inventory rows from wx_stationdataminimuminterval.
    #
    # raw_qc:
    #   Gets all raw_data rows for this station + variable + month once,
    #   groups them by day, and calculates QC counts.
    #
    # Final SELECT:
    #   Joins available_days to station_days and raw_qc so every day appears,
    #   even days with no data.
    # ---------------------------------------------------------------------
    query = """
        WITH available_days AS (
            SELECT 
                gs::date AS day_date,
                EXTRACT(DAY FROM gs)::int AS custom_day,
                EXTRACT(DOW FROM gs)::int AS dow
            FROM generate_series(
                %(query_start_datetime)s::date,
                (%(query_end_datetime)s::date - INTERVAL '1 day'),
                INTERVAL '1 day'
            ) AS gs
        ),

        station_days AS (
            SELECT 
                EXTRACT(DAY FROM station_data.datetime)::int AS day,
                TRUNC(station_data.record_count_percentage::numeric, 2) AS percentage,
                station_data.record_count,
                station_data.ideal_record_count
            FROM wx_stationdataminimuminterval AS station_data
            WHERE station_data.station_id = %(station_id)s
                AND station_data.variable_id = %(variable_id)s
                AND station_data.datetime >= %(query_start_datetime)s
                AND station_data.datetime < %(query_end_datetime)s
        ),

        raw_qc AS (
            SELECT
                EXTRACT(DAY FROM rd.datetime)::int AS day,

                -- Count all raw_data records for the day.
                -- raw_data does not have an id column, so we count datetime.
                COUNT(rd.datetime) AS qc_amount,

                -- Count only records whose effective QC flag is considered passed.
                -- COALESCE means manual_flag is used when it exists;
                -- otherwise quality_flag is used.
                COUNT(rd.datetime) FILTER (
                    WHERE COALESCE(rd.manual_flag, rd.quality_flag) IN (1, 4)
                ) AS qc_passed_amount

            FROM raw_data rd
            WHERE rd.station_id = %(station_id)s
                AND rd.variable_id = %(variable_id)s
                AND rd.datetime >= %(query_start_datetime)s
                AND rd.datetime < %(query_end_datetime)s

            GROUP BY EXTRACT(DAY FROM rd.datetime)::int
        )

        SELECT 
            available_days.custom_day,
            available_days.dow,

            COALESCE(station_days.percentage, 0) AS percentage,
            COALESCE(station_days.record_count, 0) AS record_count,
            COALESCE(station_days.ideal_record_count, 0) AS ideal_record_count,

            CASE 
                WHEN COALESCE(raw_qc.qc_amount, 0) = 0 THEN 0
                ELSE TRUNC(
                    (raw_qc.qc_passed_amount / raw_qc.qc_amount::numeric) * 100,
                    2
                )
            END AS qc_passed_percentage

        FROM available_days

        LEFT JOIN station_days
            ON station_days.day = available_days.custom_day

        LEFT JOIN raw_qc
            ON raw_qc.day = available_days.custom_day

        ORDER BY available_days.custom_day;
    """

    query_params = {
        "query_start_datetime": query_start_datetime,
        "query_end_datetime": query_end_datetime,
        "station_id": station_id,
        "variable_id": variable_id,
    }

    # ---------------------------------------------------------------------
    # Execute SQL with detailed timing.
    #
    # This helps us see exactly where the 19 seconds is happening.
    # ---------------------------------------------------------------------
    db_block_start = time()

    logger.info("data_inventory_month_view opening DB cursor...")

    cursor_open_start = time()

    with connection.cursor() as cursor:
        cursor_open_end = time()

        logger.info(
            f"data_inventory_month_view cursor opened in "
            f"{cursor_open_end - cursor_open_start:.3f} seconds."
        )

        logger.info("data_inventory_month_view executing SQL...")

        execute_start = time()

        cursor.execute(query, query_params)

        execute_end = time()

        logger.info(
            f"data_inventory_month_view SQL execute finished in "
            f"{execute_end - execute_start:.3f} seconds."
        )

        # -----------------------------------------------------------------
        # Log the exact SQL that Django sent to PostgreSQL.
        #
        # Use this SQL in psql with EXPLAIN ANALYZE to compare against
        # what Celery is actually running.
        # -----------------------------------------------------------------
        try:
            executed_sql = cursor.query

            if isinstance(executed_sql, bytes):
                executed_sql = executed_sql.decode("utf-8")

            logger.info(
                "data_inventory_month_view executed SQL:\n%s",
                executed_sql
            )

        except Exception as e:
            logger.warning(
                f"data_inventory_month_view could not log cursor.query: {e}"
            )

        logger.info("data_inventory_month_view fetching SQL rows...")

        fetch_start = time()

        rows = cursor.fetchall()

        fetch_end = time()

        logger.info(
            f"data_inventory_month_view fetchall finished in "
            f"{fetch_end - fetch_start:.3f} seconds. "
            f"Rows returned: {len(rows)}"
        )

    db_block_end = time()

    logger.info(
        f"data_inventory_month_view total DB block finished in "
        f"{db_block_end - db_block_start:.3f} seconds."
    )

    # ---------------------------------------------------------------------
    # Convert SQL rows into the same structure the frontend expects.
    # This should not change the output shape.
    # ---------------------------------------------------------------------
    build_start = time()

    days = []

    for row in rows:
        days.append({
            "day": row[0],
            "dow": row[1],
            "percentage": row[2],
            "record_count": row[3],
            "ideal_record_count": row[4],
            "qc_passed_percentage": row[5],
        })

    build_end = time()

    logger.info(
        f"data_inventory_month_view response build finished in "
        f"{build_end - build_start:.3f} seconds."
    )

    logger.info(
        f"data_inventory_month_view TOTAL finished in "
        f"{time() - task_start:.3f} seconds."
    )

    return days




@shared_task
def calculate_last24h_summary():
    # IMPORTANT: last24h summaries are run on SUSPICIOUS, GOOD and NOT CHECKED
    start_at = time()
    run_datetime = datetime.now(pytz.UTC)
    end_datetime = run_datetime
    start_datetime = end_datetime - timedelta(hours=24)

    logger.info(
        "Last 24h summary started | start_datetime=%s | end_datetime=%s",
        start_datetime,
        end_datetime,
    )

    conn = get_connection()

    try:
        with conn.cursor() as cursor:
            sql_insert = f"""
                INSERT INTO last24h_summary (
                    datetime,
                    station_id,
                    variable_id,
                    min_value,
                    max_value,
                    avg_value,
                    sum_value,
                    num_records,
                    latest_value
                )
                WITH filtered AS (
                    SELECT
                        rd.datetime,
                        rd.station_id,
                        rd.variable_id,
                        CASE
                            WHEN rd.consisted IS NOT NULL THEN rd.consisted
                            ELSE rd.measured
                        END AS value
                    FROM raw_data rd
                    WHERE rd.datetime > %(start_datetime)s
                      AND rd.datetime <= %(end_datetime)s
                      AND rd.is_daily = false
                      AND (
                            rd.consisted IS NOT NULL
                            OR rd.quality_flag IN (1, 2, 4)
                      )
                      AND (
                            CASE
                                WHEN rd.consisted IS NOT NULL THEN rd.consisted
                                ELSE rd.measured
                            END
                          ) != %(missing_value)s
                ),
                agg AS (
                    SELECT
                        station_id,
                        variable_id,
                        MIN(value) AS min_value,
                        MAX(value) AS max_value,
                        AVG(value) AS avg_value,
                        SUM(value) AS sum_value,
                        COUNT(value) AS num_records
                    FROM filtered
                    GROUP BY
                        station_id,
                        variable_id
                ),
                latest AS (
                    SELECT DISTINCT ON (station_id, variable_id)
                        station_id,
                        variable_id,
                        value AS latest_value
                    FROM filtered
                    ORDER BY
                        station_id,
                        variable_id,
                        datetime DESC
                )
                SELECT
                    %(run_datetime)s AS datetime,
                    agg.station_id,
                    agg.variable_id,
                    agg.min_value,
                    agg.max_value,
                    agg.avg_value,
                    agg.sum_value,
                    agg.num_records,
                    latest.latest_value
                FROM agg
                JOIN latest
                  ON latest.station_id = agg.station_id
                 AND latest.variable_id = agg.variable_id
                ON CONFLICT (station_id, variable_id)
                DO UPDATE SET
                    min_value = excluded.min_value,
                    max_value = excluded.max_value,
                    avg_value = excluded.avg_value,
                    sum_value = excluded.sum_value,
                    num_records = excluded.num_records,
                    latest_value = excluded.latest_value,
                    datetime = excluded.datetime;
            """

            cursor.execute(
                sql_insert,
                {
                    "start_datetime": start_datetime,
                    "end_datetime": end_datetime,
                    "run_datetime": run_datetime,
                    "missing_value": settings.MISSING_VALUE,
                },
            )

        conn.commit()

    except Exception:
        conn.rollback()
        logger.exception("Error calculating last 24h summary.")
        raise

    finally:
        conn.close()

    cache.set("last24h_summary_last_run", datetime.now(pytz.UTC), None)

    logger.info(
        "Last 24h summary finished | took=%.2fs",
        time() - start_at,
    )



@shared_task
def process_document():
    available_decoders = {
        'HOBO': read_file_hobo,
        'TOA5': read_file_toa5,
        'Weatherlink JSON': read_file_weatherlink_json,
        'HYDROLOGY': read_file_hydrology
        # Nesa
    }

    default_decoder = 'TOA5'

    document_list = Document.objects.select_related('decoder', 'station').filter(processed=False).order_by('id')[:60]

    logger.info('Documents: %s' % document_list)

    for document in document_list:
        if document.decoder:
            current_decoder = available_decoders[document.decoder.name]
        else:
            current_decoder = available_decoders[default_decoder]

        logger.info('Processing file "{0}" with "{1}" decoder.'.format(document.file.path, current_decoder))

        try:
            current_decoder(document.file.path, document.station)
        except Exception as err:
            logger.error(
                'Error Processing file "{0}" with "{1}" decoder. '.format(document.file.path, current_decoder) + repr(
                    err))
            db_logger.error(
                'Error Processing file "{0}" with "{1}" decoder. '.format(document.file.path, current_decoder) + repr(
                    err))
        else:
            document.processed = True
            document.save()


@shared_task
def dcp_tasks_scheduler():
    logger.info('Inside dcp_tasks_scheduler')

    noaa_list_to_process = []
    for noaaDcp in NoaaDcp.objects.all():
        if noaaDcp.dcp_address == '50203044':
            DEBUG = True
        else:
            DEBUG = False            

        now = pytz.UTC.localize(datetime.now())

        if noaaDcp.last_datetime is None:
            next_execution = now
        else:
            scheduled_execution = datetime(year=now.year,
                                           month=now.month,
                                           day=now.day,
                                           hour=now.hour,
                                           minute=noaaDcp.first_transmission_time.minute,
                                           second=noaaDcp.first_transmission_time.second)

            transmission_window_timedelta = timedelta(minutes=noaaDcp.transmission_window.minute,
                                                      seconds=noaaDcp.transmission_window.second)

            next_execution = scheduled_execution + transmission_window_timedelta
            next_execution = pytz.UTC.localize(next_execution)

        # if DEBUG or (next_execution <= now and (noaaDcp.last_datetime is None or noaaDcp.last_datetime < next_execution)):
        if next_execution <= now and (noaaDcp.last_datetime is None or noaaDcp.last_datetime < next_execution):
            noaa_list_to_process.append({"noaa_object": noaaDcp, "last_execution": noaaDcp.last_datetime})
            noaaDcp.last_datetime = now
            noaaDcp.save()

    for noaa_dcp in noaa_list_to_process:
        try:
            retrieve_dcp_messages(noaa_dcp)
        except Exception as e:
            logging.error(f'dcp_tasks_scheduler ERROR: {repr(e)}')


def retrieve_dcp_messages(noaa_dict):
    current_noaa_dcp = noaa_dict["noaa_object"]
    last_execution = noaa_dict["last_execution"]

    logger.info('Inside retrieve_dcp_messages ' + current_noaa_dcp.dcp_address)

    related_stations = current_noaa_dcp.noaadcpsstation_set
    related_stations_count = related_stations.count()
    if related_stations_count == 0:
        raise Exception(f"The noaa dcp '{current_noaa_dcp}' is not related to any Station.")
    elif related_stations_count != 1:
        raise Exception(f"The noaa dcp '{current_noaa_dcp}' is related to more than one Station.")

    noaa_dcp_station = related_stations.first()
    station_id = noaa_dcp_station.station_id
    decoder = noaa_dcp_station.decoder.name

    available_decoders = {
        'NESA': read_data_nesa,
        'SAT_TX325': read_data_sat_tx325,  
        # 'SURTRON': read_data_surtron,
    }

    set_search_criteria(current_noaa_dcp, last_execution)

    command = subprocess.Popen([settings.LRGS_EXECUTABLE_PATH,
                                '-h', settings.LRGS_SERVER_HOST,
                                '-p', settings.LRGS_SERVER_PORT,
                                '-u', settings.LRGS_USER,
                                '-P', settings.LRGS_PASSWORD,
                                '-f', settings.LRGS_CS_FILE_PATH
                                ], shell=False, stderr=subprocess.PIPE, stdout=subprocess.PIPE)

    output, err_message = command.communicate()
    response = output.decode('ascii')

    try:
        available_decoders[decoder](station_id, current_noaa_dcp.dcp_address, current_noaa_dcp.config_data, response, err_message)
    except Exception as err:
        logger.error(f'Error on retrieve_dcp_messages for dcp address "{current_noaa_dcp.dcp_address}". {repr(err)}')


def set_search_criteria(dcp, last_execution):
    with open(settings.LRGS_CS_FILE_PATH, 'w') as cs_file:
        if dcp.first_channel is not None:
            cs_file.write(
                f"""DRS_SINCE: now - {dcp_query_window(last_execution)} hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp.dcp_address}\nCHANNEL: |{dcp.first_channel}\n""")
        else:
            cs_file.write(
                f"""DRS_SINCE: now - {dcp_query_window(last_execution)} hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp.dcp_address}\n""")


def dcp_query_window(last_execution):
    return max(3, min(latest_received_dpc_data_in_hours(last_execution), int(settings.LRGS_MAX_INTERVAL)))


def latest_received_dpc_data_in_hours(last_execution):
    try:
        return int(((datetime.now().astimezone(pytz.UTC) - last_execution).total_seconds()) / 3600)
    except TypeError as e:
        return 999999


# set search criteria for dcp testing
def set_search_criteria_dcp_test(dcp_info):
    with open(settings.LRGS_CS_FILE_PATH, 'w') as cs_file:
        if dcp_info['first_channel'] is not None:
            cs_file.write(
                f"""DRS_SINCE: now - 1 hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp_info['dcp_address']}\nCHANNEL: |{dcp_info['first_channel']}\n""")
        else:
            cs_file.write(
                f"""DRS_SINCE: now - 1 hour\nDRS_UNTIL: now\nDCP_ADDRESS: {dcp_info['dcp_address']}\n""")


# called when the user attempts to test whether a noaaDCP is able to transmit
def test_dcp_transmit(dcp_info):
    logging.info(f"Attempting to test DCP transmition for {dcp_info['dcp_address']}")

    set_search_criteria_dcp_test(dcp_info)

    command = subprocess.Popen([settings.LRGS_EXECUTABLE_PATH,
                                '-h', settings.LRGS_SERVER_HOST,
                                '-p', settings.LRGS_SERVER_PORT,
                                '-u', settings.LRGS_USER,
                                '-P', settings.LRGS_PASSWORD,
                                '-f', settings.LRGS_CS_FILE_PATH
                                ], shell=False, stderr=subprocess.PIPE, stdout=subprocess.PIPE)

    output, err_message = command.communicate()
    # response = {
    #     "output": output.decode('ascii'),
    #     "err_message": err_message
    # }

    response = output.decode('ascii')

    logging.info(f"{dcp_info['dcp_address']} DCP test transmission output {response}")

    return response


@shared_task
def get_entl_data():
    print('LIGHTNING DATA - Starting get_entl_data task...')
    while True:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as entl_socket:
            try:
                print("LIGHTNING DATA - Connecting to ENTLN server: {}:{}".format(settings.ENTL_PRIMARY_SERVER_HOST,
                                                                                  settings.ENTL_PRIMARY_SERVER_PORT))

                entl_socket.connect((settings.ENTL_PRIMARY_SERVER_HOST, settings.ENTL_PRIMARY_SERVER_PORT))
                entl_socket.settimeout(60)
                entl_socket.setsockopt(socket.SOL_SOCKET, socket.SO_KEEPALIVE, 1)
                entl_socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPIDLE, 1)
                entl_socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPINTVL, 15)
                entl_socket.setsockopt(socket.IPPROTO_TCP, socket.TCP_KEEPCNT, 4)

                print("LIGHTNING DATA - Authenticating...")
                auth_string = '{"p":"%s","v":3,"f":2,"t":1}' % settings.ENTL_PARTNER_ID
                entl_socket.send(auth_string.encode('latin-1'))
                print("LIGHTNING DATA - Authenticated")

                process_received_data(entl_socket)

            except Exception as e:
                print("LIGHTNING DATA - An error occurred: " + repr(
                    e) + "\nLIGHTNING DATA - Reconnecting in 3 seconds...")
                sleep(3)
        print('LIGHTNING DATA - Connection error. Reconnecting in 15 seconds...')
        sleep(15)


def process_received_data(entl_socket):
    while True:
        data = entl_socket.recv(56)
        if not data:
            return

        if data[1] == '9':
            print("LIGHTNING DATA - Keep alive packet")
        else:
            latitude = int.from_bytes(data[10:14], byteorder='big', signed=True) / 1e7
            longitude = int.from_bytes(data[14:18], byteorder='big', signed=True) / 1e7
            if (15 <= latitude <= 19 and -90 <= longitude <= -87):
                save_flash_data.delay(data.decode('latin-1'))


# Used to save Flash data asynchronously
@shared_task
def save_flash_data(data_string):
    read_data_flash(data_string.encode('latin-1'))


def format_export_value_by_precision(value, precision):
    """
    Format an exported numeric value using the variable's configured precision.

    Examples:
        precision = 4
        25.3 -> "25.3000"
        25.34567 -> "25.3457"

    Notes:
        - Missing export values like "-99.9" are left unchanged.
        - Non-numeric values, such as code variables, are left unchanged.
    """

    # Keep pandas empty values unchanged
    if pandas.isna(value):
        return value

    # Keep the missing-data sentinel unchanged
    if str(value) == "-99.9":
        return "-99.9"

    # If no precision is configured, leave the value as-is
    if precision is None:
        return value

    try:
        precision = int(precision)

        # Convert through string to avoid floating-point display issues
        decimal_value = Decimal(str(value))

        # precision=0 means no decimal places
        if precision <= 0:
            quantizer = Decimal("1")
        else:
            quantizer = Decimal("1." + ("0" * precision))

        return str(decimal_value.quantize(quantizer, rounding=ROUND_HALF_UP))

    except (InvalidOperation, ValueError, TypeError):
        # If the value is not numeric, such as a code/string value, leave it alone
        return value


@shared_task
def export_data(station_id, source, start_date, end_date, variable_ids, file_id, agg, aqc_checks, mqc_checks, displayUTC):
    try:
        logger.info(f'Exporting data (file "{file_id}")')

        timezone_offset = pytz.timezone(settings.TIMEZONE_NAME)
        start_date_utc = pytz.UTC.localize(datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'))
        end_date_utc = pytz.UTC.localize(datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S'))

        station = Station.objects.get(pk=station_id)
        current_datafile = DataFile.objects.get(pk=file_id)

        variable_ids = tuple(variable_ids)
        # variable_ids = ','.join([str(x) for x in variable_ids])

        # Diferent data sources have diferents columns names for the measurement data and diferents intervals
        if source == 'raw_data':
            datetime_variable = 'datetime'
            data_source_description = 'Raw data'
            converted_start_date = start_date_utc
            converted_end_date = end_date_utc

        else:
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value
            #          WHEN var.sampling_operation_id = 3      THEN data.min_value
            #          WHEN var.sampling_operation_id = 4      THEN data.max_value
            #          WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            if source == 'hourly_summary':
                datetime_variable = 'datetime'
                date_source = f"(datetime + interval '{station.utc_offset_minutes} minutes') at time zone 'utc' as date"
                data_source_description = 'Hourly summary'
                converted_start_date = start_date_utc
                converted_end_date = end_date_utc

            elif source == 'daily_summary':
                datetime_variable = 'day'
                data_source_description = 'Daily summary'
                date_source = "day::date"
                converted_start_date = start_date_utc.astimezone(timezone_offset).date()
                converted_end_date = end_date_utc.astimezone(timezone_offset).date()

            elif source == 'monthly_summary':
                # measured_source = '''
                #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
                #         WHEN var.sampling_operation_id = 3      THEN data.min_value
                #         WHEN var.sampling_operation_id = 4      THEN data.max_value
                #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
                #     ELSE data.sum_value END as value '''
                datetime_variable = 'date'
                date_source = "date::date"
                data_source_description = 'Monthly summary'
                converted_start_date = start_date_utc.astimezone(timezone_offset).date()
                converted_end_date = end_date_utc.astimezone(timezone_offset).date()

            elif source == 'yearly_summary':
                # measured_source = '''
                #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
                #         WHEN var.sampling_operation_id = 3      THEN data.min_value
                #         WHEN var.sampling_operation_id = 4      THEN data.max_value
                #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
                #     ELSE data.sum_value END as value '''
                datetime_variable = 'date'
                date_source = "date::date"
                data_source_description = 'Yearly summary'
                converted_start_date = start_date_utc.astimezone(timezone_offset).date()
                converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        
        variable_dict = {}
        variable_precision_dict = {}
        variable_type_dict = {}
        variable_names_string = ''

        with connection.cursor() as cursor_variable:
            cursor_variable.execute(f'''
                SELECT var.symbol
                    ,var.id
                    ,CASE WHEN unit.symbol IS NOT NULL THEN CONCAT(var.symbol, ' - ', var.name, ' (', unit.symbol, ')') 
                        ELSE CONCAT(var.symbol, ' - ', var.name) END as var_name
                    ,var.precision
                    ,var.variable_type
                FROM wx_variable var 
                LEFT JOIN wx_unit unit ON var.unit_id = unit.id 
                WHERE var.id IN %s
                ORDER BY var.name
            ''', (variable_ids,))

            rows = cursor_variable.fetchall()

            for row in rows:
                variable_id = row[1]
                variable_symbol = row[0]
                variable_precision = row[3]
                variable_type = row[4]

                # Used later by pandas rename:
                # DB variable ID -> exported CSV column symbol
                variable_dict[variable_id] = variable_symbol

                # Used later to format the exported numeric values
                # after the pivot has created one column per variable.
                variable_precision_dict[variable_symbol] = variable_precision
                variable_type_dict[variable_symbol] = variable_type

                variable_names_string += f'{row[2]}   '

        # Iterate over the start and end date day by day to split the queries
        datetime_list = [converted_start_date]
        current_datetime = converted_start_date
        while current_datetime < converted_end_date and (current_datetime + timedelta(days=1)) < converted_end_date:
            current_datetime = current_datetime + timedelta(days=1)
            datetime_list.append(current_datetime)
        datetime_list.append(converted_end_date)

        # adding an extra day to the list when the source is monthly_summary or daily_summary
        # This will cause the last month/day to be included in the output
        if source in ['monthly_summary', 'daily_summary']:
            datetime_list.append(converted_end_date + timedelta(days=1))

            # in the case of querying for the monthly_summary of only 1 month
            # strip the datetime_list of the last value so that there are no duplicates
            if len(datetime_list) == 3 and source == 'monthly_summary':
                datetime_list.pop()
        # adding an extra hour to the list when the souce is hourly_summary
        elif source == 'hourly_summary':
            datetime_list.append(converted_end_date + timedelta(hours=1))
        # adding extra seconds to the list when the souce is raw_data
        # the number of seconds added is based on the data interval selected
        elif source == 'raw_data':
            datetime_list.append(converted_end_date + timedelta(seconds=current_datafile.interval_in_seconds))

        # quality control clause based on users request
        quality_clause = ""

        # if user request aqc data
        if aqc_checks and mqc_checks:
            quality_clause = """
                AND data.quality_flag = 4
                AND data.manual_flag = 4
            """
        elif aqc_checks:
            quality_clause = """
                AND data.quality_flag = 4
            """
        elif mqc_checks:
            quality_clause = """
                AND data.manual_flag = 4
            """

        query_result = []
        for i in range(0, len(datetime_list) - 1):
            current_start_datetime = datetime_list[i]
            current_end_datetime = datetime_list[i + 1]

            with connection.cursor() as cursor:
                if source == 'raw_data': 

                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_raw_data = f'''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                                {quality_clause}
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_raw_data = f'''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                                {quality_clause}
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': current_datafile.interval_in_seconds})


                    cursor.execute(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': current_datafile.interval_in_seconds})

                elif source == 'hourly_summary':
                    
                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})
                    
                elif source == 'daily_summary':
                    query_daily = '''
                        WITH processed_data AS (
                            SELECT day ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM daily_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.day >= %(start_datetime)s
                            AND data.day <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 days') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON day = generated_time AND variable.id = variable_id
                    '''

                    logging.info(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                elif source == 'monthly_summary':
                    query_monthly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM monthly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 months') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                elif source == 'yearly_summary':
                    query_yearly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM yearly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date < %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 years') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                query_result.extend(cursor.fetchall())



        filepath = f'{settings.EXPORTED_DATA_CELERY_PATH}{file_id}.csv'
        date_of_completion = datetime.utcnow()

        with open(filepath, 'w') as f: 

            # modify the displayed start and end date in the csv file based on the summary type
            # show the hour, minute, second, year, month, day in the output
            if data_source_description in ["Raw data", "Hourly summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')
            # show just the year, month and day
            elif data_source_description in ["Daily summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d')
            # show just the year and month
            elif data_source_description in ["Monthly summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m')
            # show just the year
            elif data_source_description in ["Yearly summary"]:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y')
                end_date_header = int(end_date_utc.astimezone(timezone_offset).strftime('%Y')) - 1
            # generic: show everything (hour, minute, second, year, month, day)
            else:
                start_date_header = start_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')
                end_date_header = end_date_utc.astimezone(timezone_offset).strftime('%Y-%m-%d %H:%M:%S')

            f.write(f'Station:{station.code} - {station.name}\n')
            f.write(f'Data source:{data_source_description}\n')
            f.write(f'Description:{variable_names_string}\n')
            f.write(f'Latitude:{station.latitude}\n')
            f.write(f'Longitude:{station.longitude}\n')
            f.write(f'Passed AQC Checks: {current_datafile.aqc_checks}\n')
            f.write(f'Passed MQC Checks: {current_datafile.mqc_checks}\n')
            f.write(f'Date of completion:{date_of_completion.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}\n')
            f.write(f'Prepared by:{current_datafile.prepared_by}\n')
            f.write(f'Start date:{start_date_header}, End date:{end_date_header}\n\n')
            if displayUTC:
                f.write('Dates are displayed in UTC\n')
                f.write(f'Start date in UTC:{converted_start_date.strftime("%Y-%m-%d %H:%M:%S")}, End date in UTC:{converted_end_date.strftime("%Y-%m-%d %H:%M:%S")}\n\n')

            # Check the value of the agg to inform aggregation
            if agg == "min":
                # Write the text for "min" to the file
                f.write(f'Aggregation: Minimum\n\n')
            elif agg == "max":
                # Write the text for "max" to the file
                f.write(f'Aggregation: Maximum\n\n')
            elif agg == "sum":
                # Write the text for "sum" to the file
                f.write(f'Aggregation: Sum\n\n')
            elif agg == "avg":
                # Write the text for "avg" to the file
                f.write(f'Aggregation: Average\n\n')
            else:
                # Handle unexpected values of 'agg'
                f.write(f'Aggregation: Instantaneous Raw Data\n\n')


        lines = 0
        if query_result:
            df = pandas.DataFrame(data=query_result).pivot(index=0, columns=1)
            df.rename(columns=variable_dict, inplace=True)
            df.columns = df.columns.droplevel(0)
            if source == 'daily_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                cols = df.columns.tolist()
                cols = cols[-3:] + cols[:-3]
                df = df[cols]
                df = df.drop_duplicates(subset=['Day', 'Month', 'Year'], keep='first')
            elif source == 'monthly_summary':                
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                cols = df.columns.tolist()
                cols = cols[-2:] + cols[:-2]
                df = df[cols]
                df = df.drop_duplicates(subset=['Month', 'Year'], keep='first')
            elif source == 'yearly_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                cols = df.columns.tolist()
                cols = cols[-1:] + cols[:-1]
                df = df[cols]
                df = df.drop_duplicates(subset=['Year'], keep='first')
            else:
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                df['Time'] = df.index.map(lambda x: x.strftime('%H:%M:%S'))
                cols = df.columns.tolist()
                cols = cols[-4:] + cols[:-4]
                df = df[cols]


            # Apply each variable's configured precision before writing the CSV.
            #
            # At this point, the DataFrame columns have already been renamed from
            # variable IDs to variable symbols using variable_dict.
            #
            # Example:
            #   If Air Temperature has symbol "AT" and precision=4,
            #   then values in df["AT"] will be exported like:
            #       25.3000
            #       25.3457
            #
            # Code variables are skipped because their values are not normal numeric
            # measured values.
            for variable_symbol, precision in variable_precision_dict.items():
                if variable_symbol not in df.columns:
                    continue

                variable_type = variable_type_dict.get(variable_symbol)

                if variable_type and variable_type.lower() == "code":
                    continue

                df[variable_symbol] = df[variable_symbol].map(
                    lambda value: format_export_value_by_precision(value, precision)
                )

            df.to_csv(filepath, index=False, mode='a', header=True)


            lines = len(df.index)

        current_datafile.ready = True
        current_datafile.ready_at = date_of_completion
        current_datafile.lines = lines
        current_datafile.save()
        logger.info(f'Data exported successfully (file "{file_id}")')

        # create a new .xlsx version of the file
        convert_csv_xlsx(filepath, station.name, file_id)


    except Exception as e:
        current_datafile.ready = False
        current_datafile.ready_at = datetime.utcnow()
        current_datafile.lines = 0
        current_datafile.save()
        logger.error(f'Error on export data file "{file_id}". {repr(e)}')


# converts csv files to .xlsx format
def convert_csv_xlsx(file_path, file_station, file_id):
    if os.path.exists(file_path) and file_station:
        # open the csv and read all lines
        with open(file_path, 'r') as file:
            lines = file.readlines()

        try:
            # represents the workbook
            wb = Workbook()
            # currently active default worksheet in the workbook
            ws = wb.active
            # Naming the worksheet
            ws.title = f'{file_station}'

            # Write metadata
            for i, line in enumerate(lines):
                # Stop adding metadata once the data table starts
                if "Year" in line:
                    start_data_index = i + 1
                    break
                # writing the metadata to a single column
                ws.cell(row=i + 1, column=1, value=line.strip())

            # Write tabular data headers and rows
            headers = lines[start_data_index - 1].strip().split(",")
            ws.append(headers)  # Add headers

            # Add the tabular data stating from start_data_index till the end of the list
            for line in lines[start_data_index:]:
                row_data = line.strip().split(",")
                ws.append(row_data)
        except Exception as e:
            logger.error(f'An error occured while attempting to convert {file_id}.csv to .xlsx. Error Output: {e}')

        try:
            # Save the workbook
            wb.save(f'{settings.EXPORTED_DATA_CELERY_PATH}{file_id}.xlsx')  
        except Exception as e:
            logger.error(f'An error occured while attempting to save the {file_id}.xlsx file. Error Output: {e}')


# combine .xlsx files into a single .xlsx file
@shared_task
def combine_xlsx_files(station_ids, data_source, start_date, end_date, variable_ids, aggregation, displayUTC, data_interval_seconds, prepared_by, data_source_description, aqc_checks, mqc_checks, entry_id):
    # grab the current combine xlsx data file entry
    current_datafile = CombineDataFile.objects.get(pk=entry_id)

    try:

        # list containing the data frame of each station
        station_data_frames = []

        for station_id in station_ids:
            # a list to hold the dataframes containing the data queried from the db
            export_data_df = []

            # station dataframe
            station_df = pandas.DataFrame()

            if aggregation:
                for agg in aggregation:

                    export_data_df.append(export_data_xlsx(station_id, data_source, start_date, end_date, variable_ids, agg, aqc_checks, mqc_checks, displayUTC, data_interval_seconds))
            else:
                export_data_df.append(export_data_xlsx(station_id, data_source, start_date, end_date, variable_ids, aggregation, aqc_checks, mqc_checks, displayUTC, data_interval_seconds))

            # Getting the columns in common between dataframs pertaining to a specific station
            if data_source in ['raw_data', 'hourly_summary']:
                # selecting the year, month, day, time columns which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :4]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 4:]], axis=1)

            elif data_source == 'daily_summary':
                # selecting the year, month, day columns which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :3]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 3:]], axis=1)

            elif data_source == 'monthly_summary':
                # selecting the year, month columns which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :2]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 2:]], axis=1)

            elif data_source == 'yearly_summary':
                # selecting the year column which all dataframe will have in common given the data source
                station_df = export_data_df[0].iloc[:, :1]

                # loop through the output dataframes and concat the columns containing data to the station dataframe
                for df in export_data_df:
                    station_df = pandas.concat([station_df, df.iloc[:, 1:]], axis=1)

            else:
                raise ValueError("A valid data source was not passed!")


            # add the station dataframe to a list
            station_data_frames.append(station_df)


        # Create a new workbook for the combined file
        combined_workbook = Workbook()
        combined_workbook.remove(combined_workbook.active)  # Remove default sheet

        # the name of the output file will be the celery task id
        output_file = f'{settings.EXPORTED_DATA_CELERY_PATH}combine-{entry_id}.xlsx'

        # retrieving the variable names given then variable ids and converting the names into a string
        variable_dict = {}
        variable_names_string = ''
        try: 
            with connection.cursor() as cursor_variable:
                cursor_variable.execute(f'''
                    SELECT var.symbol
                        ,var.id
                        ,CASE WHEN unit.symbol IS NOT NULL THEN CONCAT(var.symbol, ' - ', var.name, ' (', unit.symbol, ')') 
                            ELSE CONCAT(var.symbol, ' - ', var.name) END as var_name
                    FROM wx_variable var 
                    LEFT JOIN wx_unit unit ON var.unit_id = unit.id 
                    WHERE var.id IN %s
                    ORDER BY var.name
                ''', (tuple(variable_ids),))

                rows = cursor_variable.fetchall()
                for row in rows:
                    variable_dict[row[1]] = row[0]
                    variable_names_string += f'{row[2]}   '
        except Exception as e:
            logger.error(f"AN ERROR OCCURED WITH THE VARIABLE NAMES: {e}")

        # looping through the dataframes of each station
        for x, id in enumerate(station_ids):
            # get the station object
            station = Station.objects.get(pk=id)

            # the number of entries in the data
            lines = len(station_data_frames[x].index)

            # Create a new sheet with the station name
            sheet = combined_workbook.create_sheet(title=f"{station.name} - {station.code}")

            # Write DataFrame to the sheet
            for r_idx, row in enumerate(dataframe_to_rows(station_data_frames[x], index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            # insert a new row at the top of the sheet to add the aggregation options and headers
            sheet.insert_rows(1, 12) # insert two rows at row 1

            # raw data does not have any aggregation options
            if data_source != 'raw_data':
                # Get the number of columns in the DataFrame
                num_cols = station_data_frames[x].shape[1]

                # get the amount of cells to merge
                num_merge = len(variable_ids) - 1

                # retrieve how many times a merger should occur
                qty_merge = len(aggregation) - 1

                start_col = num_cols - num_merge # Second-to-last column index (1-based index)
                end_col = num_cols # Last column index

                # Define a bold border style
                bold_border = Border(
                    left=Side(style="thick"),
                    right=Side(style="thick"),
                    top=Side(style="thick"),
                    bottom=Side(style="thick")
                )

                for x in range(qty_merge,-1,-1):
                    # Merge
                    sheet.merge_cells(start_row=12, start_column=start_col, end_row=12, end_column=end_col)
                    cell = sheet.cell(row=12, column=start_col, value=aggregation[x].upper())
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Apply bold border to all merged cells
                    for row in range(12, 13):  # Only row 12
                        for col in range(start_col, end_col + 1):
                            sheet.cell(row=row, column=col).border = bold_border

                    end_col = start_col - 1 # updated end column index
                    start_col = end_col - num_merge  # updated start column index

            date_of_completion = datetime.utcnow()
            timezone_offset = pytz.timezone(settings.TIMEZONE_NAME)

            # add the file headers
            cell = sheet.cell(row=1, column=1, value=f'Station:{station.code} - {station.name}')
            cell = sheet.cell(row=2, column=1, value=f'Data source:{data_source_description}')
            cell = sheet.cell(row=3, column=1, value=f'Description:{variable_names_string}')
            cell = sheet.cell(row=4, column=1, value=f'Latitude:{station.latitude}')
            cell = sheet.cell(row=5, column=1, value=f'Longitude:{station.longitude}')
            cell = sheet.cell(row=6, column=1, value=f'Passed AQC Checks:{aqc_checks}')
            cell = sheet.cell(row=7, column=1, value=f'Passed MQC Checks:{mqc_checks}')
            cell = sheet.cell(row=8, column=1, value=f'Date of completion:{date_of_completion.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}')
            cell = sheet.cell(row=9, column=1, value=f'Prepared by:{prepared_by}')

            if displayUTC and data_source in ['raw_data','hourly_summary']:
                cell = sheet.cell(row=11, column=1, value=f'Dates are displayed in UTC')
                cell = sheet.cell(row=12, column=1, value=f'Start date:{start_date}, End date:{end_date}')
            else:
                updated_start_date = pytz.UTC.localize(datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'))
                updated_end_date = pytz.UTC.localize(datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S'))
                cell = sheet.cell(row=11, column=1, value=f'Start date:{updated_start_date.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}, End date:{updated_end_date.astimezone(timezone_offset).strftime("%Y-%m-%d %H:%M:%S")}')

        # Save the workbook
        combined_workbook.save(output_file)

        current_datafile.ready = True
        current_datafile.ready_at = date_of_completion
        current_datafile.lines = lines
        current_datafile.save()
        logger.info(f'Combine Data exported successfully (file "combine-{entry_id}")')
    

    except Exception as e:
        current_datafile.ready = False
        current_datafile.ready_at = datetime.utcnow()
        current_datafile.lines = 0
        current_datafile.save()
        logger.error(f'Error on export combine xlsx data file "combine-{entry_id}". Error -  {repr(e)}')      


# returns the data frame for each query ran in order to facilitate combining the files later
def export_data_xlsx(station_id, source, start_date, end_date, variable_ids, agg, aqc_checks, mqc_checks, displayUTC, data_interval_seconds):

    timezone_offset = pytz.timezone(settings.TIMEZONE_NAME)
    start_date_utc = pytz.UTC.localize(datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S'))
    end_date_utc = pytz.UTC.localize(datetime.strptime(end_date, '%Y-%m-%d %H:%M:%S'))

    station = Station.objects.get(pk=station_id)

    variable_ids = tuple(variable_ids)
    # variable_ids = ','.join([str(x) for x in variable_ids])

    # Diferent data sources have diferents columns names for the measurement data and diferents intervals
    if source == 'raw_data':
        datetime_variable = 'datetime'
        data_source_description = 'Raw data'
        converted_start_date = start_date_utc
        converted_end_date = end_date_utc

    else:
        # measured_source = '''
        #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value
        #          WHEN var.sampling_operation_id = 3      THEN data.min_value
        #          WHEN var.sampling_operation_id = 4      THEN data.max_value
        #          WHEN var.sampling_operation_id = 6      THEN data.sum_value
        #     ELSE data.sum_value END as value '''
        if source == 'hourly_summary':
            datetime_variable = 'datetime'
            date_source = f"(datetime + interval '{station.utc_offset_minutes} minutes') at time zone 'utc' as date"
            data_source_description = 'Hourly summary'
            converted_start_date = start_date_utc
            converted_end_date = end_date_utc

        elif source == 'daily_summary':
            datetime_variable = 'day'
            data_source_description = 'Daily summary'
            date_source = "day::date"
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        elif source == 'monthly_summary':
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
            #         WHEN var.sampling_operation_id = 3      THEN data.min_value
            #         WHEN var.sampling_operation_id = 4      THEN data.max_value
            #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            datetime_variable = 'date'
            date_source = "date::date"
            data_source_description = 'Monthly summary'
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

        elif source == 'yearly_summary':
            # measured_source = '''
            #     CASE WHEN var.sampling_operation_id in (1,2) THEN data.avg_value::real
            #         WHEN var.sampling_operation_id = 3      THEN data.min_value
            #         WHEN var.sampling_operation_id = 4      THEN data.max_value
            #         WHEN var.sampling_operation_id = 6      THEN data.sum_value
            #     ELSE data.sum_value END as value '''
            datetime_variable = 'date'
            date_source = "date::date"
            data_source_description = 'Yearly summary'
            converted_start_date = start_date_utc.astimezone(timezone_offset).date()
            converted_end_date = end_date_utc.astimezone(timezone_offset).date()

    try:
        variable_dict = {}
        variable_precision_dict = {}
        variable_type_dict = {}
        variable_names_string = ''

        with connection.cursor() as cursor_variable:
            cursor_variable.execute(f'''
                SELECT var.symbol
                    ,var.id
                    ,CASE WHEN unit.symbol IS NOT NULL THEN CONCAT(var.symbol, ' - ', var.name, ' (', unit.symbol, ')') 
                        ELSE CONCAT(var.symbol, ' - ', var.name) END as var_name
                    ,var.precision
                    ,var.variable_type
                FROM wx_variable var 
                LEFT JOIN wx_unit unit ON var.unit_id = unit.id 
                WHERE var.id IN %s
                ORDER BY var.name
            ''', (variable_ids,))

            rows = cursor_variable.fetchall()

            for row in rows:
                variable_id = row[1]
                variable_symbol = row[0]
                variable_precision = row[3]
                variable_type = row[4]

                # Used later by pandas rename:
                # DB variable ID -> exported CSV column symbol
                variable_dict[variable_id] = variable_symbol

                # Used later to format the exported numeric values
                # after the pivot has created one column per variable.
                variable_precision_dict[variable_symbol] = variable_precision
                variable_type_dict[variable_symbol] = variable_type

                variable_names_string += f'{row[2]}   '

        # Iterate over the start and end date day by day to split the queries
        datetime_list = [converted_start_date]
        current_datetime = converted_start_date
        while current_datetime < converted_end_date and (current_datetime + timedelta(days=1)) < converted_end_date:
            current_datetime = current_datetime + timedelta(days=1)
            datetime_list.append(current_datetime)
        datetime_list.append(converted_end_date)

        # adding an extra day to the list when the source is monthly_summary or daily_summary
        # This will cause the last month/day to be included in the output
        if source in ['monthly_summary', 'daily_summary']:
            datetime_list.append(converted_end_date + timedelta(days=1))

            # in the case of querying for the monthly_summary of only 1 month
            # strip the datetime_list of the last value so that there are no duplicates
            if len(datetime_list) == 3 and source == 'monthly_summary':
                datetime_list.pop()
        # adding an extra hour to the list when the souce is hourly_summary
        elif source == 'hourly_summary':
            datetime_list.append(converted_end_date + timedelta(hours=1))
        # adding extra seconds to the list when the souce is raw_data
        # the number of seconds added is based on the data interval selected
        elif source == 'raw_data':
            datetime_list.append(converted_end_date + timedelta(seconds=data_interval_seconds))

        # quality control clause based on users request
        quality_clause = ""

        # if user request aqc data
        if aqc_checks and mqc_checks:
            quality_clause = """
                AND data.quality_flag = 4
                AND data.manual_flag = 4
            """
        elif aqc_checks:
            quality_clause = """
                AND data.quality_flag = 4
            """
        elif mqc_checks:
            quality_clause = """
                AND data.manual_flag = 4
            """


        query_result = []
        for i in range(0, len(datetime_list) - 1):
            current_start_datetime = datetime_list[i]
            current_end_datetime = datetime_list[i + 1]

            with connection.cursor() as cursor:
                if source == 'raw_data': 

                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_raw_data = f'''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                                {quality_clause}
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_raw_data = f'''
                            WITH processed_data AS (
                                SELECT datetime
                                    ,var.id as variable_id
                                    ,COALESCE(CASE WHEN var.variable_type ilike 'code' THEN data.code ELSE data.measured::varchar END, '-99.9') AS value
                                FROM raw_data data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                                {quality_clause}
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')
                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '%(data_interval)s seconds') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': data_interval_seconds})


                    cursor.execute(query_raw_data, {'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'data_interval': data_interval_seconds})

                elif source == 'hourly_summary':
                    
                    # removing the offset addition from the query based on the truthines of displayUTC. Removed `+ interval '%(utc_offset)s minutes'`
                    if displayUTC:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time) at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    # keeping the offset addition in the query based on the truthines of displayUTC.
                    else:
                        query_hourly = '''
                            WITH processed_data AS (
                                SELECT datetime ,var.id as variable_id
                                ,COALESCE(
                                    CASE
                                        WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                        WHEN %(aggregation)s = 'min'      THEN data.min_value
                                        WHEN %(aggregation)s = 'max'      THEN data.max_value
                                        WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                        ELSE data.avg_value 
                                    END, '-99.9'
                                ) as value  
                                FROM hourly_summary data
                                JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                                WHERE data.datetime >= %(start_datetime)s
                                AND data.datetime <= %(end_datetime)s
                                AND data.station_id = %(station_id)s
                            )
                            SELECT (generated_time + interval '%(utc_offset)s minutes') at time zone 'utc' as datetime
                                ,variable.id
                                ,COALESCE(value, '-99.9')

                            FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 hours') generated_time
                            JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                            LEFT JOIN processed_data ON datetime = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_hourly,{'utc_offset': station.utc_offset_minutes, 'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime, 
                          'station_id': station_id, 'aggregation': agg})
                    
                elif source == 'daily_summary':
                    query_daily = '''
                        WITH processed_data AS (
                            SELECT day ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM daily_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.day >= %(start_datetime)s
                            AND data.day <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s - INTERVAL '1 seconds', INTERVAL '1 days') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON day = generated_time AND variable.id = variable_id
                    '''

                    logging.info(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_daily, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                elif source == 'monthly_summary':
                    query_monthly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM monthly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date <= %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 months') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        '''
                    
                    logging.info(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_monthly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                elif source == 'yearly_summary':
                    query_yearly = '''
                        WITH processed_data AS (
                            SELECT date ,var.id as variable_id
                            ,COALESCE(
                                CASE
                                    WHEN %(aggregation)s = 'avg'      THEN data.avg_value::real 
                                    WHEN %(aggregation)s = 'min'      THEN data.min_value
                                    WHEN %(aggregation)s = 'max'      THEN data.max_value
                                    WHEN %(aggregation)s = 'sum'      THEN data.sum_value
                                    ELSE data.avg_value 
                                END, '-99.9'
                            ) as value  
                            FROM yearly_summary data
                            JOIN wx_variable var ON data.variable_id = var.id AND var.id IN %(variable_ids)s
                            WHERE data.date >= %(start_datetime)s
                            AND data.date < %(end_datetime)s
                            AND data.station_id = %(station_id)s
                        )
                        SELECT (generated_time) as datetime
                            ,variable.id
                            ,COALESCE(value, '-99.9')
                        FROM generate_series(%(start_datetime)s, %(end_datetime)s, INTERVAL '1 years') generated_time
                        JOIN wx_variable variable ON variable.id IN %(variable_ids)s
                        LEFT JOIN processed_data ON date = generated_time AND variable.id = variable_id
                        ''' 

                    logging.info(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})

                    cursor.execute(query_yearly, {'variable_ids': variable_ids,
                          'start_datetime': current_start_datetime, 'end_datetime': current_end_datetime,
                          'station_id': station_id, 'aggregation': agg})
                
                query_result.extend(cursor.fetchall())


        if query_result:
            
            df = pandas.DataFrame(data=query_result).pivot(index=0, columns=1)
            df.rename(columns=variable_dict, inplace=True)
            df.columns = df.columns.droplevel(0)
            if source == 'daily_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                cols = df.columns.tolist()
                cols = cols[-3:] + cols[:-3]
                df = df[cols]
                df = df.drop_duplicates(subset=['Day', 'Month', 'Year'], keep='first')
            elif source == 'monthly_summary':                
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                cols = df.columns.tolist()
                cols = cols[-2:] + cols[:-2]
                df = df[cols]
                df = df.drop_duplicates(subset=['Month', 'Year'], keep='first')
            elif source == 'yearly_summary':
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                cols = df.columns.tolist()
                cols = cols[-1:] + cols[:-1]
                df = df[cols]
                df = df.drop_duplicates(subset=['Year'], keep='first')
            else:
                df['Year'] = df.index.map(lambda x: x.strftime('%Y'))
                df['Month'] = df.index.map(lambda x: x.strftime('%m'))
                df['Day'] = df.index.map(lambda x: x.strftime('%d'))
                df['Time'] = df.index.map(lambda x: x.strftime('%H:%M:%S'))
                cols = df.columns.tolist()
                cols = cols[-4:] + cols[:-4]
                df = df[cols]

            # Apply each variable's configured precision before returning the DataFrame.
            #
            # At this point:
            #   - df has already been pivoted
            #   - variable ID columns have already been renamed to variable symbols
            #   - Year/Month/Day/Time columns have already been added
            #
            # Example:
            #   If Air Temperature has symbol "AT" and precision=4,
            #   then df["AT"] will contain values like:
            #       25.3000
            #       25.3457
            #
            # Code variables are skipped because they are not normal numeric measured values.
            for variable_symbol, precision in variable_precision_dict.items():
                if variable_symbol not in df.columns:
                    continue

                variable_type = variable_type_dict.get(variable_symbol)

                if variable_type and variable_type.lower() == "code":
                    continue

                df[variable_symbol] = df[variable_symbol].map(
                    lambda value: format_export_value_by_precision(value, precision)
                )

            return df

    except Exception as e:
        logger.error(f'Error on export data as combine xlsx. Error - {repr(e)}')


############################################################
@shared_task
def ftp_ingest_standard_station_files():
    process_in_chunks = False
    hfreq_data = False   
    ftp_ingest_station_files(process_in_chunks, hfreq_data)


@shared_task
def ftp_ingest_staged_station_files():
    process_in_chunks = True
    hfreq_data = False    
    ftp_ingest_station_files(process_in_chunks, hfreq_data)


@shared_task
def ftp_ingest_highfrequency_station_files():
    process_in_chunks = False
    hfreq_data = True
    ftp_ingest_station_files(process_in_chunks, hfreq_data)


def create_station_data_file_error_from_ingestion(sfi, error_message):
    """
    Create a StationDataFile error row for an ingestion-level problem.

    This is used when the error happens before a real file is downloaded,
    such as an invalid cron_schedule on StationFileIngestion.
    """

    StationDataFile.objects.create(
        station=sfi.station,
        decoder=sfi.decoder,
        status_id=4,  # Error
        utc_offset_minutes=sfi.utc_offset_minutes,

        # No actual file exists because the error happened before FTP download.
        filepath="",
        file_hash="",
        file_size=0,

        observation=error_message[:1024],
        process_in_chunks=sfi.process_in_chunks,
        is_highfrequency_data=sfi.is_highfrequency_data,
        override_data_on_conflict=sfi.override_data_on_conflict,
    )


def ftp_ingest_station_files(process_in_chunks, highfrequency_data):
    """
        Get and process station data files via FTP protocol.

        Parameters:
            process_in_chunks (bool): flag to process staged station data files
            highfrequency_data (bool): flag to process high frequency station data files
    """

    dt = datetime.now()

    station_file_ingestions = StationFileIngestion.objects.filter(
        is_active=True,
        process_in_chunks=process_in_chunks,
        is_highfrequency_data=highfrequency_data,
    )

    valid_station_file_ingestions = []

    for s in station_file_ingestions:
        cron_schedule = (s.cron_schedule or "").strip()

        # Normalize weird spacing like tabs, newlines, double spaces, etc...
        cron_schedule = " ".join(cron_schedule.split())

        if not cron_schedule:
            error_message = (
                f"Invalid cron_schedule for StationFileIngestion id={s.id}. "
                f"The cron_schedule is blank."
            )

            logger.warning(error_message)
            db_logger.error(error_message)

            create_station_data_file_error_from_ingestion(
                sfi=s,
                error_message=error_message,
            )

            continue

        try:
            should_run = cronex.CronExpression(cron_schedule).check_trigger(
                (dt.year, dt.month, dt.day, dt.hour, dt.minute)
            )

        except Exception as e:
            error_message = (
                f"Invalid cron_schedule for StationFileIngestion id={s.id}. "
                f"cron_schedule={cron_schedule!r}. "
                f"Error: {repr(e)}"
            )

            logger.exception(error_message)
            db_logger.error(error_message)

            create_station_data_file_error_from_ingestion(
                sfi=s,
                error_message=error_message,
            )

            continue

        if should_run:
            valid_station_file_ingestions.append(s)

    station_file_ingestions = valid_station_file_ingestions

    logger.info(f"Station File Ingestions: {station_file_ingestions}")

    # No ingestion configs are scheduled for this run.
    # Since this task only processes files after scheduled FTP ingestion,
    # exit early and do not call process_station_data_files().
    if not station_file_ingestions:
        message = (
            f"No active StationFileIngestion records are scheduled to run now. "
            f"process_in_chunks={process_in_chunks}, "
            f"highfrequency_data={highfrequency_data}"
        )

        logger.info(message)
        db_logger.info(message)

        return

    # List of unique FTP servers.
    ftp_servers = list(set([s.ftp_server for s in station_file_ingestions]))

    if highfrequency_data:
        data_type = 'hf_data'
    elif process_in_chunks:
        data_type = 'process_in_chunks'
    else:
        data_type = 'raw_data'

    # Loop over FTP servers, retrieving and saving files as StationDataFile records.
    for ftp_server in ftp_servers:
        logging.info(f'Connecting to {ftp_server}')

        try:
            with FTP(timeout=60) as ftp:
                ftp.connect(ftp_server.host, ftp_server.port, timeout=60)
                ftp.login(ftp_server.username, ftp_server.password)
                ftp.set_pasv(not ftp_server.is_active_mode)

                home_folder = ftp.pwd()

                for sfi in [s for s in station_file_ingestions if s.ftp_server == ftp_server]:

                    try:
                        ftp.cwd(sfi.remote_folder)

                    except error_perm as e:
                        error_message = (
                            f'Error accessing FTP directory "{sfi.remote_folder}" '
                            f'for StationFileIngestion id={sfi.id}. {repr(e)}'
                        )

                        logger.error(error_message)
                        db_logger.error(error_message)

                        create_station_data_file_error_from_ingestion(
                            sfi=sfi,
                            error_message=error_message,
                        )

                        # IMPORTANT:
                        # Do not continue using the previous FTP directory.
                        # Skip this ingestion config and move on.
                        continue


                    try:
                        # List remote files for this ingestion config.
                        remote_files = ftp.nlst(sfi.file_pattern)

                    except Exception as e:
                        error_message = (
                            f'Error listing FTP files in "{sfi.remote_folder}" '
                            f'with pattern "{sfi.file_pattern}" '
                            f'for StationFileIngestion id={sfi.id}. {repr(e)}'
                        )

                        logger.error(error_message)
                        db_logger.error(error_message)

                        create_station_data_file_error_from_ingestion(
                            sfi=sfi,
                            error_message=error_message,
                        )

                        try:
                            ftp.cwd(home_folder)
                        except Exception:
                            pass

                        continue


                    for fname in remote_files:
                        try:
                            local_folder = '/data/documents/ingest/%s/%s/%s/%04d/%02d/%02d' % (
                                sfi.decoder.name,
                                sfi.station.code,
                                data_type,
                                dt.year,
                                dt.month,
                                dt.day,
                            )

                            local_filename = '%04d%02d%02d%02d%02d%02d_%s' % (
                                dt.year,
                                dt.month,
                                dt.day,
                                dt.hour,
                                dt.minute,
                                dt.second,
                                fname,
                            )

                            local_path = '%s/%s' % (local_folder, local_filename)

                            os.makedirs(local_folder, exist_ok=True)

                            hash_md5 = hashlib.md5()

                            if sfi.is_binary_transfer:
                                with open(local_path, 'wb') as fp_binary:
                                    ftp.retrbinary(
                                        f'RETR {fname}',
                                        lambda data: [
                                            fp_binary.write(data),
                                            hash_md5.update(data),
                                        ],
                                    )
                            else:
                                with open(local_path, 'w') as fp:
                                    ftp.retrlines(
                                        f'RETR {fname}',
                                        lambda line: [
                                            fp.write(line + '\n'),
                                            hash_md5.update(line.encode('utf8')),
                                        ],
                                    )

                            if sfi.delete_from_server:
                                try:
                                    ftp.delete(fname)

                                except error_perm as e:
                                    error_message = (
                                        f'Permission error deleting FTP server file "{fname}". '
                                        f'Local path: "{local_path}". {repr(e)}'
                                    )

                                    logger.error(error_message)
                                    db_logger.error(error_message)

                                except error_reply as e:
                                    error_message = (
                                        f'Unknown FTP reply while deleting server file "{fname}". '
                                        f'Local path: "{local_path}". {repr(e)}'
                                    )

                                    logger.error(error_message)
                                    db_logger.error(error_message)

                            # Inserts a StationDataFile object with status_id=1, Not processed.
                            station_data_file = StationDataFile(
                                station=sfi.station,
                                decoder=sfi.decoder,
                                status_id=1,
                                utc_offset_minutes=sfi.utc_offset_minutes,
                                filepath=local_path,
                                file_hash=hash_md5.hexdigest(),
                                file_size=os.path.getsize(local_path), # getsize output in BYTES
                                process_in_chunks=sfi.process_in_chunks,
                                is_highfrequency_data=sfi.is_highfrequency_data,
                                override_data_on_conflict=sfi.override_data_on_conflict,
                            )

                            station_data_file.save()

                            logging.info(f'Downloaded FTP file: {local_path}')

                        except OSError as e:
                            error_message = f'OS error while saving FTP file "{fname}". {repr(e)}'

                            logger.error(error_message)
                            db_logger.error(error_message)

                            create_station_data_file_error_from_ingestion(
                                sfi=sfi,
                                error_message=error_message,
                            )

                        except Exception as e:
                            error_message = (
                                f'Unexpected error while downloading FTP file "{fname}" '
                                f'for StationFileIngestion id={sfi.id}. {repr(e)}'
                            )

                            logger.error(error_message)
                            db_logger.error(error_message)

                            create_station_data_file_error_from_ingestion(
                                sfi=sfi,
                                error_message=error_message,
                            )

                    try:
                        ftp.cwd(home_folder)
                    except Exception as e:
                        logger.error(f'Could not return to FTP home folder "{home_folder}". {repr(e)}')
                        db_logger.error(f'Could not return to FTP home folder "{home_folder}". {repr(e)}')

        except Exception as err:
            logger.error(
                f'Error during ftp_ingest_station_files. '
                f'Host: {ftp_server.host}    '
                f'Port: {ftp_server.port}    '
                f'Username: {ftp_server.username}'
            )
            logger.error(repr(err))

            db_logger.error(
                f'Error during ftp_ingest_station_files. '
                f'Host: {ftp_server.host}    '
                f'Port: {ftp_server.port}    '
                f'Username: {ftp_server.username}'
            )
            db_logger.error(repr(err))

    try:
        process_station_data_files(process_in_chunks, highfrequency_data)

        logger.info("FTP TASK: process_station_data_files returned successfully")

    except Exception as e:
        logger.exception(
            f"FTP TASK: process_station_data_files crashed: {repr(e)}"
        )

        db_logger.error(
            f"FTP TASK: process_station_data_files crashed: {repr(e)}"
        )


def process_station_data_files(process_in_chunks=False, highfrequency_data=False, force_reprocess=False):
    """
        Process station data files.

        Parameters:
            process_in_chunks (bool): flag to process staged files in chunks
            highfrequency_data (bool): flag to process high frequency files
            force_reprocess (bool): force file to be reprocessed;
                                    don't check if the file is already processed
    """

    logger.info(
        f"PROCESS TASK: entered process_station_data_files. "
        f"process_in_chunks={process_in_chunks}, "
        f"highfrequency_data={highfrequency_data}, "
        f"force_reprocess={force_reprocess}"
    )

    available_decoders = {
        'HOBO': read_file_hobo,
        'TOA5': read_file_toa5,
        'Weatherlink JSON': read_file_weatherlink_json,
        'HYDROLOGY': read_file_hydrology,
        'BELIZE MANUAL DAILY DATA': read_file_manual_data,
        'BELIZE MANUAL HOURLY DATA': read_file_manual_data_hourly,
        'SURFACE': read_file_surface,
        'F2000': read_file_f2000,
        'DAVIS': read_file_davis,
        'R FORMAT': read_file_R_format,
    }

    # Get StationDataFiles to process.
    # Only process files with:
    #   status_id=1 -> Not processed
    #   status_id=6 -> Reprocess
    #
    # IMPORTANT:
    # We only fetch the candidates here.
    # We DO NOT mark all 60 as "Being processed" up front anymore.
    # Each file is claimed one-by-one right before processing.
    station_data_file_list = list(
        StationDataFile.objects.select_related('decoder', 'station')
        .filter(
            status_id__in=(1, 6),
            process_in_chunks=process_in_chunks,
            is_highfrequency_data=highfrequency_data,
        )
        .order_by('id')[:60]
    )

    logger.info(
        f"PROCESS TASK: found {len(station_data_file_list)} station data files to process"
    )

    logger.info(
        f"PROCESS TASK: station_data_file_list={station_data_file_list}"
    )

    duplicate_file_hashes = set()
    seen_file_hashes_in_this_batch = set()

    # Get all duplicate file hashes into a set
    if not force_reprocess and station_data_file_list:
        batch_file_ids = [s.id for s in station_data_file_list]
        batch_file_hashes = [s.file_hash for s in station_data_file_list]

        duplicate_file_hashes = set(
            StationDataFile.objects
            .filter(file_hash__in=batch_file_hashes)
            .exclude(id__in=batch_file_ids)
            .exclude(status_id__in=(4, 5))  # exclude Error and Skipped
            .values_list("file_hash", flat=True)
        )

    for station_data_file in station_data_file_list:

        # If force_reprocess is false, skip duplicate files that were already processed
        # or duplicate files already seen earlier in this same batch.
        if not force_reprocess:
            duplicate_outside_batch = station_data_file.file_hash in duplicate_file_hashes
            duplicate_inside_batch = station_data_file.file_hash in seen_file_hashes_in_this_batch

            if duplicate_outside_batch or duplicate_inside_batch:
                StationDataFile.objects.filter(
                    id=station_data_file.id
                ).update(
                    status_id=5,  # Skipped
                    updated_at=now(),
                )
                continue

            seen_file_hashes_in_this_batch.add(station_data_file.file_hash)


        # Claim this one file for processing.
        #
        # This is safer than setting station_data_file.status_id = 2 and saving,
        # because this update only succeeds if the file is STILL status 1 or 6.
        #
        # If another worker already claimed it, updated_count will be 0,
        # and we skip it.
        updated_count = (
            StationDataFile.objects
            .filter(id=station_data_file.id, status_id__in=(1, 6))
            .update(status_id=2)  # Being processed
        )

        if updated_count == 0:
            logger.info(
                f'StationDataFile id={station_data_file.id} was already claimed by another worker. Skipping.'
            )
            continue

        # Keep the in-memory object consistent with the database. Set to being processed
        station_data_file.status_id = 2

        try:
            current_decoder = available_decoders.get(station_data_file.decoder.name)

            if current_decoder is None:
                raise Exception(f'No decoder configured for "{station_data_file.decoder.name}"')

            logger.info(
                'Processing file "{0}" with "{1}" decoder.'.format(
                    station_data_file.filepath,
                    current_decoder,
                )
            )

            current_decoder(
                filename=station_data_file.filepath,
                highfrequency_data=station_data_file.is_highfrequency_data,
                process_in_chunks=station_data_file.process_in_chunks,
                station_data_file=station_data_file,
                station_object=station_data_file.station,
                utc_offset=station_data_file.utc_offset_minutes,
                override_data_on_conflict=station_data_file.override_data_on_conflict,
            )

        except Exception as err:
            # Update status_id to 4, Error.
            station_data_file.status_id = 4

            station_data_file.observation = (
                f'Error Processing file with decoder "{station_data_file.decoder.name}". '
                + repr(err)
            )[:1024]

            station_data_file.save(update_fields=['status', 'observation', 'updated_at'])

            logger.error(
                f'Error Processing file "{station_data_file.filepath}" '
                f'with decoder "{station_data_file.decoder.name}". {repr(err)}'
            )

            db_logger.error(
                f'Error Processing file "{station_data_file.filepath}" '
                f'with decoder "{station_data_file.decoder.name}". {repr(err)}'
            )

        else:
            # Update status_id to 3, Processed.
            station_data_file.status_id = 3
            station_data_file.save(update_fields=['status', 'updated_at'])


############################################################


@shared_task
def disable_unregistered_periodic_tasks():
    """
    Disable enabled django-celery-beat PeriodicTask records whose task name
    is not registered in the current Celery app.

    Why disable instead of delete?
    ------------------------------
    Sometimes a task may appear unregistered because of a temporary import issue,
    stale container, deployment mismatch, or worker startup problem.

    Disabling is safer because:
    - Beat stops sending the invalid task.
    - The record stays in Django admin for review.
    - You can re-enable it later if needed.
    """

    registered_tasks = set(current_app.tasks.keys())

    stale_tasks = (
        PeriodicTask.objects
        .filter(enabled=True)
        .exclude(task__in=registered_tasks)
    )

    rows = list(
        stale_tasks.values(
            "id",
            "name",
            "task",
        )
    )

    stale_count = len(rows)

    if stale_count == 0:
        logger.info(
            "disable_unregistered_periodic_tasks: no unregistered periodic tasks found."
        )
        return 0

    stale_tasks.update(enabled=False)

    logger.warning(
        "disable_unregistered_periodic_tasks: disabled %s unregistered periodic task(s): %s",
        stale_count,
        rows,
    )

    db_logger.warning(
        "disable_unregistered_periodic_tasks: disabled %s unregistered periodic task(s): %s",
        stale_count,
        rows,
    )

    return stale_count


############################################################


@shared_task
def process_staged_raw_data_tasks():
    """
    Process staged raw-data tasks.

    This is used for large files that should not be inserted into raw_data
    all at once.

    The staged rows were already decoded and saved into StagedRawData.
    This task sends those staged rows into insert_raw_data.insert(...) in
    smaller batches so the normal QC process can run without putting too much
    pressure on the database or Celery worker.
    """

    # Number of StagedRawDataTask records to process per Celery execution.
    # This limits how much work this task can do in one run.
    TASK_LIMIT = 50

    staged_tasks = (
        StagedRawDataTask.objects
        .filter(started_at=None, finished_at=None)
        .select_related("station_data_file", "station", "variable")
        .order_by("id")[:TASK_LIMIT]
    )

    if not staged_tasks:
        logger.info("No staged raw-data tasks found to process.")
        return 0

    processed_count = 0

    for staged_task in staged_tasks:
        process_single_staged_raw_data_task(staged_task.id)
        processed_count += 1

    logger.info(
        "Processed %s staged raw-data task(s).",
        processed_count,
    )

    return processed_count


def process_single_staged_raw_data_task(staged_task_id):
    """
    Process one StagedRawDataTask.

    This function loads staged rows for the task's station, variable, and
    datetime range, then passes them to insert_raw_data.insert(...) in batches.

    QC is not done in the staging table.
    QC is done by insert_raw_data.insert(...), exactly like normal raw-data
    processing.
    """

    # Number of staged rows to send to insert_raw_data.insert(...) at once.
    # Start conservative. You can increase this later after checking CPU/load.
    BATCH_SIZE = 500

    staged_task = (
        StagedRawDataTask.objects
        .select_related("station_data_file", "station", "variable")
        .get(id=staged_task_id)
    )

    # Claim this task safely.
    # If another worker already started it, this update will affect 0 rows.
    updated_count = (
        StagedRawDataTask.objects
        .filter(id=staged_task.id, started_at=None, finished_at=None)
        .update(
            started_at=now(),
            error=None,
        )
    )

    if updated_count == 0:
        logger.info(
            "StagedRawDataTask id=%s was already claimed by another worker. Skipping.",
            staged_task.id,
        )
        return

    # This staged/chunked raw-data pipeline is only meant for Numeric variables.
    #
    # insert_raw_data only applies QC to Numeric variables. For non-numeric
    # variables, QC is skipped and quality_flag must already be valid.
    #
    # Since StagedRawData intentionally stores only the core decoded fields
    # needed for numeric QC, non-numeric tasks should not be processed here.
    #
    # The newer insert_staged_raw_data.py should avoid creating non-numeric
    # StagedRawDataTask records, but this check protects us from old tasks that
    # may already exist in the database.
    if staged_task.variable.variable_type != "Numeric":
        message = (
            f"Skipping StagedRawDataTask id={staged_task.id} because "
            f"variable_id={staged_task.variable_id} is not Numeric."
        )

        logger.warning(message)
        db_logger.warning(message)

        StagedRawDataTask.objects.filter(id=staged_task.id).update(
            finished_at=now(),
            error=None,
        )

        return

    try:
        logger.info(
            "Started StagedRawDataTask id=%s | file_id=%s | station_id=%s | "
            "variable_id=%s | range=(%s, %s)",
            staged_task.id,
            staged_task.station_data_file_id,
            staged_task.station_id,
            staged_task.variable_id,
            staged_task.start_datetime,
            staged_task.end_datetime,
        )

        staged_rows_qs = (
            StagedRawData.objects
            .filter(
                station_data_file_id=staged_task.station_data_file_id,
                station_id=staged_task.station_id,
                variable_id=staged_task.variable_id,
                datetime__gte=staged_task.start_datetime,
                datetime__lte=staged_task.end_datetime,
            )
            .order_by("datetime")
            .values_list(
                "station_id",
                "variable_id",
                "seconds",
                "datetime",
                "measured",
            )
        )

        batch = []
        total_rows_processed = 0

        for station_id, variable_id, seconds, data_datetime, measured in staged_rows_qs.iterator(chunk_size=BATCH_SIZE):
            # insert_raw_data.insert(...) expects rows in this shape:
            #
            # [
            #   station_id,
            #   variable_id,
            #   seconds,
            #   datetime,
            #   measured,
            #   quality_flag,
            #   qc_range_quality_flag,
            #   qc_range_description,
            #   qc_step_quality_flag,
            #   qc_step_description,
            #   qc_persist_quality_flag,
            #   qc_persist_description,
            #   manual_flag,
            #   consisted,
            #   is_daily,
            # ]
            #
            # For staged raw data, we are intentionally passing the decoded
            # measurement through to the normal raw-data insert path.
            #
            # QC fields are None because insert_raw_data.insert(...) should
            # calculate/apply them for Numeric variables.
            #
            # seconds is preserved from StagedRawData because insert_raw_data
            # groups by station_id + variable_id + seconds + month while
            # preparing QC context.
            batch.append([
                station_id,
                variable_id,
                seconds,
                data_datetime,
                measured,
                None,   # quality_flag
                None,   # qc_range_quality_flag
                None,   # qc_range_description
                None,   # qc_step_quality_flag
                None,   # qc_step_description
                None,   # qc_persist_quality_flag
                None,   # qc_persist_description
                None,   # manual_flag
                None,   # consisted
                False,  # is_daily
            ])

            if len(batch) >= BATCH_SIZE:
                insert_rd(
                    batch,
                    override_data_on_conflict=staged_task.station_data_file.override_data_on_conflict,
                )

                total_rows_processed += len(batch)
                batch = []

        # Insert any rows left over after the final full batch.
        if batch:
            insert_rd(
                batch,
                override_data_on_conflict=staged_task.station_data_file.override_data_on_conflict,
            )

            total_rows_processed += len(batch)

        StagedRawDataTask.objects.filter(id=staged_task.id).update(
            finished_at=now(),
            error=None,
        )

        logger.info(
            "Finished StagedRawDataTask id=%s | rows_processed=%s",
            staged_task.id,
            total_rows_processed,
        )

    except Exception as err:
        error_message = (
            f"Error processing StagedRawDataTask id={staged_task.id}. "
            f"station_id={staged_task.station_id}, "
            f"variable_id={staged_task.variable_id}, "
            f"range=({staged_task.start_datetime}, {staged_task.end_datetime}). "
            f"{repr(err)}"
        )

        logger.exception(error_message)
        db_logger.error(error_message)

        StagedRawDataTask.objects.filter(id=staged_task.id).update(
            error=error_message[:1024],
        )


############################################################


# modify manual station data file object with the Month and Station_List
@shared_task
def ingest_manual_station_files(data_file_id_list, selected_decoder="default"):
    """
    Process station data files for Month and Station Names.
    """

    try:

        for entry_id in data_file_id_list:
            manual_station_data_file = ManualStationDataFile.objects.get(id=entry_id)

            filepath = manual_station_data_file.filepath

            logger.info(f'Processing Manual Datafile: {filepath}')

            sheet_month = None

            if selected_decoder == 'default':
                source = pandas.ExcelFile(filepath)

                stations_list = '       '.join(source.sheet_names) # combining stations list into a long sting

                header_block = source.parse(source.sheet_names[0], header=None, nrows=5, na_filter=False)

                for v in header_block.to_numpy().ravel():
                    if isinstance(v, str) and "MONTH:" in v:
                        m = re.search(r"MONTH:\s*([A-Za-z]+)", v)
                        if m:
                            sheet_month = m.group(1)
                            break
            else:
                sheet_month = "---"
                stations_list = "---"

            
            # update the manual station data file object
            manual_station_data_file.stations_list = stations_list
            manual_station_data_file.month = sheet_month
            manual_station_data_file.save()

            # process manual station data files
            process_manual_station_data_files(data_file_id_list, selected_decoder, manual_station_data_file)

    except Exception as e:
        logger.error('OS error. ' + repr(e))
        db_logger.error('OS error. ' + repr(e))


# process manual station data files
def process_manual_station_data_files(data_file_id_list, selected_decoder="default", manual_station_data_file=None):
    """
    Process manual station data files

    Parameters: 
        data_file_id_list: list containing the id of the station data files to process
    """

    match_decoder = {
        'default': 'BELIZE MANUAL DAILY DATA',
        'davis': 'DAVIS',
        'f2000': 'F2000',
        'r_format': 'RFormat',
    }

    available_decoders = {
        'BELIZE MANUAL DAILY DATA': read_file_manual_data,
        'DAVIS': read_file_davis,
        'F2000': read_file_f2000,
        'RFormat': read_file_rformat,
    }

    # Get ManualStationDataFile to process
    # Filter status id to process only StationDataFiles with code 1 (Not processed) or 6 (Reprocess)
    station_data_file_list = ManualStationDataFile.objects.filter(id__in=data_file_id_list)

    # Mark all files as Being processed
    for station_data_file in station_data_file_list:
        # Update status id to 2 (Being processed)
        station_data_file.status_id = 2
        station_data_file.save()

        try:
            current_decoder = available_decoders[match_decoder[selected_decoder]] # match the correct decoder based on the selected_decoder
            logger.info('Processing file "{0}" with "{1}" decoder.'.format(station_data_file.filepath, current_decoder))

            if selected_decoder == "default":
                current_decoder(filename=station_data_file.filepath
                                , override_data_on_conflict=station_data_file.override_data_on_conflict
                                , is_manually_validated=station_data_file.is_manually_validated)
                
            elif selected_decoder == "davis":
                current_decoder(filename=station_data_file.filepath
                                , override_data_on_conflict=station_data_file.override_data_on_conflict
                                , origin_file_name=manual_station_data_file.file_name)
                
            else:
                current_decoder(filename=station_data_file.filepath
                                , override_data_on_conflict=station_data_file.override_data_on_conflict)

        except Exception as err:
            # Update status id to 4 (Error)
            station_data_file.status_id = 4
            station_data_file.observation = (err)
            station_data_file.save()

            logger.error('Error Processing file "{0}" with "{1}" decoder. '
                         .format(station_data_file.filepath, current_decoder) + repr(err))
            db_logger.error('Error Processing file "{0}" with "{1}" decoder. '
                            .format(station_data_file.filepath, current_decoder) + repr(err))
        else:
            # Update status id to 3 (Processed)
            station_data_file.status_id = 3
            station_data_file.save()

        try:
            # attempting to delete already process manual station data file regardless of status
            os.remove(station_data_file.filepath)
            logger.info(f'Manual station data file deleted after processing. File path: {station_data_file.filepath}')
        except Exception as e:
            logger.error('An error occured whill attempting to delete manual station data file.')
            logger.error(f'Error - {repr(e)}')


# Persist Logic Starts here
# Despite its name this fxn doesn't get only hourly_raw_data
# persist checks called during hourly AND daily summaries call this fxn with daily summaries retrieving data for the whole day.
def get_hourly_raw_data(start_datetime, end_datetime, station_ids):
    con = get_connection()

    # note that in the query bellow the 'qc_persist_quality_flag IS NULL' condition IS COMMENTED OUT!
    sql = '''SELECT *
             FROM raw_data
             WHERE datetime BETWEEN %(start_datetime)s AND %(end_datetime)s
               AND station_id = ANY(%(station_ids)s) 
               -- AND qc_persist_quality_flag IS NULL
             ORDER BY datetime DESC
          '''
    params = {"station_ids": station_ids, "start_datetime": start_datetime, "end_datetime": end_datetime}
    

    # sql_query = pd.read_sql_query(sql=sql, con=con, params=params)
    # con.close()

    # df = pd.DataFrame(sql_query)

    df = pd.read_sql_query(sql=sql, con=ENGINE, params=params)

    return df


# Build a dictionary of stations and their variables with observed datetime ranges
def get_hourly_sv_dict(start_datetime, end_datetime, station_ids):

    # Retrieve raw_data for the given time range and station IDs
    df = get_hourly_raw_data(start_datetime, end_datetime, station_ids)

    dict_sv = {}

    # Loop through each station
    # Note: stations with no data in the time range will map to an empty dict
    for station_id in station_ids:

        # Get unique variable IDs observed at this station
        variable_ids = df[df['station_id']==station_id].variable_id.unique()

        dict_v = {}

        # For each variable, determine the earliest and latest datetime observed
        for variable_id in variable_ids:

            s_datetime = df[(df['station_id']==station_id) & (df['variable_id']==variable_id)].datetime.min()

            e_datetime = df[(df['station_id']==station_id) & (df['variable_id']==variable_id)].datetime.max()

            dict_v[variable_id] = (s_datetime, e_datetime)

        # Assign variable ranges to the station
        dict_sv[station_id] = dict_v

    # Resulting format of dict_sv:
    # {
    #     station_id: {
    #         variable_id: (start_datetime, end_datetime)
    #     }
    # }

    return dict_sv



def get_hourly_sv_data(start_datetime, end_datetime, station_id, variable_id, window):
    sql = '''
        SELECT *
        FROM raw_data
        WHERE (datetime BETWEEN %(start_datetime)s AND %(end_datetime)s)
          AND station_id = %(station_id)s
          AND variable_id = %(variable_id)s
    '''

    params = {
        "station_id": station_id,
        "variable_id": int(variable_id),
        "start_datetime": start_datetime - timedelta(seconds=window),
        "end_datetime": end_datetime,
    }

    df = pd.read_sql_query(sql=sql, con=ENGINE, params=params)

    # sql_query = pd.read_sql_query(sql=sql, con=con, params=params)
    # con.close()

    # df = pd.DataFrame(sql_query)

    if not df.empty:
        mask = df.datetime.between(start_datetime, end_datetime)
        df["updated"] = False
        df.loc[mask, "updated"] = True

    return df



def most_frequent(List):
    return max(set(List), key = List.count)

# Interval
# This function computes the most common time difference (in seconds) between consecutive datetime 
# entries in a DataFrame, effectively detecting the dataset’s sampling interval.
# If each interval occurs exactly once, it returns the first one it encounters.
def get_interval(df):
    interval_list = df.datetime.diff(periods=1).dt.total_seconds().replace(np.nan, 0).astype("int32")
    interval_list = list(interval_list)
    interval = most_frequent(interval_list)
    return abs(interval)

# Window
def get_window(station_id, variable_id):

    # default window of 3600 seconds (1 hour) if the station/variable doesn't exist
    try:
        _station = Station.objects.get(id=station_id)
        _variable = Variable.objects.get(id=variable_id) 
    except ObjectDoesNotExist:
        return 3600
        
    if  _station.is_automatic:
        hours = _variable.persistence_window_hourly

        if hours is None:
            hours = 1 # 1 Hour
    else:
        hours = _variable.persistence_window

        if hours is None:
            hours = 96 # 4 Days

    return hours*3600

# Thresholds
def get_thresholds(station_id, variable_id, interval, window):
    thresholds = {}
    if window == 0 or interval==0:
        return thresholds

    try:
        # Trying to set persist thresholds using current station
        _persist = QcPersistThreshold.objects.get(station_id=station_id, variable_id=variable_id, interval=interval, window=window)
        thresholds['persist_min'] = _persist.minimum_variance
        thresholds['persist_des'] = 'Custom station Threshold'
    except ObjectDoesNotExist:
        try:
            # Trying to set persist thresholds using current station with NULL intervall
            _range = QcPersistThreshold.objects.get(station_id=station_id, variable_id=variable_id, interval__isnull=True, window=window)        
            thresholds['persist_min'] = _persist.minimum_variance
            thresholds['persist_des'] = 'Custom station Threshold'
        except ObjectDoesNotExist:
            try:
                # Trying to set persist thresholds using current station
                _station = Station.objects.get(pk=station_id)
                _persist = QcPersistThreshold.objects.get(station_id=_station.reference_station_id, variable_id=variable_id, interval=interval, window=window)
                thresholds['persist_min'] = _persist.minimum_variance
                thresholds['persist_des'] = 'Reference station threshold'
            except ObjectDoesNotExist:
                try:
                    # Trying to set persist thresholds using current station with NULL intervall
                    _station = Station.objects.get(pk=station_id)
                    _persist = QcPersistThreshold.objects.get(station_id=_station.reference_station_id, variable_id=variable_id, interval__isnull=True, window=window)        
                    thresholds['persist_min'] = _persist.minimum_variance
                    thresholds['persist_des'] = 'Reference station threshold'
                except ObjectDoesNotExist:
                    try:
                        # Trying to set persistence thresholds using global ranges
                        _station = Station.objects.get(pk=station_id)
                        _persist = Variable.objects.get(pk=variable_id)
                        if _station.is_automatic:
                            thresholds['persist_min'] = _persist.persistence_hourly
                            thresholds['persist_des'] = 'Global threshold (Automatic)'
                        else:
                            thresholds['persist_min'] = _persist.persistence
                            thresholds['persist_des'] = 'Global threshold (Manual)'
                    except ObjectDoesNotExist:
                        pass;                    
    return thresholds

# Persistance function and calculation
def persit_function(values):
    return abs(max(values)-min(values))

def get_persist(row, df, window):
    datetime = row.datetime

    s_datetime = datetime-timedelta(seconds=window)
    e_datetime = datetime

    mask = df.datetime.between(s_datetime, e_datetime)
    List = list(df[mask]['measured'])

    persist = persit_function(List)
    return persist

def qc_persist(value, thresholds):
    if 'persist_min' not in thresholds:
        return NOT_CHECKED, "Threshold not found"

    p_min = thresholds['persist_min']
    p_des = thresholds['persist_des']
           
    if p_min is None:
        return NOT_CHECKED, "Threshold not found"           

    if value >= p_min:
        return GOOD, p_des
    else:
        return BAD, p_des    

def set_persist_sus(row, df, window, persist_flag):
    updated = row.updated
    datetime = row.datetime

    if persist_flag in [BAD, SUSPICIOUS]:
        return persist_flag, updated

    s_datetime = datetime
    e_datetime = datetime+timedelta(seconds=window)

    mask = df.datetime.between(s_datetime, e_datetime)
    List = list(df[mask]['qc_persist_quality_flag'])

    if BAD in List:
        return SUSPICIOUS, True
    
    return persist_flag, updated

def qc_final(row, persist_flag):
    range_flag = row.qc_range_quality_flag
    step_flag = row.qc_step_quality_flag

    flags = (persist_flag, range_flag, step_flag)
    if BAD in flags:
        return BAD
    elif GOOD in flags:
        return GOOD
    elif SUSPICIOUS in flags:
        return SUSPICIOUS        
    return NOT_CHECKED

def set_persist(row, df, s_datetime, e_datetime, interval, window, thresholds):
    if s_datetime <= row.datetime <= e_datetime:
        persist = get_persist(row, df, window)
        persist_flag, persist_des = qc_persist(persist, thresholds)

    else:
        persist_flag, persist_des = row.qc_persist_quality_flag, row.qc_persist_description

    # By the pandas apply function order, the previous rows are calculated so we can set suspicous flags using them
    persist_flag, updated = set_persist_sus(row, df, window, persist_flag)

    # Finally compute the final flag using all three flags, range, step, and persistence
    final_flag =  qc_final(row, persist_flag)

    return updated, persist_flag, persist_des, final_flag

# Persistance update
def update_insert_persist(df):
    data = df.to_dict('records')
    query = '''
            INSERT INTO raw_data(created_at, updated_at, datetime, station_id, variable_id, measured, qc_persist_quality_flag, qc_persist_description, quality_flag)
                VALUES (CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %(datetime)s, %(station_id)s, %(variable_id)s, 0, %(qc_persist_quality_flag)s, %(qc_persist_description)s, %(quality_flag)s)
            ON CONFLICT (datetime, station_id, variable_id) DO UPDATE SET
                updated_at=CURRENT_TIMESTAMP,
                qc_persist_quality_flag = %(qc_persist_quality_flag)s,
                qc_persist_description = %(qc_persist_description)s,
                quality_flag = %(quality_flag)s
            '''
    con = get_connection()            
    with con.cursor() as cursor:
        cursor.executemany(query, data)
    con.commit()
    con.close()

# Main Persist Function
def update_qc_persist(start_datetime, end_datetime, station_ids, summary_type):

    # Dictionary of stations and their variables with observed datetime ranges
    dict_sv = get_hourly_sv_dict(start_datetime, end_datetime, station_ids)

    for station_id in dict_sv:

        for variable_id in dict_sv[station_id]:
            # grab datetime ranges
            s_datetime, e_datetime = dict_sv[station_id][variable_id]

            # Get the window from persistence window hourly or persistence window from the variable
            window = get_window(station_id, variable_id)

            df = get_hourly_sv_data(s_datetime, e_datetime, station_id, variable_id, window)
            
            if not df.empty:
                interval = get_interval(df)

                thresholds = get_thresholds(station_id, variable_id, interval, window)   
                
                columns = ['updated', 'qc_persist_quality_flag', 'qc_persist_description', 'quality_flag']
                
                df[columns] = df.apply(
                    lambda row: set_persist(row, df, s_datetime, e_datetime, interval, window, thresholds), 
                    axis=1, 
                    result_type="expand")

                df = df[df['updated']==True]                

                columns = ['station_id', 'variable_id', 'qc_persist_quality_flag', 'qc_persist_description', 'quality_flag', 'datetime']

                update_insert_persist(df[columns])

                recalculate_summary(df, station_id, s_datetime, e_datetime, summary_type)

def insert_summay(date, station_id, summary_type):
    query_hourly = '''
                    INSERT INTO wx_hourlysummarytask(created_at, updated_at, datetime, started_at, finished_at, station_id)
                       VALUES (CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %(date)s, NULL, NULL, %(station_id)s)
                   '''
    query_daily = '''
                    INSERT INTO wx_dailysummarytask(created_at, updated_at, date, started_at, finished_at, station_id)
                       VALUES (CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, %(date)s, NULL, NULL, %(station_id)s)
                   '''

    if summary_type == 'hourly':
        query = query_hourly
    elif summary_type == 'daily':
        query = query_daily


    data = {'station_id': station_id, 'date': date}
    con = get_connection()            
    with con.cursor() as cursor:
        cursor.execute(query, data)
    con.commit()
    con.close()

def recalculate_summary(df, station_id, s_datetime, e_datetime, summary_type):
    mask = df.datetime.between(s_datetime, e_datetime)    
    datetimes = list(df[~mask]['datetime'])

    if datetimes:
        if summary_type == 'hourly':
            datetimes = set([dt.replace(microsecond=0, second=0, minute=0) for dt in datetimes])
            for datetime in datetimes:
                _hourlysummaries = HourlySummaryTask.objects.filter(station_id=station_id, datetime=datetime, started_at__isnull=True)
                if not _hourlysummaries:
                    insert_summay(datetime, station_id, summary_type)

        elif summary_type == 'daily':
            dates = set([dt.date() for dt in datetimes])
            for date in dates:
                _dailysummaries = DailySummaryTask.objects.filter(station_id=station_id, date=date, started_at__isnull=True)
                if not _dailysummaries:

                    insert_summay(date, station_id, summary_type)

def hourly_summary(hourly_summary_tasks_ids, station_ids, s_datetime, e_datetime):
    try:
        HourlySummaryTask.objects.filter(id__in=hourly_summary_tasks_ids).update(started_at=datetime.now(tz=pytz.UTC))
        calculate_hourly_summary(s_datetime, e_datetime, station_id_list=station_ids)
    except Exception as err:
        logger.error('Error calculation hourly summary for hour "{0}". '.format(s_datetime) + repr(err))
        db_logger.error('Error calculation hourly summary for hour "{0}". '.format(s_datetime) + repr(err))
    else:
        HourlySummaryTask.objects.filter(id__in=hourly_summary_tasks_ids).update(finished_at=datetime.now(tz=pytz.UTC))

@shared_task
def process_hourly_summary_tasks():
    # Process only 501 hourly summaries per execution
    unprocessed_hourly_summary_datetimes = HourlySummaryTask.objects.filter(started_at=None).values_list('datetime',flat=True).distinct()[:501]
    
    for hourly_summary_datetime in unprocessed_hourly_summary_datetimes:
        start_datetime = hourly_summary_datetime.replace(minute=0, second=0, microsecond=0)
        end_datetime = start_datetime + timedelta(hours=1)

        hourly_summary_tasks = HourlySummaryTask.objects.filter(started_at=None, datetime=hourly_summary_datetime)
        hourly_summary_tasks_ids = list(hourly_summary_tasks.values_list('id', flat=True))
        station_ids = list(hourly_summary_tasks.values_list('station_id', flat=True).distinct())

        # legacy way of calculating QC persist
        # if station_ids:
        #     update_qc_persist(start_datetime, end_datetime, station_ids, 'hourly')

        # Updating Hourly summary task
        hourly_summary(hourly_summary_tasks_ids, station_ids, start_datetime, end_datetime)
        

def daily_summary(daily_summary_tasks_ids, station_ids, s_datetime, e_datetime):
    try:
        # Mark tasks as started before calculation begins.
        DailySummaryTask.objects.filter(
            id__in=daily_summary_tasks_ids
        ).update(
            started_at=datetime.now(tz=pytz.UTC)
        )

        # Calculate/recalculate the daily summary rows.
        calculate_daily_summary(
            s_datetime,
            e_datetime,
            station_id_list=station_ids
        )

        # Calculate station minimum interval after the daily summary.
        for station_id in station_ids:
            calculate_station_minimum_interval(
                s_datetime,
                e_datetime,
                station_id_list=(station_id,)
            )

    except Exception as err:
        logger.exception(
            'Error calculation daily summary for day "%s".',
            s_datetime
        )

        db_logger.error(
            'Error calculation daily summary for day "{0}". {1}'.format(
                s_datetime,
                repr(err)
            )
        )

        # Re-raise so Celery marks the task as failed.
        raise

    else:
        # Only mark tasks as finished if everything completed successfully.
        DailySummaryTask.objects.filter(
            id__in=daily_summary_tasks_ids
        ).update(
            finished_at=datetime.now(tz=pytz.UTC)
        )



@shared_task
def process_daily_summary_tasks():
    try:
        # Process only 500 daily summary dates per execution.
        # The slice uses 501 in the original code, so I am preserving that behavior.
        unprocessed_daily_summary_dates = (
            DailySummaryTask.objects
            .filter(started_at=None)
            .values_list('date', flat=True)
            .distinct()[:501]
        )

        for daily_summary_date in unprocessed_daily_summary_dates:
            start_date = daily_summary_date
            end_date = daily_summary_date + timedelta(days=1)

            daily_summary_tasks = DailySummaryTask.objects.filter(
                started_at=None,
                date=daily_summary_date
            )

            daily_summary_tasks_ids = list(
                daily_summary_tasks.values_list('id', flat=True)
            )

            station_ids = list(
                daily_summary_tasks
                .values_list('station_id', flat=True)
                .distinct()
            )

            daily_summary(
                daily_summary_tasks_ids,
                station_ids,
                start_date,
                end_date
            )

    except Exception:
        logger.exception('Error processing daily summary.')
        raise



def predict_data(start_datetime, end_datetime, prediction_id, station_ids, target_station_id, variable_id,
                 data_period_in_minutes, interval_in_minutes, result_mapping):
    data_frequency = (interval_in_minutes // data_period_in_minutes) - 1
    date_dict = {}

    logger.info(
        f"predict_data= start_datetime: {start_datetime}, end_datetime: {end_datetime}, station_ids: {station_ids}, variable_id: {variable_id}, data_period_in_minutes: {data_period_in_minutes}, interval_in_minutes: {interval_in_minutes}, data_frequency: {data_frequency}"
        )
    query = """
        WITH acc_query AS (SELECT datetime
                                ,station_id
                                ,SUM(measured) OVER (PARTITION BY station_id ORDER BY datetime ROWS BETWEEN %(data_frequency)s PRECEDING AND CURRENT ROW) AS acc
                                ,LAG(datetime, %(data_frequency)s) OVER (PARTITION BY station_id ORDER BY datetime) AS earliest_datetime
                           FROM raw_data
                           WHERE station_id = ANY(%(station_ids)s) 
                             AND variable_id = %(variable_id)s
                             AND datetime   >= %(start_datetime)s
                             AND datetime   <= %(end_datetime)s
                             AND measured   != %(MISSING_VALUE)s)
        SELECT acc_query.datetime
              ,acc_query.station_id
              ,acc_query.acc
        FROM acc_query
        WHERE acc_query.datetime - acc_query.earliest_datetime < INTERVAL '%(interval_in_minutes)s MINUTES';
    """

    params = {
        "start_datetime": start_datetime,
        "end_datetime": end_datetime,
        "station_ids": station_ids,
        "variable_id": variable_id,
        "data_frequency": data_frequency,
        "interval_in_minutes": interval_in_minutes,
        "MISSING_VALUE": settings.MISSING_VALUE
    }

    formated_list = []
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(query, params)

        # Group records in a dictionary by datetime
        # rows[0] = datetime
        # rows[1] = station_id
        # rows[2] = acc

        rows = cursor.fetchall()
        if len(rows) == 0:
            raise Exception('No data found')

        for row in rows:
            current_datetime = row[0]
            current_station_id = row[1]
            current_value = row[2]

            if current_datetime not in date_dict:
                date_dict[current_datetime] = {}

            date_dict[current_datetime][current_station_id] = current_value

        # Validate if a datetime contains all stations measurements, calculate avg and format value 
        for datetime, station_data_dict in date_dict.items():

            current_record_station_ids = tuple(station_data_dict.keys())
            if any(station_id not in current_record_station_ids for station_id in station_ids):
                continue

            current_record_dict = {'datetime': datetime.isoformat()}
            station_count = 0
            value_acc = 0.0

            for station_id, measured in station_data_dict.items():
                current_record_dict[station_id] = measured
                station_count += 1
                value_acc += measured

            if station_count > 0:
                current_record_dict['avg'] = value_acc / station_count
                formated_list.append(current_record_dict)

    # Format output request data
    request_data = {
        "prediction_id": prediction_id,
        "data": formated_list,
    }

    logger.info(f'request_data: {repr(request_data)}')

    request = requests.post(settings.HYDROML_URL, json=request_data)

    if request.status_code != 200:
        logger.error(f'Error on predict data via HydroML, {request.status_code}')
        return

    formated_response = []
    response = json.loads(request.json())

    # Format predicted values
    for record in response:
        try:
            result = result_mapping[str(record['prediction'])]

            formated_response.append({
                "datetime": dateutil.parser.isoparse(record['datetime']),
                "target_station_id": target_station_id,
                "variable_id": variable_id,
                "result": result,
            })
        except KeyError as e:
            logger.error(
                f'Error on predict_data for prediction "{prediction_id}": Invalid mapping for result "{record["prediction"]}".')
            raise Exception(e)

    # Update records' labels
    try:
        with conn.cursor() as cursor:
            cursor.executemany(f"""
                UPDATE raw_data 
                SET ml_flag = %(result)s
                WHERE station_id = %(target_station_id)s
                  AND variable_id = %(variable_id)s 
                  AND datetime = %(datetime)s;
            """, formated_response)
        conn.commit()
    except Exception as e:
        logger.error(f'Error on update raw_data: {repr(e)}')

@shared_task
def predict_preciptation_data():
    hydroml_params = HydroMLPredictionStation.objects.all()

    end_datetime = datetime.utcnow()
    start_datetime = end_datetime - timedelta(hours=2, minutes=30)

    for hydroml_param in hydroml_params:
        current_prediction = hydroml_param.prediction
        logger.info(f"Processing Prediction: {current_prediction.name}")
        station_ids = list(hydroml_param.neighborhood.neighborhood_stations.all().values_list('station_id', flat=True))

        result_mapping = {}
        mappings = HydroMLPredictionMapping.objects.filter(hydroml_prediction_id=hydroml_param.id)
        for mapping in mappings:
            result_mapping[mapping.prediction_result] = mapping.quality_flag.id

        try:
            predict_data(start_datetime=start_datetime,
                         end_datetime=end_datetime,
                         prediction_id=current_prediction.hydroml_prediction_id,
                         station_ids=station_ids,
                         target_station_id=hydroml_param.target_station.id,
                         variable_id=current_prediction.variable_id,
                         data_period_in_minutes=hydroml_param.data_period_in_minutes,
                         interval_in_minutes=hydroml_param.interval_in_minutes,
                         result_mapping=result_mapping)
        except Exception as e:
            logger.error(f'Error on predict_preciptation_data for "{current_prediction.name}": {repr(e)}')

@shared_task
def process_hfdata_summary_tasks():
    # process only 500 HF summaries per execution
    unprocessed_hf_summaries = HFSummaryTask.objects.filter(started_at=None).values_list('station', 'variable', 'start_datetime', 'end_datetime').distinct()[:501]

    for  station_id, variable_id, s_datetime, e_datetime in unprocessed_hf_summaries:

        hf_summary_task = HFSummaryTask.objects.filter(started_at=None, station_id=station_id, variable_id=variable_id, start_datetime=s_datetime, end_datetime=e_datetime)
        hf_summary_task_ids = list(hf_summary_task.values_list('id', flat=True))

        # Updating Hourly summary task
        hfdata_summary(hf_summary_task_ids, station_id, variable_id, s_datetime, e_datetime)


def hfdata_summary(hf_summary_task_ids, station_id, variable_id, s_datetime, e_datetime):
    try:
        HFSummaryTask.objects.filter(id__in=hf_summary_task_ids).update(started_at=datetime.now(tz=pytz.UTC))
        calculate_hfdata_summary(station_id, variable_id, s_datetime, e_datetime)
    except Exception as err:
        logger.error('Error calculation hfdata summary for variable "{0}" and range ("{1}","{2}"). '.format(variable_id, s_datetime, s_datetime) + repr(err))
        db_logger.error('Error calculation hfdata summary for variable "{0}" and range ("{1}","{2}"). '.format(variable_id, s_datetime, s_datetime) + repr(err))
    else:
        HFSummaryTask.objects.filter(id__in=hf_summary_task_ids).update(finished_at=datetime.now(tz=pytz.UTC))


def calculate_hfdata_summary(station_id, variable_id, s_datetime, e_datetime):
    start_at = time()

    logger.info(f"HighFrequency summary started at {datetime.now(pytz.UTC)}")
    logger.info(
        f"HighFrequency summary parameters: "
        f"{station_id} {variable_id} {s_datetime} {e_datetime}"
    )

    _variable = Variable.objects.get(id=variable_id)

    if _variable.name != "Sea Level":
        message = (
            f"HighFrequency processing is currently only implemented for Sea Level. "
            f"Skipping station_id={station_id}, variable_id={variable_id}, "
            f"variable_name={_variable.name}, range=({s_datetime}, {e_datetime})."
        )
        logger.warning(message)
        db_logger.warning(message)
        return

    _hfdata = HighFrequencyData.objects.filter(
        station_id=station_id,
        variable_id=variable_id,
        datetime__gte=s_datetime,
        datetime__lte=e_datetime,
    )

    datetimes = list(_hfdata.values_list("datetime", flat=True))

    if not datetimes:
        message = (
            f"No HighFrequencyData found for Sea Level processing. "
            f"station_id={station_id}, variable_id={variable_id}, "
            f"range=({s_datetime}, {e_datetime})."
        )
        logger.warning(message)
        db_logger.warning(message)
        return

    seconds = (max(datetimes) - min(datetimes)).total_seconds() + 1

    data = list(_hfdata.values_list("measured", flat=True))

    reads = process_wave_data(station_id, e_datetime, data, seconds)

    if not reads:
        logger.warning(
            "process_wave_data produced 0 reads for station_id=%s, variable_id=%s, range=(%s, %s)",
            station_id,
            variable_id,
            s_datetime,
            e_datetime,
        )
        return

    insert_rd(reads)

    logger.info(
        f"HighFrequency summary finished at {datetime.now(pytz.UTC)}. "
        f"Took {time() - start_at} seconds."
    )
    

class wave(): # Wave object
    def __init__(self, frequency: float, height: float, phase_rad: float):
        self.height = height # Wave height in cm
        self.frequency = frequency # Frequency in Hz
        self.phase_rad = phase_rad # Phase is radians
        self.time = None
        self.wave = None

    def gen_sinewave(self, time):
        self.time = time
        self.wave = self.height*np.sin(self.frequency*2*np.pi*time+self.phase_rad)
        return self.wave

def fft_decompose(data, DEBUG = False):
    MEASUREMENT_PERIOD = 1 #1 Second

    MIN_AMPLITUDE = 0.01 #In m

    MIN_FREQUENCY = 0.0001 #In Hz
    MAX_FREQUENCY = 0.3 #In Hz

    # https://stackoverflow.com/questions/59725933/plot-fft-as-a-set-of-sine-waves-in-python
    t = np.arange(0, len(data)*MEASUREMENT_PERIOD, MEASUREMENT_PERIOD)

    fft = np.fft.fft(data)
    fftfreq = np.fft.fftfreq(len(t), MEASUREMENT_PERIOD)

    wave_list = []
    sinewave_list = []
    
    mid = len(t)//2 + 1    
    for i in range(mid):
        phase_rad = cmath.phase(fft[i])+np.pi/2
        amplitude = 2*abs(fft[i])/len(t)
        frequency = fftfreq[i]

        if MIN_FREQUENCY <= frequency <= MAX_FREQUENCY and MIN_AMPLITUDE <= amplitude:
            W = wave(frequency=frequency, height=amplitude, phase_rad=phase_rad)

            wave_list.append(W)
            sinewave_list.append(W.gen_sinewave(t))

            if DEBUG:
                logger.info(f"WAVE  {len(sinewave_list)}:")
                logger.info(f"Height: {amplitude}:")
                logger.info(f"Frequency: {frequency}:")
                logger.info(f"Phase: {math.degrees(phase_rad)}:")
    return wave_list

def get_top_wave_components(wave_list):
    # Top 5 amp
    wave_list.sort(key=lambda W: abs(W.height), reverse=True)
    top_wave_components = wave_list[:5]
    return top_wave_components

def process_wave_data(station_id, e_datetime, data, seconds):
    avg = np.mean(data)
    std = np.std(data)

    reads = [['Sea Level [MIN]', np.min(data)],
             ['Sea Level [MAX]', np.max(data)],
             ['Sea Level [AVG]', avg],
             ['Sea Level [STDV]', std],
             ['Significant Wave Height', 4*std]]

    # Normalizing Sea Level by average
    norm_data = data-avg

    wave_list = fft_decompose(norm_data, DEBUG = True)
    wave_list = get_top_wave_components(wave_list)

    for i, W in enumerate(wave_list):
        name = 'Wave Component '+str(i+1)
        reads += [[name+' Frequency', W.frequency],
                  [name+' Amplitude', W.height],
                  [name+' Phase', math.degrees(W.phase_rad) % 360]]

    for read in reads:
        read[0] = Variable.objects.get(name=read[0]).id

    now = datetime.now()

    df = pd.DataFrame(reads, columns=['variable_id', 'measured'])
    df['station_id'] = station_id
    df['datetime'] = e_datetime
    df['updated_at'] = now
    df['created_at'] = now
    df['seconds'] = seconds
    df['is_daily'] = False
    df['quality_flag'] = None
    df['qc_range_quality_flag'] = None
    df['qc_range_description'] = None
    df['qc_step_quality_flag'] = None
    df['qc_step_description'] = None
    df['qc_persist_quality_flag'] = None
    df['qc_persist_description'] = None
    df['manual_flag'] = None
    df['consisted'] = None

    cols = ["station_id", "variable_id", "seconds", "datetime", "measured", "quality_flag", "qc_range_quality_flag",
           "qc_range_description", "qc_step_quality_flag", "qc_step_description", "qc_persist_quality_flag",
           "qc_persist_description", "manual_flag", "consisted", "is_daily"]

    reads = df[cols].values.tolist()

    return reads




######################################################################
# logic to submit station information to OSCAR given a request
@shared_task
def export_station_to_oscar(request):
    # get wigos ID's
    selected_ids = request.POST.getlist('selected_ids[]')

    api_token = request.POST.get('api_token')

    # Log the wigos id
    print(f"Selected IDs: {selected_ids}")
    # print(f"Token entered: {api_token}")

    if selected_ids:

        stations = Station.objects.filter(wigos__in=selected_ids)

        oscar_status_msg_list = []

        try:
            for obj in stations:
                station_info = []
                
                station_info.append(str(obj.wigos))
                station_info.append(str(obj.name))
                station_info.append(str(datetime.strptime(f'{obj.begin_date}', '%Y-%m-%d %H:%M:%S%z').strftime('%Y-%m-%dT%H:%M:%SZ')))
                station_info.append(str(obj.latitude))
                station_info.append(str(obj.longitude))
                station_info.append(str(obj.elevation))
                station_info.append(str(obj.wmo_region.notation))
                station_info.append(str(obj.wmo_station_type.notation))
                station_info.append(str(obj.reporting_status.notation))
                station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')))
                station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dZ')))
                station_info.append(str(obj.country.notation))

                # export station information to OSCAR
                oscar_status_msg_list.append(exso.surface_to_oscar(station_info, api_token=api_token))

        except Exception as e:
            oscar_status_msg_list.append({'code': 406, 'description': f'{e}'})
        
    else:
        oscar_status_msg_list.append({'code': 412, 'description': 'No WIGOS ID was provided'})

    return  oscar_status_msg_list


# logic to submit station information to OSCAR given the wigos id and the api token
@shared_task
def export_station_to_oscar_wigos(selected_ids, api_token, cleaned_data):

    # Log the wigos id
    print(f"Selected IDs: {selected_ids[0]}")
    # print(f"Token entered: {api_token}")

    if selected_ids[0]:

        try:
            station_info = []
            
            station_info.append(str(selected_ids[0]))
            station_info.append(str(cleaned_data['name']))
            station_info.append(str(datetime.strptime(f"{cleaned_data['begin_date']}", '%Y-%m-%d %H:%M:%S%z').strftime('%Y-%m-%dT%H:%M:%SZ')))
            station_info.append(str(cleaned_data['latitude']))
            station_info.append(str(cleaned_data['longitude']))
            station_info.append(str(cleaned_data['elevation']))
            station_info.append(str(WMORegion.objects.filter(name=cleaned_data['wmo_region']).values_list('notation', flat=True).first()))
            station_info.append(str(WMOStationType.objects.filter(name=cleaned_data['wmo_station_type']).values_list('notation', flat=True).first()))
            station_info.append(str(WMOReportingStatus.objects.filter(name=cleaned_data['reporting_status']).values_list('notation', flat=True).first()))
            station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')))
            station_info.append(str(datetime.utcnow().strftime('%Y-%m-%dZ')))
            station_info.append(str(Country.objects.filter(name=cleaned_data['country']).values_list('notation', flat=True).first()))

            # export station information to OSCAR
            return exso.surface_to_oscar(station_info, api_token=api_token)

        except Exception as e:
            return {'code': 406, 'description': f'{e}'}
        
    else:
        return {'code': 412, 'description': 'No WIGOS ID was provided'}


# MANUAL STATIONS ONLY** Tasks the looks for stations that are set for international exchange (see STATION model) and transmits them to wis2box
def manual_transmit_wis2box(id, station_name, regional_transmit, local_transmit, add_gts, wsi):
    wis2_logs = []  # Stores log messages for debugging and reporting
    wis2_message = []
    regional_success = False
    local_success = False

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        wis2_logs.append(formatted_message)  # Store it in wis2_logs
        
    # query the data base and get all available values and log messages
    data_row, manual_logs, mws_checks = get_maunal_data(station_name, id)

    # update the wis2_logs with the logs
    for log_msg in manual_logs:
        wis2_logs.append(log_msg)

    # if mws checks passed continue on with transmission
    if not mws_checks:
        return wis2_logs, wis2_message, False

    # Helper function to upload CSV to MinIO storage
    def upload_to_minio(credentials, label, path):
        try:
            if label == "secondary":
                client = Minio(
                    endpoint=f"{credentials.regional_wis2_ip_address}:{credentials.regional_wis2_port}",
                    access_key=credentials.regional_wis2_username,
                    secret_key=str(credentials.get_regional_password()),
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
                
            else:
                client = Minio(
                    endpoint=f"{credentials.local_wis2_ip_address}:{credentials.local_wis2_port}",
                    access_key=credentials.local_wis2_username,
                    secret_key=str(credentials.get_local_password()), 
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
            
            if add_gts:
                # Function to set tag for file based on hour
                def set_tag(hour, minute):
                    if hour % 6 == 0 and minute == 0:
                        return 'main'
                    elif hour % 3 == 0 and minute == 0:
                        return 'intermediate'
                    elif minute == 0:
                        return 'non-standard'
                    else:
                        return 'other'
                    
                # Now get the tag
                tag = set_tag(data_row['hour'], data_row['minute'])

                # setting the base stations id as the id used in the file name
                filename = f"wmo_data_{wsi}_{tag}.csv"
            else:
                filename = f'wmo_data_{wsi}.csv'

            # Note that 'settings.WIS2BOX_TOPIC_HIERARCHY + filename' simply just tacks on the filename no '/' separation
            # see 'https://docs.wis2box.wis.wmo.int/en/latest/user/data-ingest.html' for and example
            # client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + filename, temp_csv_path)
            # Note: The filename and the path are not the same, but i don't believe this affects wis2 reception if an error occurs consider 
            # having the names match
            client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + '/' + filename, path)

            # Add the file name
            log_message(f"File sent to WIS2BOX under name: {filename}")

            logging.info(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            log_message(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            return True
        except Exception as e:
            logging.error(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            log_message(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            return False

    try:
        # Save the data into a temporary CSV file
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w+', delete=False) as temp_csv:
            csv_writer = csv.DictWriter(temp_csv, fieldnames=data_row.keys())
            csv_writer.writeheader()
            csv_writer.writerow(data_row)
            temp_csv_path = temp_csv.name

            # Capture the CSV content into a string
            temp_csv.seek(0)  # Move to the beginning of the file
            csv_content = temp_csv.read()

        # Transmit data to secondary (regional) and primary (local) WIS2Box if required
        if regional_transmit:
            regional_success = upload_to_minio(RegionalWisCredentials.objects.first(), "secondary", temp_csv_path)
            if regional_success:
                log_message('Push to the Secondary WIS2BOX was successful')
            else:
                log_message('Push to the Secondary WIS2BOX failed')
        if local_transmit:
            local_success = upload_to_minio(LocalWisCredentials.objects.first(), "primary", temp_csv_path)
            if local_success:
                log_message('Push to the Primary WIS2BOX was successful')
            else:
                log_message('Push to the Primary WIS2BOX failed')

        # Add the CSV data to the list
        wis2_message.append(csv_content)
    finally:
        # Clean up the temporary CSV file
        os.remove(temp_csv_path)
        logging.info(f'AWS Transmission for Station: {station_name} complete')
        log_message(f'AWS Transmission for Station: {station_name} complete')

        # if both regional and local transmission are selected both pushes must be success to get a 'success status'
        if regional_transmit and local_transmit:
            success_push = regional_success and local_success
        # else if only regional transmission is set the 'success status' is based on only regional transmission
        elif regional_transmit:
            success_push = regional_success
        # else if only local transmission is set the 'success status' is based on only local transmission
        elif local_transmit:
            success_push = local_success
        else:
            success_push = False

    return wis2_logs, wis2_message, success_push


# AWS STATIONS ONLY** Tasks the looks for stations that are set for international exchange (see STATION model) and transmits them to wis2box
def aws_transmit_wis2box(id, station_name, regional_transmit, local_transmit, add_gts, wsi):
    wis2_logs = []  # Stores log messages for debugging and reporting
    wis2_message = []
    regional_success = False
    local_success = False

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        wis2_logs.append(formatted_message)  # Store it in wis2_logs

    # query the data base and get all available values and log messages
    data_row, aws_logs, aws_checks = get_aws_data(station_name, id)

    # update the wis2_logs with the logs
    for log_msg in aws_logs:
        wis2_logs.append(log_msg)

    # if aws checks passed continue on with transmission
    if not aws_checks:
        return wis2_logs, wis2_message, False

    # Helper function to upload CSV to MinIO storage
    def upload_to_minio(credentials, label, path):
        try:
            if label == "regional":
                client = Minio(
                    endpoint=f"{credentials.regional_wis2_ip_address}:{credentials.regional_wis2_port}",
                    access_key=credentials.regional_wis2_username,
                    secret_key=str(credentials.get_regional_password()),
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes
                        )
                    )
                )
                
            else:
                client = Minio(
                    endpoint=f"{credentials.local_wis2_ip_address}:{credentials.local_wis2_port}",
                    access_key=credentials.local_wis2_username,
                    secret_key=str(credentials.get_local_password()), 
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes
                        )
                    )
                )
                
            if add_gts:
                # Function to set tag for file based on hour
                def set_tag(hour, minute):
                    if hour % 6 == 0 and minute == 0:
                        return 'main'
                    elif hour % 3 == 0 and minute == 0:
                        return 'intermediate'
                    elif minute == 0:
                        return 'non-standard'
                    else:
                        return 'other'
                    
                # Now get the tag
                tag = set_tag(data_row['hour'], data_row['minute'])

                filename = f"wmo_data_{wsi}_{tag}.csv"
            else:
                filename = f'wmo_data_{wsi}.csv'

            # Note that 'settings.WIS2BOX_TOPIC_HIERARCHY + filename' simply just tacks on the filename no '/' separation
            # see 'https://docs.wis2box.wis.wmo.int/en/latest/user/data-ingest.html' for and example
            # client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + filename, temp_csv_path)
            client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + '/' + filename, path)

            # Add the file name
            log_message(f"File sent to WIS2BOX under name: {filename}")

            logging.info(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            log_message(f"Data transfer to {label} WIS2Box successful for Station: {station_name}")
            return True
        except Exception as e:
            logging.error(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            log_message(f"{label} WIS2Box Connection Error for station {station_name}: {e}")
            return False

    try:
        # Save the data into a temporary CSV file
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w+', delete=False) as temp_csv:
            csv_writer = csv.DictWriter(temp_csv, fieldnames=data_row.keys())
            csv_writer.writeheader()
            csv_writer.writerow(data_row)
            temp_csv_path = temp_csv.name

            # Capture the CSV content into a string
            temp_csv.seek(0)  # Move to the beginning of the file
            csv_content = temp_csv.read()

        # Transmit data to regional and local WIS2Box if required
        if regional_transmit:
            regional_success = upload_to_minio(RegionalWisCredentials.objects.first(), "regional", temp_csv_path)
            if regional_success:
                log_message('Push to the Regional WIS2BOX was successfull')
            else:
                log_message('Push to the Regional WIS2BOX failed')
        if local_transmit:
            local_success = upload_to_minio(LocalWisCredentials.objects.first(), "local", temp_csv_path)
            if local_success:
                log_message('Push to the Local WIS2BOX was successfull')
            else:
                log_message('Push to the Local WIS2BOX failed')

        # Add the CSV data to the list
        wis2_message.append(csv_content)

    finally:
        # Clean up the temporary CSV file
        os.remove(temp_csv_path)
        logging.info(f'AWS Transmission for Station: {station_name} complete')
        log_message(f'AWS Transmission for Station: {station_name} complete')
        
        # if both regional and local transmission are selected both pushes must be success to get a 'success status'
        if regional_transmit and local_transmit:
            success_push = regional_success and local_success
        # else if only regional transmission is set the 'success status' is based on only regional transmission
        elif regional_transmit:
            success_push = regional_success
        # else if only local transmission is set the 'success status' is based on only local transmission
        elif local_transmit:
            success_push = local_success
        else:
            success_push = False

    return wis2_logs, wis2_message, success_push


# transmit for hybrid stations
def hybrid_transmit_wis2box(id, station_name, regional_transmit, local_transmit, hybrid_station_id, base_aws, add_gts, hybrid_station_name, wsi):
    wis2_logs = []  # Stores log messages for debugging and reporting
    wis2_message = []
    regional_success = False
    local_success = False

    # Get UTC time and truncate to the hour
    unformatted_now = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # Format it as 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    formatted_now = unformatted_now.strftime('%Y-%m-%d %H:%M:%S+00')

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        wis2_logs.append(formatted_message)  # Store it in wis2_logs

    # setting which station to get AWS data from depending on the base_aws data
    # see the Wis2BoxPublish model for more information
    if not base_aws:
        secondary_station = id
        id = hybrid_station_id
    else:
        secondary_station = hybrid_station_id
    

    # query the data base and get all available values and log messages
    aws_data_row, aws_logs, aws_checks = get_aws_data(f"HYBRID STATION:[{station_name}][{hybrid_station_name}]", id)

    # query the data base and get all available values and log messages
    mws_data_row, manual_logs, mws_checks = get_maunal_data(f"HYBRID STATION:[{station_name}][{hybrid_station_name}]", secondary_station)

    # update the wis2_logs with the aws logs
    for log_msg in aws_logs:
        wis2_logs.append(log_msg)

    # update the wis2_logs with the mws logs
    for log_msg in manual_logs:
        wis2_logs.append(log_msg)


    # IF BOTH AWS STATION CHECKS AND MWS STATION CHECKS FAIL A HYBRID TRANSMISSION CANNOT HAPPEN
    if not aws_checks and not mws_checks:
        return wis2_logs, wis2_message, False
    
    # dictionary to hold all the values along with the column names going in the message. (a combination of AWS and MWS data)
    data_row = {}

    # load data_row with data from the AWS station
    for col in aws_data_row:
        # if aws station pressure isn't complete (meaning msl and/or station pressure is missing) 
        # use all three pressure readings from the manual station. Pressure from AWS stations and 
        # pressure from MWS stations are calculated with slightly different instruments, this method
        # avoids any inaccuracies.
        # either take all pressure readings from AWS or MWS but don't mix both
        if col == "station_pressure":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row[col]
                data_row["msl_pressure"] = mws_data_row["msl_pressure"]
                data_row["24_hour_barometric_change"] = mws_data_row["24_hour_barometric_change"]
            else:
                data_row[col] = aws_data_row[col]

        elif col == "msl_pressure":
            if aws_data_row["station_pressure"] == None:
                pass
            else:
                data_row[col] = aws_data_row[col]

        elif col == "air_temperature":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["air_temperature"]
            else:
                data_row[col] = aws_data_row[col]

        elif col == "dewpoint_temperature":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["dewpoint_temperature"]
            else:
                data_row[col] = aws_data_row[col]
            
            data_row["maximum_air_temperature"] = mws_data_row["maximum_air_temperature"]
            data_row["minimum_air_temperature"] = mws_data_row["minimum_air_temperature"]
            data_row["wetbulb_temperature"] = mws_data_row["wetbulb_temperature"]

        elif col == "relative_humidity":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["relative_humidity"]
            else:
                data_row[col] = aws_data_row[col]

        elif col == "wind_direction":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["wind_direction"]
            else:
                data_row[col] = aws_data_row[col]

        elif col == "wind_speed":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["wind_speed"]
            else:
                data_row[col] = aws_data_row[col]

        elif col == "total_precipitation_1_hour":
            data_row["precipitation_indicator"] = mws_data_row["precipitation_indicator"]
            data_row["precipitation_period_duration"] = mws_data_row["precipitation_period_duration"]

            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["precipitation"]
            else:
                data_row[col] = aws_data_row[col]

        elif col == "total_precipitation_24_hours":
            if aws_data_row[col] == None:
                data_row[col] = mws_data_row["total_precipitation_24_hours"]
            else:
                data_row[col] = aws_data_row[col]

        else:
            data_row[col] = aws_data_row[col]

    # load data_row with remaining data from the MWS station
    remaining_mws_col = ["lowest_cloud_height","visibility","cloud_cover_tot",
                         "present_weather","past_weather_w_1","past_weather_w_2","observed_cloud_coverage",
                         "low_cloud_type","middle_cloud_type","high_cloud_type","state_of_sky","direction_of_cl_clouds",
                         "direction_of_cm_clouds","direction_of_ch_clouds","cloud_coverage_l1","genus_of_cloud_l1",
                         "height_of_cloud_base_l1","cloud_coverage_l2","genus_of_cloud_l2","height_of_cloud_base_l2",
                         "cloud_coverage_l3","genus_of_cloud_l3","height_of_cloud_base_l3","cloud_coverage_l4",
                         "genus_of_cloud_l4","height_of_cloud_base_l4","special_phenomenon"
                        ]
    
    for col in remaining_mws_col:
        data_row[col] = mws_data_row[col]

    # confirm the station metadata was added if not use the MWS metadata
    mws_metadata = ["wsi_series","wsi_issuer","wsi_issue_number","wsi_local",
                    "wmo_block_number","wmo_station_number","station_type",
                    "year","month","day","hour","minute","latitude","longitude",
                    'thermometer_height', 'anemometer_height', 'station_height_above_msl', 
                    'barometer_height_above_msl', 'rain_sensor_height']
    
    # When AWS data is missing (hour values are None),
    # this logic falls back to using data from the MWS station
    # to ensure the dataset remains complete.
    if data_row["hour"] == None:
        for col in mws_metadata:
                data_row[col] = mws_data_row[col] 

    # set the station type to hybrid. 2 is the code for hybrid (both manned and automatic)
    log_message("This is a hybrid transmision therefore station_type will be changed:")
    data_row["station_type"] = 2
    log_message("Data for 'station_type' value [2].")


    # Helper function to upload CSV to MinIO storage
    def upload_to_minio(credentials, label, path):
        try:
            if label == "regional":
                client = Minio(
                    endpoint=f"{credentials.regional_wis2_ip_address}:{credentials.regional_wis2_port}",
                    access_key=credentials.regional_wis2_username,
                    secret_key=str(credentials.get_regional_password()),
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
                
            else:
                client = Minio(
                    endpoint=f"{credentials.local_wis2_ip_address}:{credentials.local_wis2_port}",
                    access_key=credentials.local_wis2_username,
                    secret_key=str(credentials.get_local_password()), 
                    secure=False,
                    http_client=urllib3.PoolManager(
                        timeout=10.0,  # 10-second timeout
                        retries=urllib3.util.Retry(
                            total=1,  # Retry once
                            backoff_factor=0,  # No backoff (instant retry)
                            raise_on_status=False  # Do not raise exceptions on certain status codes (optional)
                        )
                    )
                )
            
            if add_gts:
                # Function to set tag for file based on hour
                def set_tag(hour, minute):
                    if hour % 6 == 0 and minute == 0:
                        return 'main'
                    elif hour % 3 == 0 and minute == 0:
                        return 'intermediate'
                    elif minute == 0:
                        return 'non-standard'
                    else:
                        return 'other'
                    
                # Now get the tag
                tag = set_tag(data_row['hour'], data_row['minute'])

                # setting the base stations id as the id used in the file name
                if base_aws:
                    filename = f"wmo_data_{wsi}_{tag}.csv"
                else:
                    filename = f"wmo_data_{wsi}_{tag}.csv"
            else:
                filename = f'wmo_data_{wsi}.csv'

            # Note that 'settings.WIS2BOX_TOPIC_HIERARCHY + filename' simply just tacks on the filename no '/' separation
            # see 'https://docs.wis2box.wis.wmo.int/en/latest/user/data-ingest.html' for and example
            # client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + filename, temp_csv_path)
            # Note: The filename and the path are not the same, but i don't believe this affects wis2 reception if an error occurs consider 
            # having the names match
            client.fput_object('wis2box-incoming', settings.WIS2BOX_TOPIC_HIERARCHY + '/' + filename, path)

            # Add the file name
            log_message(f"File sent to WIS2BOX under name: {filename}")

            logging.info(f"Data transfer to {label} WIS2Box successful for Station: HYBRID STATION:[{station_name}][{hybrid_station_name}]")
            log_message(f"Data transfer to {label} WIS2Box successful for Station: HYBRID STATION:[{station_name}][{hybrid_station_name}]")
            return True
        except Exception as e:
            logging.error(f"{label} WIS2Box Connection Error for station HYBRID STATION:[{station_name}][{hybrid_station_name}]: {e}")
            log_message(f"{label} WIS2Box Connection Error for station HYBRID STATION:[{station_name}][{hybrid_station_name}]: {e}")
            return False

    try:
        # Save the data into a temporary CSV file
        with tempfile.NamedTemporaryFile(suffix='.csv', mode='w+', delete=False) as temp_csv:
            csv_writer = csv.DictWriter(temp_csv, fieldnames=data_row.keys())
            csv_writer.writeheader()
            csv_writer.writerow(data_row)
            temp_csv_path = temp_csv.name

            # Capture the CSV content into a string
            temp_csv.seek(0)  # Move to the beginning of the file
            csv_content = temp_csv.read()

        # Transmit data to regional and local WIS2Box if required
        if regional_transmit:
            regional_success = upload_to_minio(RegionalWisCredentials.objects.first(), "regional", temp_csv_path)
            if regional_success:
                log_message('Push to the Regional WIS2BOX was successfull')
            else:
                log_message('Push to the Regional WIS2BOX failed')
        if local_transmit:
            local_success = upload_to_minio(LocalWisCredentials.objects.first(), "local", temp_csv_path)
            if local_success:
                log_message('Push to the Local WIS2BOX was successfull')
            else:
                log_message('Push to the Local WIS2BOX failed')

        # Add the CSV data to the list
        wis2_message.append(csv_content)
    finally:
        # Clean up the temporary CSV file
        os.remove(temp_csv_path)
        logging.info(f'AWS Transmission for Station: {station_name} complete')
        log_message(f'AWS Transmission for Station: {station_name} complete')

        # if both regional and local transmission are selected both pushes must be success to get a 'success status'
        if regional_transmit and local_transmit:
            success_push = regional_success and local_success
        # else if only regional transmission is set the 'success status' is based on only regional transmission
        elif regional_transmit:
            success_push = regional_success
        # else if only local transmission is set the 'success status' is based on only local transmission
        elif local_transmit:
            success_push = local_success
        else:
            success_push = False

    return wis2_logs, wis2_message, success_push


@shared_task
def wis2publish_cleanup():
    # clean up logs older than 24 hours
    try:
        # Delete logs older than 24 hours
        time_threshold = now() - timedelta(hours=24)
        deleted_count, _ = Wis2BoxPublishLogs.objects.filter(created_at__lt=time_threshold).delete()

        logger.info(f"WIS2 Publish Cleanup tasks: Successfully deleted {deleted_count} logs older than 24 hours")
    except Exception as e:
        logger.error("WIS2 Publish Cleanup tasks: Failed to delete logs older than 24 hours")
        logger.exception(f"Error: {e}")

    # Update publish logs number and fail logs number
    try:
        # Annotate each Wis2BoxPublish record with the count of related success and fail logs
        wis2publish_updates = Wis2BoxPublish.objects.annotate(
            success_count=Count('wis2boxpublishlogs', filter=Q(wis2boxpublishlogs__success_log=True)),
            fail_count=Count('wis2boxpublishlogs', filter=Q(wis2boxpublishlogs__success_log=False))
        )

        update_count = 0  # Counter to track the number of successful updates

        # Iterate over each annotated Wis2BoxPublish record and update the success/fail counts
        for record in wis2publish_updates:
            Wis2BoxPublish.objects.filter(id=record.id).update(
                publish_success=record.success_count,  # Set the success count
                publish_fail=record.fail_count  # Set the fail count
            )
            update_count += 1  # Increment the count of updated records
        
        # Log the number of records successfully updated
        logger.info(f"WIS2 Publish Cleanup tasks: Successfully updated publish_fail and publish_success for {update_count} records in the wis2publish table")
    except Exception as e:
        # Log an error message if the update process fails
        logger.error("WIS2 Publish Cleanup tasks: Failed to update publish_fail and publish_success for some records")
        logger.exception(f"Error: {e}")  # Log the full exception traceback for debugging


@shared_task
def wis2publish_task():
    now = datetime.now(pytz.UTC)

    # retrieving wis2 tranmit station which are set to publishing
    _publishing_stations = Wis2BoxPublish.objects.filter(publishing=True)

    for _entry in _publishing_stations:
        # check if the cron schedule for the station matches the time now
        if croniter.match(_entry.publishing_offset.cron_schedule, now):
            # check if hybrid station or not
            if _entry.hybrid:
                wis2push_hybrid.delay(_entry.id, now)
            else:
                wis2push_regular.delay(_entry.id, now)


@shared_task
# for instant pushing
def wis2publish_task_now(station_id):
    try:
        now = datetime.now(pytz.UTC)

        # grabbing a single entry base on the id
        _entry = Wis2BoxPublish.objects.get(id=station_id)

        if _entry.hybrid:
            wis2push_hybrid(_entry.id, now)
        else:
            wis2push_regular(_entry.id, now)
    except Exception as e:
        raise e


# push to wis2 for non hybrid stations
@shared_task
def wis2push_regular(entry_id, now):
    # get the entry
    _entry = Wis2BoxPublish.objects.get(id=entry_id)

    # get station obj
    _station_obj = Station.objects.get(id=_entry.station.id)

    # get the wsi and prepare it for use in the transmission
    wigos = _entry.station.wigos

    if wigos:
        wsi = wigos.replace('-','_').replace(' ', '') # replacing '-' with '_' (wsi will be used in the transmission file name)
    else:
        wsi = f"no_wsi_found_{_entry.station.id}"

    # differentiating between manual and automatic stations
    if _station_obj.is_automatic:
        # get the data from the transmission
        wis2_logs, wis2_message, success_push = aws_transmit_wis2box(_entry.station.id, 
                                                                        _entry.station.name, 
                                                                        _entry.regional_wis2, 
                                                                        _entry.local_wis2,
                                                                        _entry.add_gts,
                                                                        wsi)
    else:
        # get the data from the transmission
        wis2_logs, wis2_message, success_push = manual_transmit_wis2box(_entry.station.id, 
                                                                            _entry.station.name, 
                                                                            _entry.regional_wis2, 
                                                                            _entry.local_wis2,
                                                                            _entry.add_gts,
                                                                            wsi)

    message_exist = bool(wis2_message)

    # create the logs entry
    Wis2BoxPublishLogs.objects.create(
        created_at=now, 
        last_modified= datetime.now(tz=pytz.UTC),
        publish_station_id=_entry.id,
        success_log=success_push,
        log="\n\n".join(wis2_logs),
        wis2message_exist=message_exist,
        wis2message="\n".join(wis2_message)
    )

    # call clean up tasks
    wis2publish_cleanup() 


# push to wis2 for hybrid stations
@shared_task
def wis2push_hybrid(entry_id, now):
    # get the entry
    _entry = Wis2BoxPublish.objects.get(id=entry_id)

    # get the wsi and prepare it for use in the transmission
    wigos = _entry.station.wigos

    if wigos:
        wsi = wigos.replace('-','_').replace(' ', '') # replacing '-' with '_' (wsi will be used in the transmission file name)
    else:
        wigos = _entry.hybrid_station.wigos

        if wigos:
            wsi = wigos.replace('-','_').replace(' ', '') # replacing '-' with '_' (wsi will be used in the transmission file name)
        else:
            wsi = f"no_wsi_found_{_entry.station.id}"

    # get data from the transmission
    wis2_logs, wis2_message, success_push = hybrid_transmit_wis2box(_entry.station.id, 
                                                                        _entry.station.name, 
                                                                        _entry.regional_wis2, 
                                                                        _entry.local_wis2, 
                                                                        _entry.hybrid_station.id,
                                                                        _entry.base_aws,
                                                                        _entry.add_gts,
                                                                        _entry.hybrid_station.name,
                                                                        wsi)

    message_exist = bool(wis2_message)

    # create the logs entry
    Wis2BoxPublishLogs.objects.create(
        created_at=now, 
        last_modified= datetime.now(tz=pytz.UTC),
        publish_station_id=_entry.id,
        success_log=success_push,
        log="\n\n".join(wis2_logs),
        wis2message_exist=message_exist,
        wis2message="\n".join(wis2_message)
    )

    # call clean up tasks
    wis2publish_cleanup() 


# Helper function to execute database queries safely
def execute_query(query, params, error_message):
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            cursor_results = cursor.fetchall()
            # log_message(f"Query result: {cursor_results}")
            return [list(row) for row in cursor_results]
    except Exception as e:
        logger.error(f"{error_message}: {e}")
        return []
    
        
# get manual data for WIS2 transmission
def get_maunal_data(station_name, id):
    # log messages
    logs_manual_wis2_transmit = []

    # Get UTC time and truncate to the hour
    unformatted_now = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # Format it as 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    formatted_now = unformatted_now.strftime('%Y-%m-%d %H:%M:%S+00')

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        logs_manual_wis2_transmit.append(formatted_message)  # Store it in logs_manual_wis2_transmit
        

    logging.info(f'MWS Transmission: Processing data for Station: {station_name}')
    log_message(f'MWS Transmission: Processing data for Station: {station_name}')

    # Initialize data_row dictionary with default None values for each weather parameter
    data_row = {key: None for key in [
        'wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 'wmo_block_number',
        'wmo_station_number', 'station_type', 'year', 'month', 'day', 'hour', 'minute',
        'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl',
        'station_pressure', 'msl_pressure', '24_hour_barometric_change', 'geopotential_height', 
        'thermometer_height', 'air_temperature', 'maximum_air_temperature', 'minimum_air_temperature', 
        'dewpoint_temperature', 'wetbulb_temperature', 'relative_humidity', 'method_of_ground_state_measurement', 
        'ground_state', 'method_of_snow_depth_measurement', 'snow_depth', 'present_weather', 
        'precipitation_intensity', 'precipitation_indicator', 'precipitation', 'precipitation_period_duration', 
        'anemometer_height', 'time_period_of_wind', 'wind_direction', 'wind_speed', 'maximum_wind_gust_direction_10_minutes',
        'maximum_wind_gust_speed_10_minutes', 'maximum_wind_gust_direction_1_hour',
        'maximum_wind_gust_speed_1_hour', 'maximum_wind_gust_direction_3_hours',
        'maximum_wind_gust_speed_3_hours', 'rain_sensor_height', 'total_precipitation_1_hour',
        'total_precipitation_3_hours', 'total_precipitation_6_hours',
        'total_precipitation_12_hours', 'total_precipitation_24_hours', 'solar_radiation',
        'soil_temperature_depth_1', 'soil_temperature_depth_2', 'soil_temperature_1', 
        'soil_temperature_2', 'soil_moisture_1', 'soil_moisture_2', 'visibility', 'cloud_cover_tot', 'lowest_cloud_height',
        'past_weather_w_1', 'past_weather_w_2', 'observed_cloud_coverage',
        'low_cloud_type', 'middle_cloud_type', 'high_cloud_type', 'state_of_sky',
        'direction_of_cl_clouds', 'direction_of_cm_clouds', 'direction_of_ch_clouds',
        'cloud_coverage_l1', 'genus_of_cloud_l1',
        'height_of_cloud_base_l1', 'cloud_coverage_l2', 'genus_of_cloud_l2',
        'height_of_cloud_base_l2', 'cloud_coverage_l3', 'genus_of_cloud_l3',
        'height_of_cloud_base_l3', 'cloud_coverage_l4', 'genus_of_cloud_l4',
        'height_of_cloud_base_l4', 'special_phenomenon']}
    

    # checking if there is data in the database at the top of the hour (if not the output will be an error)
    # Query latest datetime entries for both tables
    # hourly_summary_status = execute_query("SELECT MAX(datetime) FROM hourly_summary WHERE station_id = %s", [id], "Error fetching data")
    raw_data_status = execute_query("SELECT MAX(datetime) FROM raw_data WHERE station_id = %s", [id], "Error fetching data")

    # Check if either has data for the current hour
    # if hourly_summary_status or raw_data_status:
    if raw_data_status:
        if raw_data_status[0][0] is None or raw_data_status[0][0] < unformatted_now:
            log_message(f"Comparing DB time {raw_data_status[0][0]} with current time {unformatted_now}")
            log_message(f"NO data exists in the Raw Data Table at timeslot [{formatted_now}] for Station: {station_name}")
            return data_row, logs_manual_wis2_transmit, False
    else:
        log_message(f"NO data exists in the Database at timeslot [{formatted_now}] for Station: {station_name}")
        return data_row, logs_manual_wis2_transmit, False
    

    ##########################################################################################
    # Fetch station information from the database
    station_query = """
        SELECT wigos, latitude, longitude, elevation, is_automatic
        FROM wx_station WHERE id = %s
    """

    log_message("-------------------------STATION METADATA QUERY-------------------------")

    station_data = execute_query(station_query, [id], "Error fetching station info")
    if station_data:
        row = station_data[0]
        wsi_parts = row[0].split("-") if row[0] else [None] * 4
        data_row.update({
            'wsi_series': wsi_parts[0], 
            'wsi_issuer': wsi_parts[1],
            'wsi_issue_number': wsi_parts[2], 
            'wsi_local': wsi_parts[3],
            'latitude': round(row[1], 5), 
            'longitude': round(row[2], 5),
            'station_height_above_msl': row[3], 
            'barometer_height_above_msl': round(row[3] + 1.7, 2), # 1.7 metre above station_height_above_msl
            # the settings below will be hard coded until further notice
            'thermometer_height': 1.7, 
            'anemometer_height': 10, 
            'time_period_of_wind': -10, # for this particular one the user should be able to change this
            'rain_sensor_height': 1.5, 
            'station_type': 1,
        })

    # write info out to logs
    station_metadata_keys = ['wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 
                             'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl', 
                             'thermometer_height', 'anemometer_height', 'time_period_of_wind', 'rain_sensor_height', 
                             'station_type']

    for key in station_metadata_keys:
        log_message(f"Data for '{key}' value [{data_row[key]}].")

    # Populate year, month, day, hour, minute in data_row
    if unformatted_now:
        # Add the extracted date components to data_row
        data_row['year'] = unformatted_now.year
        data_row['month'] = unformatted_now.month
        data_row['day'] = unformatted_now.day
        data_row['hour'] = unformatted_now.hour
        data_row['minute'] = unformatted_now.minute

    ##########################################################################################

    # ------------- GETTING MANUAL STATIONS DATA START ------------- #
    # query to get the measured values
    manual_data_query = """
        SELECT variable_id, measured
        FROM raw_data
        WHERE station_id = %s 
        AND datetime = %s
        AND variable_id IN (0, 10, 14, 16, 18, 19, 30, 50, 55, 60, 61,
                            4062, 4055, 4056, 4057, 4058, 4059, 4060, 4061);
    """

    # query to get the code values
    manual_data_query_code = """
        SELECT variable_id, code
        FROM raw_data
        WHERE station_id = %s 
        AND datetime = %s
        AND variable_id IN (1001, 1002, 1003, 1004, 1006, 1007, 1008, 1009, 1010, 1012,
                            1013, 1015, 1016, 1018, 1019, 4005, 4006, 
                            4041, 4044, 4045, 4046, 4047, 4043);
    """

    # Execute the query
    secondary_station_raw_data = execute_query(manual_data_query, [id, formatted_now], "Error fetching non coded manual data")
    secondary_station_raw_data_code = execute_query(manual_data_query_code, [id, formatted_now], "Error fetching coded manual data")

    # Mapping of variable_id to dictionary keys
    variable_mapping = {
        4041: 'precipitation_indicator',
        4062: 'lowest_cloud_height', 4056: 'visibility', 1001: 'cloud_cover_tot', 
        55: 'wind_direction', 50: 'wind_speed', 10: 'air_temperature', 
        19: 'dewpoint_temperature', 18: 'wetbulb_temperature', 30: 'relative_humidity', 
        60: 'station_pressure', 61: 'msl_pressure', 0: 'precipitation', 
        4043: 'precipitation_period_duration', 4057: '24_hour_barometric_change', 4055: 'total_precipitation_24_hours', 
        1002: 'present_weather', 1003: 'past_weather_w_1', 1004: 'past_weather_w_2', 
        4005: 'observed_cloud_coverage', 1006: 'low_cloud_type', 1007: 'middle_cloud_type', 
        1008: 'high_cloud_type', 4044: 'state_of_sky', 4045: 'direction_of_cl_clouds', 14: 'minimum_air_temperature',
        4046: 'direction_of_cm_clouds', 4047: 'direction_of_ch_clouds', 16: 'maximum_air_temperature', 
        1009: 'cloud_coverage_l1', 1010: 'genus_of_cloud_l1', 4058: 'height_of_cloud_base_l1', 
        1012: 'cloud_coverage_l2', 1013: 'genus_of_cloud_l2', 4059: 'height_of_cloud_base_l2', 
        1015: 'cloud_coverage_l3', 1016: 'genus_of_cloud_l3', 4060: 'height_of_cloud_base_l3', 
        1018: 'cloud_coverage_l4', 1019: 'genus_of_cloud_l4', 4061: 'height_of_cloud_base_l4', 
        4006: 'special_phenomenon'}


    # Populate the dictionary with retrieved data
    log_message("-------------------------RAW DATA QUERY-------------------------")
    for row in secondary_station_raw_data:
        variable_id, measured_value = row

        # if the value retrieve is -99.9 that means the value is missing therefore leave the entry empty
        if str(measured_value) == "-99.9":
            if measured_value == "/":
                measured_value = ""

            if variable_id in [4058, 4059, 4060, 4061]: # replacing all cloud base height values which are blank (-99.9) to '/'
                measured_value = ""
            else:
                measured_value = None

        if variable_id in variable_mapping:
            if measured_value and variable_id == 50: # converting the wind speed in knots to meters per second
                data_row[variable_mapping[variable_id]] = round(float(measured_value / 1.944), 2)

            elif measured_value and variable_id in [10, 14, 16, 18, 19]:  # converting all temperature values to kelvin
                data_row[variable_mapping[variable_id]] = round(float(measured_value) + 273.15, 2)

            elif measured_value and variable_id in [60, 61, 4057]:  # converting all pressure values to pascals
                data_row[variable_mapping[variable_id]] = round(float(measured_value) * 100)

            else:
                data_row[variable_mapping[variable_id]] = measured_value   

            log_message(f"Data for '{variable_mapping[variable_id]}' value [{data_row[variable_mapping[variable_id]]}].") # loggig that data

    # Populate the dictionary with retrieved data (only data entries which store codes)
    log_message("-------------------------CODE QUERY-------------------------")
    for row in secondary_station_raw_data_code:
        variable_id, code_value = row

        if str(code_value) == "/":
            code_value = "" # '/' should not be included in the csv file

        if variable_id in variable_mapping:
            if code_value and variable_id == 4043 and code_value != "/":
                data_row[variable_mapping[variable_id]] = f"-{int(code_value) * 6}"
            else:
                data_row[variable_mapping[variable_id]] = code_value   

            log_message(f"Data for '{variable_mapping[variable_id]}' value [{data_row[variable_mapping[variable_id]]}].") # loggig that data
    # ------------- GETTING MANUAL STATIONS DATA END ------------- #

    log_message("-----------------------------------------------------------------")
    logging.info(f'MWS Transmission: Processing data for Station Complete: {station_name}')
    log_message(f'MWS Transmission: Processing data for Station Complete: {station_name}')


    return data_row, logs_manual_wis2_transmit, True


# get aws data for WIS2 transmission
def get_aws_data(station_name, id):
    # log messages
    logs_aws_wis2_transmit = []

    # Get UTC time and truncate to the hour
    unformatted_now = datetime.now(timezone.utc).replace(minute=0, second=0, microsecond=0)

    # Format it as 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    formatted_now = unformatted_now.strftime('%Y-%m-%d %H:%M:%S+00')

    # current time in the format 'YYYY-MM-DD HH:MM:SS+00' for use in the queries
    current_now = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S+00')

    # Formatted time 24 hrs ago
    formatted_past_24_hrs = (unformatted_now - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S+00')

    # Formatted time 3 hrs ago
    formatted_past_3_hrs = (unformatted_now - timedelta(hours=3)).strftime('%Y-%m-%d %H:%M:%S+00')

    # Formatted time 1 hr ago
    formatted_past_1_hr = (unformatted_now - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S+00')

    # Formatted time 10 min ago
    formatted_past_10_min = (unformatted_now - timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S+00')

    # keep track of the logs to send back
    def log_message(message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"
        logs_aws_wis2_transmit.append(formatted_message)  # Store it in logs_aws_wis2_transmit

    logging.info(f'AWS Transmission: Processing data for Station: {station_name}')
    log_message(f'AWS Transmission: Processing data for Station: {station_name}')

    # Initialize data_row dictionary with default None values for each weather parameter
    data_row = {key: None for key in [
        'wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 'wmo_block_number',
        'wmo_station_number', 'station_type', 'year', 'month', 'day', 'hour', 'minute',
        'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl',
        'station_pressure', 'msl_pressure', '24_hour_barometric_change', 'geopotential_height', 
        'thermometer_height', 'air_temperature', 'maximum_air_temperature', 'minimum_air_temperature', 
        'dewpoint_temperature', 'wetbulb_temperature', 'relative_humidity', 'method_of_ground_state_measurement', 
        'ground_state', 'method_of_snow_depth_measurement', 'snow_depth', 'present_weather', 
        'precipitation_intensity', 'precipitation_indicator', 'precipitation', 'precipitation_period_duration', 
        'anemometer_height', 'time_period_of_wind', 'wind_direction', 'wind_speed', 'maximum_wind_gust_direction_10_minutes',
        'maximum_wind_gust_speed_10_minutes', 'maximum_wind_gust_direction_1_hour',
        'maximum_wind_gust_speed_1_hour', 'maximum_wind_gust_direction_3_hours',
        'maximum_wind_gust_speed_3_hours', 'rain_sensor_height', 'total_precipitation_1_hour',
        'total_precipitation_3_hours', 'total_precipitation_6_hours',
        'total_precipitation_12_hours', 'total_precipitation_24_hours', 'solar_radiation',
        'soil_temperature_depth_1', 'soil_temperature_depth_2', 'soil_temperature_1', 
        'soil_temperature_2', 'soil_moisture_1', 'soil_moisture_2', 'visibility', 'cloud_cover_tot', 'lowest_cloud_height',
        'past_weather_w_1', 'past_weather_w_2', 'observed_cloud_coverage',
        'low_cloud_type', 'middle_cloud_type', 'high_cloud_type', 'state_of_sky',
        'direction_of_cl_clouds', 'direction_of_cm_clouds', 'direction_of_ch_clouds',
        'cloud_coverage_l1', 'genus_of_cloud_l1',
        'height_of_cloud_base_l1', 'cloud_coverage_l2', 'genus_of_cloud_l2',
        'height_of_cloud_base_l2', 'cloud_coverage_l3', 'genus_of_cloud_l3',
        'height_of_cloud_base_l3', 'cloud_coverage_l4', 'genus_of_cloud_l4',
        'height_of_cloud_base_l4', 'special_phenomenon']}


    # checking if there is data in the database at the top of the hour (if not the output will be an error)
    # Query latest datetime entries for both tables
    hourly_summary_status = execute_query("SELECT MAX(datetime) FROM hourly_summary WHERE station_id = %s", [id], "Error fetching data")
    raw_data_status = execute_query("SELECT MAX(datetime) FROM raw_data WHERE station_id = %s", [id], "Error fetching data")

    # Check if either has data for the current hour
    if hourly_summary_status or raw_data_status:
        if (
            not hourly_summary_status or
            hourly_summary_status[0][0] is None or
            hourly_summary_status[0][0] < unformatted_now
        ):
            log_message(f"Comparing DB time {hourly_summary_status[0][0] if hourly_summary_status else 'None'} with current time {unformatted_now}")
            log_message(f"NO data exists in the Hourly Summary Table at timeslot [{formatted_now}] for Station: {station_name}")
            return data_row, logs_aws_wis2_transmit, False

        if (
            not raw_data_status or
            raw_data_status[0][0] is None or
            raw_data_status[0][0] < unformatted_now
        ):
            log_message(f"Comparing DB time {raw_data_status[0][0] if raw_data_status else 'None'} with current time {unformatted_now}")
            log_message(f"NO data exists in the Raw Data Table at timeslot [{formatted_now}] for Station: {station_name}")
            return data_row, logs_aws_wis2_transmit, False
    else:
        log_message(f"NO data exists in the Database at timeslot [{formatted_now}] for Station: {station_name}")
        return data_row, logs_aws_wis2_transmit, False
    

    ##########################################################################################
    # Fetch station information from the database
    station_query = """
        SELECT wigos, latitude, longitude, elevation, is_automatic
        FROM wx_station 
        WHERE id = %s
    """

    log_message("-------------------------STATION METADATA QUERY-------------------------")

    station_data = execute_query(station_query, [id], "Error fetching station info")

    if station_data:
        row = station_data[0]
        wsi_parts = row[0].split("-") if row[0] else [None] * 4
        data_row.update({
            'wsi_series': wsi_parts[0], 
            'wsi_issuer': wsi_parts[1],
            'wsi_issue_number': wsi_parts[2], 
            'wsi_local': wsi_parts[3],
            'latitude': round(row[1], 5), 
            'longitude': round(row[2], 5),
            'station_height_above_msl': row[3], 
            'barometer_height_above_msl': round(row[3] + 1.7, 2), # 1.7 metre above station_height_above_msl
            # the settings below will be hard coded until further notice
            'thermometer_height': 1.7, 
            'anemometer_height': 10, 
            'time_period_of_wind': -5,
            'rain_sensor_height': 1.5, 
            'station_type': 0,
            'soil_temperature_depth_1' : 0.30,  # 1 feet in meters
            'soil_temperature_depth_2' : 1.22 # 4 feet in meters
        })

    # write info out to logs
    station_metadata_keys = ['wsi_series', 'wsi_issuer', 'wsi_issue_number', 'wsi_local', 
                             'latitude', 'longitude', 'station_height_above_msl', 'barometer_height_above_msl', 
                             'thermometer_height', 'anemometer_height', 'time_period_of_wind', 'rain_sensor_height', 
                             'station_type']

    for key in station_metadata_keys:
        log_message(f"Data for '{key}' value [{data_row[key]}].")

    ##########################################################################################
    # Fetch the latest raw data measurements for specific variables
    log_message("-------------------------RAW DATA QUERY-------------------------")

    raw_data_query = """
        SELECT variable_id, measured, datetime FROM raw_data
        WHERE station_id = %s AND datetime = %s
        AND variable_id IN (10, 19, 21, 23, 30, 51, 56, 60, 61, 72);
    """
    raw_data = execute_query(raw_data_query, [id, formatted_now], "Error fetching raw data")

    for row in raw_data:
        variable_id, measured, db_datetime = row

        key, value = {
            10: ('air_temperature', round(measured + 273.15, 2)),
            19: ('dewpoint_temperature', round(measured + 273.15, 2)),
            30: ('relative_humidity', round(measured)),
            51: ('wind_speed', round(measured, 1)),
            56: ('wind_direction', int(measured)),
            60: ('station_pressure', measured * 100),
            61: ('msl_pressure', measured * 100),
            72: ('solar_radiation', measured),
            21: ('soil_temperature_1', measured),
            23: ('soil_temperature_2', measured)
            # 00: ('soil_moisture_1', measured),
            # 00: ('soil_moisture_2', measured)
        }.get(variable_id, (None, None))

        if key:  # Only update if a valid key was found
            data_row[key] = value
            log_message(f"Data for '{key}' value [{data_row[key]}].") # wrtie info out to logs
            # log_message(f"Data for '{key}' value [{data_row[key]}]. Updated from datetime: {db_datetime}")

    # Populate year, month, day, hour, minute in data_row
    if unformatted_now:
        # Add the extracted date components to data_row
        data_row['year'] = unformatted_now.year
        data_row['month'] = unformatted_now.month
        data_row['day'] = unformatted_now.day
        data_row['hour'] = unformatted_now.hour
        data_row['minute'] = unformatted_now.minute


    ##########################################################################################
    # Fetch precipitation data for the past 24 hours
    precipitation_query = """
        SELECT sum_value 
        FROM hourly_summary 
        WHERE station_id = %s 
        AND variable_id = 0 
        AND datetime BETWEEN %s AND %s
        ORDER BY datetime DESC;
    """
    
    log_message("-------------------------PRECIPITATION QUERY-------------------------")
    precipitation_data = execute_query(precipitation_query, [id, formatted_past_24_hrs, formatted_now], "Error fetching precipitation data")

    precip_values = [row[0] for row in precipitation_data if row]
    if precip_values:

        # 3 hours
        if len(precip_values) >= 3:
            data_row.update({
                'total_precipitation_3_hours': round(sum(precip_values[:3]), 1)
            })
        else:
            data_row.update({
                'total_precipitation_3_hours': None
            })

        # 6 hours
        if len(precip_values) >= 6:
            data_row.update({
                'total_precipitation_6_hours': round(sum(precip_values[:6]), 1)
            })
        else:
            data_row.update({
                'total_precipitation_6_hours': None
            })

        # 12 hours
        if len(precip_values) >= 12:
            data_row.update({
                'total_precipitation_12_hours': round(sum(precip_values[:12]), 1)
            })
        else:
            data_row.update({
                'total_precipitation_12_hours': None
            })

        data_row.update({
            'total_precipitation_1_hour': round(precip_values[0], 1) if len(precip_values) >= 1 else None,
            'total_precipitation_24_hours': round(sum(precip_values), 1) if len(precip_values) >= 24 else None
        })

    # write info out to logs
    station_metadata_keys = ['total_precipitation_1_hour', 'total_precipitation_3_hours', 
                             'total_precipitation_6_hours', 'total_precipitation_12_hours', 
                             'total_precipitation_24_hours'
                            ]

    for key in station_metadata_keys:
        log_message(f"Data for '{key}' value [{data_row[key]}].")


    ##########################################################################################
    def get_max_wind_data(time_label, variable_id_speed, variable_id_direction, start_time, end_time):
        # Get max wind speed
        wind_query = f"""
            SELECT datetime, measured
            FROM raw_data
            WHERE station_id = %s
            AND variable_id = %s
            AND datetime BETWEEN %s AND %s
            ORDER BY datetime DESC;
        """

        log_message(f"-------{time_label} DATA QUERY-------")
        wind_data = execute_query(wind_query, [id, variable_id_speed, start_time, end_time], f"Error fetching {time_label} wind speed data")

        if not wind_data:
            return

        # Get max entry by wind speed
        max_entry = max(wind_data, key=lambda x: x[1])

        if not max_entry:
            return

        speed_value = round(max_entry[1], 1)
        data_row[f'maximum_wind_gust_speed_{time_label}'] = speed_value

        log_message(f"Data for 'maximum_wind_gust_speed_{time_label}' value [{speed_value}].") # write info out to logs

        # Get direction using datetime of max wind
        max_datetime_str = max_entry[0].strftime('%Y-%m-%d %H:%M:%S+00')

        direction_query = """
            SELECT measured
            FROM raw_data
            WHERE station_id = %s
            AND variable_id = %s
            AND datetime = %s
            LIMIT 1;
        """
        direction_data = execute_query(direction_query, [id, variable_id_direction, max_datetime_str], f"Error fetching {time_label} wind direction data")

        if direction_data:
            data_row[f'maximum_wind_gust_direction_{time_label}'] = int(direction_data[0][0])
            log_message(f"Data for 'maximum_wind_gust_direction_{time_label}' value [{int(direction_data[0][0])}].") # write info out to logs

    # --- Constants ---
    VARIABLE_ID_WIND_SPEED = 51
    VARIABLE_ID_WIND_DIRECTION = 56
    VARIABLE_ID_WIND_SPEED_MAX = 53
    VARIABLE_ID_WIND_DIRECTION_DOM = 4063

    log_message("-------------------------WIND DATA QUERY-------------------------")
    # --- 3 HOURS ---
    get_max_wind_data("3_hours", VARIABLE_ID_WIND_SPEED, VARIABLE_ID_WIND_DIRECTION, formatted_past_3_hrs, formatted_now)

    # --- 1 HOUR ---
    get_max_wind_data("1_hour", VARIABLE_ID_WIND_SPEED, VARIABLE_ID_WIND_DIRECTION, formatted_past_1_hr, formatted_now)

    # --- 10 MINUTES (Last 2 readings) ---
    get_max_wind_data("10_minutes", VARIABLE_ID_WIND_SPEED_MAX, VARIABLE_ID_WIND_DIRECTION_DOM, formatted_past_10_min, formatted_now)

    log_message("-----------------------------------------------------------------")
    logging.info(f'AWS Transmission: Processing data for Station Complete: {station_name}')
    log_message(f'AWS Transmission: Processing data for Station Complete: {station_name}')

    return data_row, logs_aws_wis2_transmit, True



# from wx.tasks import aquacrop_gen_daily_data
@shared_task
def aquacrop_gen_daily_data():
    try:
        yesterday = datetime.now().date() - timedelta(days=1)

        man_station = Station.objects.get(code='TEST_STATION_MAN_1')
        aut_station = Station.objects.get(code='TEST_STATION_AUT_1')

        sql_query_check = f"""
            SELECT
                station_id
                ,COUNT(*)
            FROM daily_summary ds
            WHERE station_id IN ({man_station.id}, {aut_station.id})
            AND DAY = '{yesterday}'
            GROUP BY station_id;
        """
        config = settings.SURFACE_CONNECTION_STRING
        with psycopg2.connect(config) as conn:
            df = pd.read_sql(sql_query_check, conn)

        if df.empty:
            aut_exists = False
            man_exists = False
        else:
            man_exists = df[df['station_id']==man_station.id].shape[0] > 0
            aut_exists = df[df['station_id']==aut_station.id].shape[0] > 0

        if not man_exists:
            sql_query_man = f"""
                WITH daily_data AS (
                    SELECT
                        day
                        ,station_id
                        ,variable_id
                        ,min_value
                        ,max_value
                        ,avg_value
                        ,sum_value
                        ,num_records
                    FROM daily_summary ds
                    JOIN wx_variable vr ON ds.variable_id = vr.id
                    WHERE vr.symbol IN ('TEMPMIN', 'PRECIP', 'TEMPMAX') 
                    AND ds.station_id=14
                    AND day = '2008-{yesterday.day}-{yesterday.month}'
                )
                INSERT INTO daily_summary(
                    created_at
                    ,updated_at
                    ,day
                    ,station_id
                    ,variable_id
                    ,min_value
                    ,max_value
                    ,avg_value
                    ,sum_value
                    ,num_records
                )
                SELECT
                    NOW()
                    ,NOW()
                    ,'{yesterday}' AS day
                    ,{man_station.id} AS station_id
                    ,variable_id
                    ,min_value
                    ,max_value
                    ,avg_value
                    ,sum_value
                    ,num_records
                FROM daily_data
            """

            config = settings.SURFACE_CONNECTION_STRING
            with psycopg2.connect(config) as conn:
                with conn.cursor() as cursor:
                    logging.info(sql_query_man)
                    cursor.execute(sql_query_man)

        if not aut_exists:
            sql_query_aut = f"""
                WITH daily_data AS (
                    SELECT
                        day
                        ,station_id
                        ,variable_id
                        ,min_value
                        ,max_value
                        ,avg_value
                        ,sum_value
                        ,num_records
                    FROM daily_summary ds
                    JOIN wx_variable vr ON ds.variable_id = vr.id
                    WHERE vr.symbol IN ('TEMP', 'PRECIP', 'PRESSTN', 'WNDSPAVG', 'SOLARRAD', 'RH')
                    AND ds.station_id=123
                    AND day = '2024-{yesterday.day}-{yesterday.month}'
                )
                INSERT INTO daily_summary(
                    created_at
                    ,updated_at
                    ,day
                    ,station_id
                    ,variable_id
                    ,min_value
                    ,max_value
                    ,avg_value
                    ,sum_value
                    ,num_records
                )
                SELECT
                    NOW()
                    ,NOW()
                    ,'{yesterday}' as day
                    ,{aut_station.id} as station_id
                    ,variable_id
                    ,min_value
                    ,max_value
                    ,avg_value
                    ,sum_value
                    ,num_records
                FROM daily_data
            """

            config = settings.SURFACE_CONNECTION_STRING
            with psycopg2.connect(config) as conn:
                with conn.cursor() as cursor:
                    logging.info(sql_query_aut)
                    cursor.execute(sql_query_aut)
    except Exception as e:
        logger.error(f'Error on aquacrop_gen_daily_data: {repr(e)}')


# fetch interpolation data
@shared_task
def fetch_interpolation_data_task(
    *,
    start_datetime,
    end_datetime,
    variable_id,
    agg="instant",
    source="raw_data",
    quality_flags=None
):
    """
    Fetches and aggregates climate interpolation data.
    Returns a Python dict (Celery-safe).
    """

    where_query = ""

    # RAW DATA SOURCE LOGIC
    if source == "raw_data":
        dt_query = "datetime"
        value_query = "measured"
        source_query = "raw_data"

        if quality_flags:
            try:
                [int(qf) for qf in quality_flags.split(',')]
            except ValueError:
                raise ValueError("Invalid quality_flags value.")

            where_query = (
                f" measured != {settings.MISSING_VALUE} "
                f"AND quality_flag IN ({quality_flags}) AND "
            )
        else:
            where_query = f" measured != {settings.MISSING_VALUE} AND "

    # SUMMARY DATA SOURCES
    else:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sampling_operation_id
                FROM wx_variable
                WHERE id=%(variable_id)s
            """, params={'variable_id': variable_id})
            sampling_operation = cursor.fetchone()[0]

        if sampling_operation in [6, 7]:
            value_query = "sum_value"
        elif sampling_operation == 3:
            value_query = "min_value"
        elif sampling_operation == 4:
            value_query = "max_value"
        else:
            value_query = "avg_value"

        if source == "hourly":
            dt_query = "datetime"
            source_query = "hourly_summary"
        elif source == "daily":
            dt_query = "day"
            source_query = "daily_summary"
        elif source == "monthly":
            dt_query = "date"
            source_query = "monthly_summary"
        elif source == "yearly":
            dt_query = "date"
            source_query = "yearly_summary"

    # TIME FILTERING
    if agg == "instant":
        where_query += (
            "variable_id=%(variable_id)s AND "
            + dt_query + "=%(datetime)s"
        )
        params = {
            'datetime': start_datetime,
            'variable_id': variable_id
        }
    else:
        where_query += (
            "variable_id=%(variable_id)s AND "
            + dt_query + " >= %(start_datetime)s AND "
            + dt_query + " <= %(end_datetime)s"
        )
        params = {
            'start_datetime': start_datetime,
            'end_datetime': end_datetime,
            'variable_id': variable_id
        }

    # QUERY EXECUTION
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                a.station_id,
                b.name,
                b.code,
                b.latitude,
                b.longitude,
                a.""" + value_query + """ AS measured
            FROM """ + source_query + """ a
            INNER JOIN wx_station b ON a.station_id=b.id
            WHERE """ + where_query,
            params=params
        )
        raw_data = cursor.fetchall()

    data = []

    for row in raw_data:
        data.append({
            'station_id': row[0],
            'name': row[1],
            'code': row[2],
            'latitude': row[3],
            'longitude': row[4],
            'measured': row[5],
        })

    # AGGREGATION
    if agg != "instant" and raw_data:
        df = pd.json_normalize(data)
        data = json.loads(
            df.groupby(
                ['station_id', 'name', 'code', 'longitude', 'latitude']
            )
            .agg(agg)
            .reset_index()
            .sort_values('name')
            .to_json(orient="records")
        )

    return {"data": data}



# fetch interpolation image
@shared_task
def fetch_interpolation_img_task(
    start_datetime,
    end_datetime,
    variable_id,
    cmap,
    hres,
    minimum_neighbors,
    search_radius,
    agg,
    source,
    vmin,
    vmax,
    quality_flags,
    stands_llat,
    stands_llon,
    stands_ulat,
    stands_ulon
):

    # dir where spatial analysis image overlay can be found. Create it if it doesn't exits else just move on
    out_dir = Path("/surface/static/images/spatial_img") 
    out_dir.mkdir(parents=True, exist_ok=True)
    
    # dir where spatial analysis image overlay can be found. Create it if it doesn't exits else just move on
    out_dir = Path("/surface/static/images/spatial_img") 
    out_dir.mkdir(parents=True, exist_ok=True)
    
    stations_df = pd.read_sql_query("""
        SELECT id,name,alias_name,code,latitude,longitude
        FROM wx_station
        WHERE longitude!=0
        """,
                                    con=connection
                                    )
    
    stations = geopandas.GeoDataFrame(
        stations_df, 
        geometry=geopandas.points_from_xy(stations_df.longitude, stations_df.latitude)
    )
    
    stations.crs = 'epsg:4326'

    where_query = ""
    if source == "raw_data":
        dt_query = "datetime"
        value_query = "measured"
        source_query = "raw_data"
        if quality_flags:
            try:
                [int(qf) for qf in quality_flags.split(',')]
            except ValueError:
                raise ValueError("Invalid quality_flags value.")

            where_query = f" measured != {settings.MISSING_VALUE} AND quality_flag IN ({quality_flags}) AND "
        else:
            where_query = f" measured != {settings.MISSING_VALUE} AND "
    else:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sampling_operation_id
                FROM wx_variable
                WHERE id=%(variable_id)s
                """,
                           params={'variable_id': variable_id}
                           )
            sampling_operation = cursor.fetchone()[0]

        if sampling_operation in [6, 7]:
            value_query = "sum_value"
        elif sampling_operation == 3:
            value_query = "min_value"
        elif sampling_operation == 4:
            value_query = "max_value"
        else:
            value_query = "avg_value"

        if source == "hourly":
            dt_query = "datetime"
            source_query = "hourly_summary"

        elif source == "daily":
            dt_query = "day"
            source_query = "daily_summary"

        elif source == "monthly":
            dt_query = "date"
            source_query = "monthly_summary"

        elif source == "yearly":
            dt_query = "date"
            source_query = "yearly_summary"

    if agg == "instant":
        where_query += "variable_id=%(variable_id)s AND " + dt_query + "=%(datetime)s"
        params = {'datetime': start_datetime, 'variable_id': variable_id}
    else:
        where_query += "variable_id=%(variable_id)s AND " + dt_query + " >= %(start_datetime)s AND " + dt_query + " <= %(end_datetime)s"
        params = {'start_datetime': start_datetime, 'end_datetime': end_datetime, 'variable_id': variable_id}

    climate_data = pd.read_sql_query(
        "SELECT station_id,variable_id," + dt_query + "," + value_query + """
        FROM """ + source_query + """
        WHERE """ + where_query + "",
        params=params,
        con=connection
    )

    if len(climate_data) == 0:
        return {"data":"/static/images/no-interpolated-data.png"}

    df_merged = pd.merge(left=climate_data, right=stations, how='left', left_on='station_id', right_on='id')
    df_climate = df_merged[["station_id", dt_query, "longitude", "latitude", value_query]]

    if agg != "instant":
        df_climate = (
            df_climate
            .groupby(['station_id', 'longitude', 'latitude'], as_index=False)
            .agg({value_query: agg})
        )

    gx, gy, img = interpolate_to_grid(
        df_climate["longitude"],
        df_climate["latitude"],
        df_climate[value_query],
        interp_type='cressman',
        minimum_neighbors=int(minimum_neighbors),
        hres=float(hres),
        search_radius=float(search_radius),
        boundary_coords={'west': stands_llon, 'east': stands_ulon, 'south': stands_llat, 'north': stands_ulat}
    )

    fig = plt.figure(frameon=False)
    ax = plt.Axes(fig, [0., 0., 1., 1.])
    ax.set_axis_off()
    fig.add_axes(ax)
    ax.imshow(img, origin='lower', cmap=cmap, vmin=vmin, vmax=vmax)
    fname = str(uuid.uuid4())
    fig.savefig("/surface/static/images/spatial_img/" + fname + ".png", dpi='figure', format='png', transparent=True,
                bbox_inches=Bbox.from_bounds(2, 0, 2.333, 4.013))

    image1 = cv2.imread("/surface/static/images/spatial_img/" + fname + ".png", cv2.IMREAD_UNCHANGED)
    image2 = cv2.imread(settings.SPATIAL_ANALYSIS_SHAPE_FILE_PATH, cv2.IMREAD_UNCHANGED)
    image1 = cv2.resize(image1, dsize=(image2.shape[1], image2.shape[0]))
    for i in range(image1.shape[0]):
        for j in range(image1.shape[1]):
            image1[i][j][3] = image2[i][j][3]
    cv2.imwrite("/surface/static/images/spatial_img/" + fname + "-output.png", image1)

    return {"data":f"/static/images/spatial_img/{fname}-output.png"}


# fetch interpolation image after custom value has been added
@shared_task
def fetch_mod_interpolation_img_task(
    stands_llat,
    stands_llon,
    stands_ulat,
    stands_ulon,
    parameters,
    vmin,
    vmax,
    json_body_data
):

    df_climate = pd.json_normalize(json_body_data)

    # dir where spatial analysis image overlay can be found. Create it if it doesn't exits else just move on
    out_dir = Path("/surface/static/images/spatial_img") 
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        df_climate = df_climate[["station_id", "longitude", "latitude", "measured"]]
    except KeyError:
        return {"data":"/static/images/no-interpolated-data.png"}

    gx, gy, img = interpolate_to_grid(
        df_climate["longitude"],
        df_climate["latitude"],
        df_climate["measured"],
        interp_type='cressman',
        minimum_neighbors=int(parameters["minimum_neighbors"]),
        hres=float(parameters["hres"]),
        search_radius=float(parameters["search_radius"]),
        boundary_coords={'west': stands_llon, 'east': stands_ulon, 'south': stands_llat, 'north': stands_ulat}
    )

    fig = plt.figure(frameon=False)
    ax = plt.Axes(fig, [0., 0., 1., 1.])
    ax.set_axis_off()
    fig.add_axes(ax)
    ax.imshow(img, origin='lower', cmap=parameters["cmap"]["value"], vmin=vmin, vmax=vmax)
    
    fname = datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ_") + str(uuid.uuid4())

    fig.savefig("/surface/static/images/spatial_img/" + fname + ".png", dpi='figure', format='png', transparent=True,
                bbox_inches=Bbox.from_bounds(2, 0, 2.333, 4.013))
    
    image1 = cv2.imread("/surface/static/images/spatial_img/" + fname + ".png", cv2.IMREAD_UNCHANGED)
    image2 = cv2.imread(settings.SPATIAL_ANALYSIS_SHAPE_FILE_PATH, cv2.IMREAD_UNCHANGED)
    image1 = cv2.resize(image1, dsize=(image2.shape[1], image2.shape[0]))

    for i in range(image1.shape[0]):
        for j in range(image1.shape[1]):
            image1[i][j][3] = image2[i][j][3]

    cv2.imwrite("/surface/static/images/spatial_img/" + fname + "-output.png", image1)

    return {"data":f"/static/images/spatial_img/{fname}-output.png"}


# Remove stale interpolation image clean up tasks
@shared_task
def interpolation_img_cleanup():
    try:
        for filename in os.listdir('/surface/static/images/spatial_img/'):
            os.remove('/surface/static/images/spatial_img/' + filename)
    except Exception as e:
        # Log an error message if the clean up task fails
        logger.error("Interpolation image clean up task failed!")
        logger.exception(f"Error: {e}")  # Log the full exception traceback for debugging



ALLOWED_UPDATE_FIELDS = {"manual_flag", "consisted", "remarks"}

def _build_set_clause(updates: dict):
    """
    Returns (set_clause_sql, set_params)
    Example: ("manual_flag=%s, remarks=%s", [4, "note"])
    """
    parts = []
    params = []

    for key in ("manual_flag", "consisted", "remarks"):
        if key not in updates:
            continue
        if key not in ALLOWED_UPDATE_FIELDS:
            continue

        parts.append(sql.SQL("{}=%s").format(sql.Identifier(key)))
        params.append(updates[key])

    if not parts:
        raise ValueError("No valid fields to update.")

    return sql.SQL(", ").join(parts), params


# Update the QC manual checks
@shared_task(bind=True)
def qc_manual_checks(self, station_id: int, variable_id: int, req_datetime: str, updates: dict):
    """
    req_datetime: ISO string like '2026-02-05T12:00:00Z'
    updates: {"manual_flag": 4, "consisted": 1.23, "remarks": "ok"}
    """
    # Parse datetime (UTC)
    dt = datetime.strptime(req_datetime, "%Y-%m-%dT%H:%M:%SZ").replace(
        tzinfo=timezone.utc
    )
    now_time = datetime.now(timezone.utc)

    set_sql, set_params = _build_set_clause(updates)

    with transaction.atomic():
        with connection.cursor() as cursor:
            # 1) Find chunk containing dt
            cursor.execute(
                """
                SELECT chunk_schema, chunk_name, is_compressed
                FROM timescaledb_information.chunks
                WHERE hypertable_name = 'raw_data'
                  AND %s::timestamptz >= range_start
                  AND %s::timestamptz <  range_end
                LIMIT 1
                """,
                (dt, dt),
            )
            row = cursor.fetchone()
            if not row:
                raise ValueError("No chunk found for datetime (unexpected).")

            chunk_schema, chunk_name, is_compressed = row
            if is_compressed:
                # Guardrail for the future: compressed chunks are not safe to update directly
                raise ValueError("Target chunk is compressed; cannot apply manual QC update.")

            # 2) Update chunk directly
            update_stmt = sql.SQL("""
                UPDATE {chunk}
                SET {set_clause}
                WHERE station_id = %s
                  AND variable_id = %s
                  AND datetime = %s::timestamptz
            """).format(
                chunk=sql.Identifier(chunk_schema, chunk_name),
                set_clause=set_sql,
            )

            params = list(set_params) + [station_id, variable_id, dt]
            cursor.execute(update_stmt, params)

            if cursor.rowcount == 0:
                raise ValueError("No matching record found to update.")

            # 3) Schedule summaries
            cursor.execute("""
                INSERT INTO wx_hourlysummarytask (station_id, datetime, updated_at, created_at)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (station_id, dt, now_time, now_time))

            cursor.execute("""
                INSERT INTO wx_dailysummarytask (station_id, date, updated_at, created_at)
                SELECT
                    %s,
                    ( (%s AT TIME ZONE 'UTC') + make_interval(mins => COALESCE(s.utc_offset_minutes, 0)) )::date,
                    %s, %s
                FROM wx_station s
                WHERE s.id = %s
                ON CONFLICT DO NOTHING
            """, (station_id, dt, now_time, now_time, station_id))

    return {"updated": True}


@shared_task(bind=True)
def qc_manual_checks_bulk(
    self,
    station_id: int,
    variable_id: int,
    manual_flag: int,
    remarks: str,
    datetime_list: list,
):
    """
    Bulk QC update:
    - Updates raw_data rows for a station+variable at many datetimes
    - Uses Timescale chunk lookup to update chunk tables directly (fast)
    - Groups datetimes by chunk to reduce chunk lookups and UPDATE statements
    - Dedupe hourly/daily summary task inserts (insert once per unique hour/day)

    Inputs:
      datetime_list: list of ISO-Z strings like '2026-02-05T12:00:00Z'
    """

    # -----------------------------
    # 1) Normalize / parse datetimes
    # -----------------------------
    if not datetime_list:
        raise ValueError("datetime_list is empty.")

    # Parse to aware UTC datetimes
    dts = []
    for dt_str in datetime_list:
        try:
            dt = datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%SZ").replace(
                tzinfo=timezone.utc
            )
            dts.append(dt)
        except Exception:
            raise ValueError(f"Invalid datetime format: {dt_str}")

    # remove duplicates early
    dts = sorted(set(dts))

    # Compute min/max for chunk range query
    min_dt = dts[0]
    max_dt = dts[-1]

    now_time = datetime.now(timezone.utc)

    # -----------------------------
    # 2) Run everything atomically
    # -----------------------------
    with transaction.atomic():
        with connection.cursor() as cursor:

            # ------------------------------------------------------------
            # 2a) Fetch station offset ONCE (needed for daily summary date)
            # ------------------------------------------------------------
            cursor.execute(
                "SELECT utc_offset_minutes FROM wx_station WHERE id = %s",
                (station_id,),
            )
            row = cursor.fetchone()
            utc_offset_minutes = row[0] if row and row[0] is not None else 0

            # ------------------------------------------------------------
            # 2b) Fetch ONLY chunks that overlap our [min_dt, max_dt] range
            #     This avoids scanning 5867 chunks in Python.
            # ------------------------------------------------------------
            cursor.execute(
                """
                SELECT chunk_schema, chunk_name, range_start, range_end, is_compressed
                FROM timescaledb_information.chunks
                WHERE hypertable_name = 'raw_data'
                  -- Overlap check for half-open ranges:
                  -- chunk overlaps [min_dt, max_dt] if:
                  --   range_end > min_dt AND range_start <= max_dt
                  AND range_end  > %s::timestamptz
                  AND range_start <= %s::timestamptz
                ORDER BY range_start
                """,
                (min_dt, max_dt),
            )
            chunk_rows = cursor.fetchall()
            if not chunk_rows:
                # This would be unusual unless the datetimes are outside the hypertable data range
                raise ValueError("No chunks found for the requested datetime range.")

            # Guardrail: if any relevant chunk is compressed, bulk update should stop.
            # for (_, _, _, _, is_compressed) in chunk_rows:
            #     if is_compressed:
            #         raise ValueError("One or more target chunks are compressed; cannot apply bulk update.")

            # ------------------------------------------------------------
            # 2c) Group each datetime into its owning chunk
            #     Uses the standard half-open test:
            #       range_start <= dt < range_end
            # ------------------------------------------------------------
            chunk_map = {}  # (schema, name) -> list[datetime]

            # Convert chunk ranges into a list of dicts for easier looping
            chunks = [
                {
                    "schema": r[0],
                    "name": r[1],
                    "start": r[2],
                    "end": r[3],
                }
                for r in chunk_rows
            ]

            # Assign each dt to exactly one chunk
            # (Given standard chunking, each dt should match exactly one chunk.)
            for dt in dts:
                matched = False
                for ch in chunks:
                    if ch["start"] <= dt < ch["end"]:
                        key = (ch["schema"], ch["name"])
                        chunk_map.setdefault(key, []).append(dt)
                        matched = True
                        break
                if not matched:
                    # If this happens, it usually indicates a boundary/format issue.
                    raise ValueError(f"No chunk matched datetime {dt.isoformat()}")

            # ------------------------------------------------------------
            # 2d) Perform one UPDATE per chunk using datetime = ANY(array)
            # ------------------------------------------------------------
            # We'll track total updated rows for debugging/monitoring.
            total_updated = 0

            for (chunk_schema, chunk_name), chunk_dts in chunk_map.items():
                # Build safe SQL for the chunk table name
                update_stmt = sql.SQL("""
                    UPDATE {chunk}
                    SET manual_flag = %s,
                        remarks = %s,
                        consisted = NULL
                    WHERE station_id = %s
                      AND variable_id = %s
                      AND datetime = ANY(%s::timestamptz[])
                """).format(
                    chunk=sql.Identifier(chunk_schema, chunk_name)
                )

                # IMPORTANT:
                # - psycopg2 can adapt Python lists into PostgreSQL arrays.
                # - We cast explicitly to timestamptz[] so Postgres knows the type.
                cursor.execute(
                    update_stmt,
                    (manual_flag, remarks, station_id, variable_id, chunk_dts),
                )

                # rowcount is how many rows were updated in this statement
                total_updated += cursor.rowcount

            if total_updated == 0:
                # Nothing matched. Could be wrong station/variable, or datetimes not present.
                raise ValueError("No matching records found to update in bulk.")

            # ------------------------------------------------------------
            # 2e) Dedupe hourly + daily summary task scheduling
            # ------------------------------------------------------------
            # Hourly: dedupe by hour bucket (UTC hour)
            # Daily: dedupe by station-local date (offset minutes)
            hourly_set = set()
            daily_set = set()

            for dt in dts:
                # Convert UTC dt -> station-local datetime
                local_dt = dt + timedelta(minutes=utc_offset_minutes)

                # DAILY bucket: if dt is exactly at 00:00:00, treat it as belonging to the previous day
                # (matches daily summary logic that effectively assigns midnight boundary values to the prior day)
                if (local_dt.hour == 0 and local_dt.minute == 0 and local_dt.second == 0 and local_dt.microsecond == 0):
                    local_dt = local_dt - timedelta(seconds=1)

                daily_set.add(local_dt.date())

                # Hour bucket stays UTC (unchanged)
                hour_dt = dt.replace(minute=0, second=0, microsecond=0)
                hourly_set.add(hour_dt)

            # Insert hourly summary tasks once per unique hour
            cursor.executemany(
                """
                INSERT INTO wx_hourlysummarytask (station_id, datetime, updated_at, created_at)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                """,
                [(station_id, hour_dt, now_time, now_time) for hour_dt in sorted(hourly_set)],
            )

            # Insert daily summary tasks once per unique day
            cursor.executemany(
                """
                INSERT INTO wx_dailysummarytask (station_id, date, updated_at, created_at)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                """,
                [(station_id, day, now_time, now_time) for day in sorted(daily_set)],
            )

    # Return useful info for debugging/polling responses
    return {
        "updated_rows": total_updated,
        "unique_hours": len(hourly_set),
        "unique_days": len(daily_set),
        "chunks_touched": len(chunk_map),
    }


# Get data for the QC manual checks
@shared_task
def get_qc_data(sql_string, where_parameters, response):
    with connection.cursor() as cursor:

        cursor.execute(sql_string, where_parameters)
        rows = cursor.fetchall()

        for row in rows:
            obj = {
                'datetime': row[0],
                'measured': row[1],
                'consisted': row[2],
                'automatic_flag': row[3],
                'manual_flag': row[4],
                'station_id': row[5],
                'variable_id': row[6],
                'remarks': row[7],
                'qc_range_quality_flag': row[8],
                'qc_step_quality_flag': row[9],
                'qc_persist_quality_flag': row[10]
            }

            response['results'].append(obj)

    return response


# IMPORTANT:
# Monthly capture replaces all existing raw_data rows for the same
# station + variable(s) + displayed local day before inserting the new
# daily/manual value.
#
# This intentionally does NOT filter by is_daily because this form is
# used for manual daily stations, and any higher-frequency rows present
# for these stations are treated as replaceable/test/noise data.
def delete_existing_monthly_daily_rows(records_list, utc_offset_minutes):
    """
    Delete existing monthly/daily raw_data rows before inserting updated values.

    This matches rows by:
        station_id
        variable_id
        station-local day

    Performance notes:
        - Uses datetime range filtering instead of:
              (datetime + offset)::date = local_date

          This is faster because PostgreSQL/Timescale can use indexes
          and chunk pruning on the datetime column.

        - Does NOT use RETURNING because we only need the number of rows
          deleted. cursor.rowcount is cheaper than fetching deleted rows.
    """

    if not records_list:
        return

    delete_groups = {}

    for record in records_list:
        station_id = int(record[0])
        variable_id = int(record[1])
        data_datetime = record[3]

        # record[3] should already be the station-local datetime
        # for the monthly form row.
        #
        # Example:
        #   2026-05-02 00:00:00-06:00
        day_start = data_datetime.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
        )

        # Half-open interval:
        #   day_start <= datetime < day_end
        #
        # This safely captures all records for that local day.
        day_end = day_start + timedelta(days=1)

        key = (station_id, day_start, day_end)

        if key not in delete_groups:
            delete_groups[key] = set()

        delete_groups[key].add(variable_id)

    delete_query = """
        DELETE FROM raw_data
        WHERE station_id = %s
          AND variable_id = ANY(%s)
          AND datetime >= %s
          AND datetime < %s
    """

    with psycopg2.connect(settings.SURFACE_CONNECTION_STRING) as conn:
        with conn.cursor() as cursor:
            for (station_id, day_start, day_end), variable_ids in delete_groups.items():
                variable_ids = list(variable_ids)

                cursor.execute(
                    delete_query,
                    [
                        station_id,
                        variable_ids,
                        day_start,
                        day_end,
                    ]
                )

                deleted_count = cursor.rowcount

                logger.info(
                    "Monthly pre-insert cleanup deleted %s rows | station_id=%s | day_start=%s | day_end=%s | variable_ids=%s",
                    deleted_count,
                    station_id,
                    day_start,
                    day_end,
                    variable_ids,
                )

        conn.commit()


# Process saving of the monthly form
@shared_task(bind=True)
def save_monthly_form(self, records_list, first_day, station_id, override_data_on_conflict, utc_offset_minutes):
    try:
        first_day = pandas.to_datetime(first_day)

        fixed_records = []

        for record in records_list:
            record = list(record)

            # index 3 = data_datetime
            dt = record[3]

            if isinstance(dt, str):
                record[3] = pandas.to_datetime(dt)

            fixed_records.append(tuple(record))

        # If there are no records, do not call insert_rd_daily.
        # This can happen if the user only cleared cells or only changed remarks
        # on a row that has no measurement values to save.
        if not fixed_records:
            logger.info(
                "Monthly form save skipped insert because records_list is empty | first_day=%s | station_id=%s",
                first_day,
                station_id,
            )
            return {
                "inserted": 0,
                "message": "No monthly records to insert."
            }

        # Delete existing values for the same displayed local day first
        delete_existing_monthly_daily_rows(
            fixed_records,
            utc_offset_minutes
        )

        insert_rd_daily(
            raw_data_list=fixed_records,
            date=first_day,
            station_id=station_id,
            override_data_on_conflict=True,
            utc_offset_minutes=utc_offset_minutes
        )

        return {
            "inserted": len(fixed_records),
            "message": "Monthly records saved successfully."
        }

    except Exception as e:
        logger.error(f"Monthly form save task failed: {repr(e)}")
        raise e


# Process saving of the synop form
@shared_task(bind=True)
def save_synop_form(self, records_list, day, station_id, station_utc_offset):
    try:
        # Convert datetime strings back to datetime objects
        fixed_records = []

        for record in records_list:
            record = list(record)

            # index 3 = data_datetime
            dt = record[3]

            if isinstance(dt, str):
                record[3] = dateutil.parser.isoparse(dt)

            fixed_records.append(tuple(record))

        insert_rd_synop(
            raw_data_list=fixed_records,
            date=day,
            station_id=station_id,
            override_data_on_conflict=True,
            utc_offset_minutes=station_utc_offset
        )

    except Exception as e:
        logger.error(f"Task failed: {repr(e)}")
        raise e